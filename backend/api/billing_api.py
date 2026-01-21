# backend/api/billing_api.py (添加考勤录入API)

from flask import Blueprint, jsonify, current_app, request, send_from_directory, url_for
from flask_jwt_extended import jwt_required
from sqlalchemy import or_, case, and_, not_, func, distinct, cast, String
from sqlalchemy.orm import with_polymorphic, attributes
from flask_jwt_extended import get_jwt_identity
from dateutil.parser import parse as date_parse
import re
import io # <-- 添加此行
import calendar
from datetime import date, timedelta, datetime
import decimal
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from collections import defaultdict
from pypinyin import pinyin, Style, lazy_pinyin
import os
import uuid
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import send_file

# --- 新增 AttendanceRecord 的导入 ---
from backend.models import (
    db,
    BaseContract,
    User,
    ServicePersonnel,
    MaternityNurseContract,
    NannyTrialContract,
    AttendanceRecord,
    CustomerBill,
    EmployeePayroll,
    FinancialAdjustment,
    FinancialActivityLog,
    InvoiceRecord,
    PaymentStatus,
    PaymentRecord,
    PayoutStatus,
    PayoutRecord,
    AdjustmentType,
    ExternalSubstitutionContract,
    NannyContract, 
    TrialOutcome,
    CompanyBankAccount,
    BankTransaction, 
    BankTransactionStatus,
    PayerAlias,
    SubstituteRecord,
    Customer,
    SigningStatus
)
from backend.tasks import (
    sync_all_contracts_task,
    calculate_monthly_billing_task,
    post_virtual_contract_creation_task,
    generate_all_bills_for_contract_task,
)  # 导入新任务
from backend.services.billing_engine import BillingEngine, _update_bill_payment_status, _update_payroll_payout_status, calculate_substitute_management_fee
from backend.services.contract_service import (
    create_maternity_nurse_contract_adjustments,
    cancel_substitute_bill_due_to_transfer,
    update_salary_history_on_contract_activation
)
from backend.api.utils import get_billing_details_internal, get_contract_type_details, _log_activity
from backend.services.contract_service import _find_successor_contract_internal
from backend.services.payment_message_generator import PaymentMessageGenerator
from backend.services.bank_statement_service import BankStatementService


D = decimal.Decimal

billing_bp = Blueprint("billing_api", __name__, url_prefix="/api/billing")

ADJUSTMENT_TYPE_LABELS = {
    AdjustmentType.CUSTOMER_INCREASE: "客户增款",
    AdjustmentType.CUSTOMER_DECREASE: "退客户款",
    AdjustmentType.CUSTOMER_DISCOUNT: "优惠",
    AdjustmentType.EMPLOYEE_INCREASE: "员工增款",
    AdjustmentType.EMPLOYEE_DECREASE: "员工减款",
    AdjustmentType.DEFERRED_FEE: "上期顺延费用",
    AdjustmentType.EMPLOYEE_COMMISSION: "员工佣金",
    AdjustmentType.EMPLOYEE_COMMISSION_OFFSET: "佣金冲账",
    AdjustmentType.INTRODUCTION_FEE: "介绍费",
    AdjustmentType.DEPOSIT: "定金",
    AdjustmentType.EMPLOYEE_CLIENT_PAYMENT: "客户直付"
}

def _get_or_create_personnel_ref(name: str, phone: str = None):
    """
    根据姓名和（可选）手机号，查找或创建服务人员。
    始终返回一个包含 ServicePersonnel ID 的字典。
    """
    sp = None
    # 1. 优先在 ServicePersonnel 表中查找
    if phone:
        sp = ServicePersonnel.query.filter_by(phone_number=phone).first()
    if not sp and name:
        sp = ServicePersonnel.query.filter_by(name=name).first()

    # 2. 如果在 ServicePersonnel 中找到，直接返回
    if sp:
        return {"type": "service_personnel", "id": sp.id}

    # 3. 如果在 ServicePersonnel 中未找到，尝试在 User 表中查找
    user = None
    if phone:
        user = User.query.filter_by(phone_number=phone).first()
    if not user and name:
        user = User.query.filter_by(username=name).first()

    # 4. 如果在 User 中找到，则创建或关联 ServicePersonnel
    if user:
        # 检查 User 是否已经关联了 ServicePersonnel
        if user.service_personnel_profile:
            return {"type": "service_personnel", "id": user.service_personnel_profile.id}
        else:
            # User 存在但没有关联 ServicePersonnel，创建新的 ServicePersonnel 并关联
            name_pinyin_full = "".join(lazy_pinyin(name))
            name_pinyin_initials = "".join(item[0] for item in pinyin(name, style=Style.FIRST_LETTER))
            name_pinyin_combined = f"{name_pinyin_full}{name_pinyin_initials}"

            new_sp = ServicePersonnel(
                name=name,
                phone_number=phone,
                id_card_number=user.id_card_number, # 尝试同步身份证号
                user_id=user.id,
                name_pinyin=name_pinyin_combined
            )
            db.session.add(new_sp)
            db.session.flush()
            current_app.logger.info(f"为现有用户 {user.username} 创建并关联了新的服务人员: {name} (ID: {new_sp.id})")
            return {"type": "service_personnel", "id": new_sp.id}

    # 5. 如果 User 和 ServicePersonnel 都没有找到，则创建新的 ServicePersonnel
    name_pinyin_full = "".join(lazy_pinyin(name))
    name_pinyin_initials = "".join(item[0] for item in pinyin(name, style=Style.FIRST_LETTER))
    name_pinyin_combined = f"{name_pinyin_full}{name_pinyin_initials}"

    new_sp = ServicePersonnel(
        name=name,
        phone_number=phone,
        name_pinyin=name_pinyin_combined
    )
    db.session.add(new_sp)
    db.session.flush()
    current_app.logger.info(f"创建了新的服务人员: {name} (ID: {new_sp.id}, Pinyin: {name_pinyin_combined})")
    return {"type": "service_personnel", "id": new_sp.id}


def admin_required(fn):
    return jwt_required()(fn)


@billing_bp.route("/sync-contracts", methods=["POST"])
@admin_required
def trigger_sync_contracts():
    try:
        task = sync_all_contracts_task.delay()
        return jsonify(
            {"message": "合同同步任务已成功提交到后台处理。", "task_id": task.id}
        ), 202
    except Exception:
        return jsonify({"error": "提交后台任务时发生错误"}), 500


@billing_bp.route("/calculate-bills", methods=["POST"])
@admin_required
def trigger_calculate_bills():
    # 这个接口用于【批量计算】，应该触发【异步任务】
    data = request.get_json()
    year = data.get("year")
    month = data.get("month")
    if not all([year, month]):
        return jsonify({"error": "缺少 year 和 month 参数"}), 400

    # 调用 Celery 任务，使用 .delay()
    task = calculate_monthly_billing_task.delay(
        year=year, month=month, force_recalculate=False
    )
    return jsonify({"task_id": task.id, "message": "批量计算任务已提交"})


# 新增一个用于强制重算的接口
@billing_bp.route("/force-recalculate", methods=["POST"])
@admin_required
def force_recalculate_bill():
    # 这个接口用于【单个强制重算】，也应该触发【异步任务】，以避免请求超时
    data = request.get_json()
    contract_id = data.get("contract_id")
    year = data.get("year")
    month = data.get("month")
    if not all([contract_id, year, month]):
        return jsonify({"error": "缺少必要参数"}), 400

    # 调用 Celery 任务，并传递所有参数
    task = calculate_monthly_billing_task.delay(
        year=year, month=month, contract_id=contract_id, force_recalculate=True
    )
    return jsonify({"message": "强制重算任务已提交", "task_id": task.id})


@billing_bp.route("/attendance", methods=["POST"])
@admin_required
def save_attendance():
    data = request.get_json()
    # 简化输入，只需要加班天数
    required_fields = [
        "contract_id",
        "cycle_start_date",
        "cycle_end_date",
        "overtime_days",
        "billing_year",
        "billing_month",
    ]
    if not all(field in data for field in required_fields):
        return jsonify({"error": "缺少必要字段"}), 400

    try:
        contract_id = data["contract_id"]
        cycle_start = datetime.strptime(data["cycle_start_date"], "%Y-%m-%d").date()
        cycle_end = datetime.strptime(data["cycle_end_date"], "%Y-%m-%d").date()
        overtime_days = int(data["overtime_days"])
        billing_year = int(data["billing_year"])
        billing_month = int(data["billing_month"])

        contract = BaseContract.query.get(contract_id)
        if not contract:
            return jsonify({"error": "关联合同未找到"}), 404

        # employee_id = contract.user_id or contract.service_personnel_id
        employee_id = contract.service_personnel_id

        if not employee_id:
            return jsonify({"error": "合同未关联任何员工"}), 400

        attendance_record = AttendanceRecord.query.filter_by(
            contract_id=contract_id, cycle_start_date=cycle_start
        ).first()

        if attendance_record:
            attendance_record.overtime_days = overtime_days
            # 删掉 statutory_holiday_days 的处理，统一为 overtime_days
            attendance_record.statutory_holiday_days = 0
            attendance_record.total_days_worked = 26 + overtime_days
            msg = "考勤记录更新成功"
        else:
            attendance_record = AttendanceRecord(
                employee_id=employee_id,
                contract_id=contract_id,
                cycle_start_date=cycle_start,
                cycle_end_date=cycle_end,
                overtime_days=overtime_days,
                statutory_holiday_days=0,
                total_days_worked=26 + overtime_days,
            )
            db.session.add(attendance_record)
            msg = "考勤记录创建成功"

        db.session.commit()

        # 自动触发重算
        engine = BillingEngine()
        engine.calculate_for_month(billing_year, billing_month)

        latest_details = get_billing_details_internal(
            contract_id=contract_id,
            year=billing_year,
            month=billing_month,
            cycle_start_date_from_bill=cycle_start,
            is_substitute_bill=False,
        )

        return jsonify(
            {"message": f"{msg}，账单已自动重算。", "latest_details": latest_details}
        )

    except Exception as e:
        current_app.logger.error(f"保存考勤记录失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500


@billing_bp.route("/bills", methods=["GET"])
@admin_required
def get_bills():
    try:
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 100, type=int)
        search_term = request.args.get("search", "").strip()
        contract_type = request.args.get("type", "")
        status = request.args.get("status", "")
        billing_month_str = request.args.get("billing_month")
        payment_status_filter = request.args.get("payment_status", "")
        payout_status_filter = request.args.get("payout_status", "") # V2 修改

        if not billing_month_str:
            return jsonify({"error": "必须提供账单月份 (billing_month) 参数"}), 400

        billing_year, billing_month = map(int, billing_month_str.split("-"))

        contract_poly = with_polymorphic(BaseContract, "*")
        query = (
            db.session.query(CustomerBill, contract_poly)
            .select_from(CustomerBill)
            .join(contract_poly, CustomerBill.contract_id == contract_poly.id)
            .outerjoin(
                ServicePersonnel,
                contract_poly.service_personnel_id == ServicePersonnel.id,
            )
        )

        query = query.filter(
            CustomerBill.year == billing_year, CustomerBill.month == billing_month
        )

        if status:
            query = query.filter(contract_poly.status == status)
        if contract_type:
            query = query.filter(contract_poly.type == contract_type)
        if search_term:
            query = query.filter(
                db.or_(
                    contract_poly.customer_name.ilike(f"%{search_term}%"),
                    contract_poly.customer_name_pinyin.ilike(f"%{search_term}%"),
                    ServicePersonnel.name.ilike(f"%{search_term}%"),
                    ServicePersonnel.name_pinyin.ilike(f"%{search_term}%"),
                )
            )

        if payment_status_filter:
            try:
                status_enum = PaymentStatus(payment_status_filter)
                query = query.filter(CustomerBill.payment_status == status_enum)
            except ValueError:
                pass

        # 【V2 修改】根据新的 payout_status 枚举进行过滤
        if payout_status_filter:
            query = query.join(
                EmployeePayroll,
                and_(
                    EmployeePayroll.contract_id == CustomerBill.contract_id,
                    EmployeePayroll.cycle_start_date == CustomerBill.cycle_start_date,
                ),
            )
            try:
                status_enum = PayoutStatus(payout_status_filter)
                query = query.filter(EmployeePayroll.payout_status == status_enum)
            except ValueError:
                pass

        query = query.order_by(CustomerBill.total_due.desc())

        filtered_bill_ids_query = query.with_entities(CustomerBill.id)
        filtered_bill_ids = [item[0] for item in filtered_bill_ids_query.all()]


        # --- Linus 改造开始：计算新的财务摘要 ---
        total_management_fee = D(0)
        total_receivable = D(0)
        total_deposit = D(0)
        total_introduction_fee = D(0)
        total_customer_increase = D(0)
        total_employee_payable = D(0)
        total_security_deposit = D(0)

        if filtered_bill_ids:
            # 1. 先把原始管理费算出来
            original_management_fee_query = db.session.query(
                func.sum(CustomerBill.calculation_details['management_fee'].as_float())
            ).filter(
                CustomerBill.id.in_(filtered_bill_ids),
                CustomerBill.calculation_details.has_key('management_fee')
            )
            original_total_management_fee =D(original_management_fee_query.scalar() or 0)

            # 2. 找到与这些账单关联的所有薪酬单ID
            payroll_ids_query = db.session.query(EmployeePayroll.id).join(
                CustomerBill,
                and_(
                    EmployeePayroll.contract_id == CustomerBill.contract_id,
                    EmployeePayroll.cycle_start_date ==CustomerBill.cycle_start_date,
                    EmployeePayroll.is_substitute_payroll ==CustomerBill.is_substitute_bill
                )
            ).filter(CustomerBill.id.in_(filtered_bill_ids))
            payroll_ids = [item[0] for item in payroll_ids_query.all()]

            # 3. 一次性查询所有相关的财务调整项
            adjustments_query = db.session.query(
                FinancialAdjustment.adjustment_type,
                FinancialAdjustment.amount,
                FinancialAdjustment.description
            ).filter(
                or_(FinancialAdjustment.customer_bill_id.in_(filtered_bill_ids),
                    FinancialAdjustment.employee_payroll_id.in_(payroll_ids)
                )
            )
            all_adjustments = adjustments_query.all()

            # 4. 遍历调整项，分类汇总
            for adj in all_adjustments:
                if adj.adjustment_type == AdjustmentType.DEPOSIT:
                    total_deposit += adj.amount
                elif adj.adjustment_type == AdjustmentType.INTRODUCTION_FEE:
                    total_introduction_fee += adj.amount
                elif adj.adjustment_type == AdjustmentType.CUSTOMER_INCREASE:
                    total_customer_increase += adj.amount
                elif adj.adjustment_type == AdjustmentType.EMPLOYEE_COMMISSION: # 员工首月10%返佣
                    total_employee_payable += adj.amount

                # 根据描述文字处理“保证金”
                if adj.description and '保证金退款' in adj.description:
                    total_security_deposit -= adj.amount
                elif adj.description and '保证金' in adj.description:
                    total_security_deposit += adj.amount

            # 5. 计算最终的指标
            # 新的管理费总计 = 原管理费 + 员工应缴款
            total_management_fee = original_total_management_fee +total_employee_payable

            # 应收款总计 = 各分项之和
            total_receivable = (
                total_deposit +
                total_introduction_fee +
                original_total_management_fee + # 注意：这里用的是【原始】管理费
                total_customer_increase +
                total_security_deposit +
                total_employee_payable
            )
        # --- Linus 改造结束 ---

        paginated_results = query.paginate(
            page=page, per_page=per_page, error_out=False
        )

        results = []
        engine = BillingEngine()

        # 【V2 修改】定义状态到中文的映射
        status_map = {
            PaymentStatus.PAID: "已支付",
            PaymentStatus.UNPAID: "未支付",
            PaymentStatus.PARTIALLY_PAID: "部分支付",
            PaymentStatus.OVERPAID: "超额支付",
        }

        payout_status_map = {
            PayoutStatus.PAID: "已发放",
            PayoutStatus.UNPAID: "未发放",
            PayoutStatus.PARTIALLY_PAID: "部分发放",
        }

        for i, (bill, contract) in enumerate(paginated_results.items):
            payroll = EmployeePayroll.query.filter_by(
                contract_id=bill.contract_id, cycle_start_date=bill.cycle_start_date, is_substitute_payroll=bill.is_substitute_bill
            ).first()

            invoice_balance = engine.calculate_invoice_balance(str(bill.id))

            # --- 新增逻辑：查询员工应缴款 ---
            employee_payable_amount = D(0)
            employee_payable_is_settled = True # 如果没有应缴款，默认为已结清
            if payroll:
                # --- 这是修改点 ---
                payable_adjustments = FinancialAdjustment.query.filter(
                    FinancialAdjustment.employee_payroll_id == payroll.id,
                    FinancialAdjustment.adjustment_type.in_([
                        AdjustmentType.EMPLOYEE_DECREASE,
                        AdjustmentType.EMPLOYEE_COMMISSION
                    ])
                ).all()

                if payable_adjustments:
                    total_payable = sum(adj.amount for adj in payable_adjustments)
                    employee_payable_amount = total_payable
                    # 只有当所有应缴款都结清时，总状态才算“已缴纳”
                    employee_payable_is_settled = all(adj.is_settled for adj in payable_adjustments)
            # --- 新增逻辑结束 ---

            item = {
                "id": str(bill.id),
                "contract_id": str(contract.id),
                "customer_name": contract.customer_name,
                "status": contract.status,
                "customer_payable": str(bill.total_due),
                "customer_total_paid": str(bill.total_paid),
                "customer_is_paid": bill.payment_status == PaymentStatus.PAID,
                "is_deferred": False,
                "employee_payout": str(payroll.total_due) if payroll else "待计算",
                "employee_is_paid": payroll.payout_status == PayoutStatus.PAID if payroll else False,

                # --- 新增字段 ---
                "employee_payable_amount": str(employee_payable_amount),
                "employee_payable_is_settled": employee_payable_is_settled,
                # --- 新增结束 ---

                "is_substitute_bill": bill.is_substitute_bill,
                "contract_type_label": get_contract_type_details(contract.type),
                "is_monthly_auto_renew": getattr(contract, 'is_monthly_auto_renew', False),
                "contract_type_value": contract.type,
                "employee_level": str(contract.employee_level or "0"),
                "active_cycle_start": bill.cycle_start_date.isoformat() if bill.cycle_start_date else None,
                "active_cycle_end": bill.cycle_end_date.isoformat() if bill.cycle_end_date else None,
                "invoice_needed": invoice_balance.get("auto_invoice_needed", False),
                "remaining_invoice_amount": str(invoice_balance.get("remaining_un_invoiced","0.00")),
                "payment_status_label": status_map.get(bill.payment_status, "未知"),
                "payout_status_label": payout_status_map.get(payroll.payout_status, "未知") if payroll else "未知",
            }

            start = contract.start_date.isoformat() if contract.start_date else "—"
            end = contract.end_date.isoformat() if contract.end_date else "—"
            item["contract_period"] = f"{start} ~ {end}"

            original_employee = contract.service_personnel
            item["employee_name"] = (
                original_employee.name if original_employee else "未知员工"
            )

            if bill.is_substitute_bill:
                item["contract_type_label"] = (
                    f"{item.get('contract_type_label', '未知类型')} (替)"
                )
                sub_record = bill.source_substitute_record
                if sub_record:
                    sub_employee = (
                        sub_record.substitute_user or sub_record.substitute_personnel
                    )
                    item["employee_name"] = getattr(
                        sub_employee,
                        "username",
                        getattr(sub_employee, "name", "未知替班员工"),
                    )
                    item["employee_level"] = str(sub_record.substitute_salary or "0")
                    item["active_cycle_start"] = sub_record.start_date.isoformat()
                    item["active_cycle_end"] = sub_record.end_date.isoformat()
                else:
                    item["employee_name"] = "替班(记录丢失)"
                    item["employee_level"] = "N/A"

            results.append(item)

        return jsonify(
            {
                "items": results,
                "total": paginated_results.total,
                "page": paginated_results.page,
                "per_page": paginated_results.per_page,
                "pages": paginated_results.pages,
                "summary": {
                    "total_management_fee": str(total_management_fee.quantize(D('0.01'))),
                    "total_receivable": str(total_receivable.quantize(D('0.01')))
                }
            }
        )
    except Exception as e:
        current_app.logger.error(f"获取账单列表失败: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# --- 新增：批量获取账单摘要的API ---
@billing_bp.route("/summary", methods=["POST"])
@admin_required
def get_billing_summaries():
    data = request.get_json()
    contract_ids = data.get("contract_ids", [])
    billing_month_str = data.get("billing_month")
    if not contract_ids or not billing_month_str:
        return jsonify({"error": "缺少合同ID列表或账单月份"}), 400

    billing_year, billing_month = map(int, billing_month_str.split("-"))

    try:
        summaries = (
            db.session.query(CustomerBill.contract_id, CustomerBill.total_due)
            .filter(
                CustomerBill.contract_id.in_(contract_ids),
                CustomerBill.year == billing_year,
                CustomerBill.month == billing_month,
            )
            .all()
        )

        # 将结果转换为字典，方便前端查找
        summary_map = {
            str(contract_id): str(total_due)
            for contract_id, total_due in summaries
        }

        return jsonify(summary_map)
    except Exception as e:
        current_app.logger.error(f"获取账单摘要失败: {e}", exc_info=True)
        return jsonify({"error": "获取账单摘要时发生服务器内部错误"}), 500


@billing_bp.route("/details", methods=["GET"])
@admin_required
def get_billing_details():
    bill_id = request.args.get("bill_id")
    contract_id = request.args.get("contract_id")
    month_str = request.args.get("month")

    target_bill = None

    if bill_id:
        target_bill = db.session.get(CustomerBill, bill_id)
    elif contract_id and month_str:
        try:
            year, month = map(int, month_str.split('-'))
            # 假设在月份导航时，我们总是关注主账单而非替班账单
            target_bill = CustomerBill.query.filter_by(
                contract_id=contract_id,
                year=year,
                month=month,
                is_substitute_bill=False
            ).order_by(CustomerBill.cycle_start_date.desc()).first()
        except ValueError:
            return jsonify({"error": "无效的月份格式，应为 YYYY-MM"}), 400
    
    if not target_bill:
        error_msg = f"账单未找到。查询参数: bill_id={bill_id}, contract_id={contract_id}, month={month_str}"
        return jsonify({"error": error_msg}), 404

    # 从这里开始，我们确保已经获取到了 target_bill，后续逻辑可以统一
    bill_id = str(target_bill.id)

    try:
        details = get_billing_details_internal(bill_id=bill_id)
        if details is None:
            return jsonify({"error": "获取账单详情失败"}), 404

        engine = BillingEngine()
        invoice_balance = engine.calculate_invoice_balance(bill_id)
        details["invoice_balance"] = invoice_balance
        details["invoice_needed"] = (target_bill.payment_details or {}).get("invoice_needed", False)

        return jsonify(details)
    except Exception as e:
        current_app.logger.error(f"获取账单详情失败 (bill_id: {bill_id}): {e}", exc_info=True)
        return jsonify({"error": "获取账单详情时发生服务器内部错误"}), 500


# --- 新增：计算前预检查的API ---
@billing_bp.route("/pre-check", methods=["POST"])  # <--- 方法从 GET 改为 POST
@admin_required
def pre_check_billing():
    # --- 核心修正：从请求体中获取合同ID列表 ---
    data = request.get_json()
    if not data:
        return jsonify({"error": "缺少请求体"}), 400

    contract_ids = data.get("contract_ids", [])
    if not contract_ids:
        # 如果前端列表为空，直接返回空结果，是正常情况
        return jsonify([])

    # 查找这些合同中，是月嫂合同且缺少实际上户日期的
    missing_date_contracts = (
        MaternityNurseContract.query.filter(
            MaternityNurseContract.id.in_(contract_ids),  # <--- 只在提供的ID中查找
            MaternityNurseContract.actual_onboarding_date is None,
        )
        .join(
            ServicePersonnel,
            BaseContract.service_personnel_id == ServicePersonnel.id,
            isouter=True,
        )
        .add_columns(ServicePersonnel.name.label("sp_name"))
        .all()
    )

    results = []
    for contract, sp_name in missing_date_contracts:
        results.append(
            {
                "id": str(contract.id),
                "customer_name": contract.customer_name,
                "employee_name": sp_name or "未知员工",
                "provisional_start_date": contract.provisional_start_date.isoformat()
                if contract.provisional_start_date
                else None,
            }
        )

    return jsonify(results)


@billing_bp.route("/contracts/<uuid:contract_id>", methods=["PUT"])
@admin_required
def update_single_contract(contract_id):
    """
    一个通用的、用于更新单个合同字段的API。
    当月嫂合同的实际上户日期被设置时，会触发一个后台任务来串行生成所有账单。
    """
    contract = db.session.get(BaseContract, str(contract_id))
    if not contract:
        return jsonify({"error": "合同未找到"}), 404

    data = request.get_json()
    if not data:
        return jsonify({"error": "缺少更新数据"}), 400

    try:
        should_generate_bills = False
        should_recalculate_first_bill = False
        log_details = {}

        if "actual_onboarding_date" in data:
            new_onboarding_date_str = data["actual_onboarding_date"]
            if new_onboarding_date_str:
                new_onboarding_date = datetime.strptime(new_onboarding_date_str, "%Y-%m-%d")
                if isinstance(contract, MaternityNurseContract):

                    existing_date_obj = contract.actual_onboarding_date
                    existing_date = existing_date_obj.date() if isinstance(existing_date_obj,datetime) else existing_date_obj

                    if existing_date != new_onboarding_date.date():
                        current_app.logger.info(f"====合同==== {contract.id} 的实际上户日期从 {existing_date} 更新为 {new_onboarding_date.date()}")
                        contract.actual_onboarding_date = new_onboarding_date

                        provisional_start_obj = contract.provisional_start_date
                        provisional_start = provisional_start_obj.date() if isinstance(provisional_start_obj, datetime) else provisional_start_obj

                        original_end_obj = contract.end_date
                        original_end = original_end_obj.date() if isinstance(original_end_obj,datetime) else original_end_obj

                        if provisional_start and original_end:
                            total_days = (original_end - provisional_start).days
                            contract.expected_offboarding_date = new_onboarding_date +timedelta(days=total_days)
                        else:
                            contract.expected_offboarding_date = new_onboarding_date +timedelta(days=26)
                        log_details['实际上户日期'] = {'from': str(existing_date), 'to': str(new_onboarding_date.date())}
                        should_generate_bills = True
                else:
                    return jsonify({"error": "只有月嫂合同才能设置实际上户日期"}), 400
            else:
                if isinstance(contract, MaternityNurseContract):
                    contract.actual_onboarding_date = None
                    contract.expected_offboarding_date = None

        if "introduction_fee" in data and hasattr(contract, 'introduction_fee'):
            new_fee = D(data["introduction_fee"] or 0)
            should_recalculate_first_bill = True
            # upsert_introduction_fee_adjustment(contract)
            if contract.introduction_fee != new_fee:
                # should_generate_bills = True
                log_details['介绍费'] = {'from': str(contract.introduction_fee), 'to': str(new_fee)}
                contract.introduction_fee = new_fee

        if 'requires_signature' in data:
            requires_signature = data['requires_signature']
            contract.requires_signature = requires_signature
            if requires_signature is False:
                contract.signing_status = SigningStatus.NOT_REQUIRED
                # 无需签署时，自动激活合同
                if contract.status not in ['active', 'finished', 'terminated']:
                    contract.status = 'active'
                    update_salary_history_on_contract_activation(contract)
                    current_app.logger.info(f"合同 {contract_id} 已激活（无需签署），已更新薪资历史")
            elif requires_signature is True:
                contract.signing_status = SigningStatus.UNSIGNED
            current_app.logger.info(
                f"合同 {contract_id} 的签署需求更新为: {requires_signature}, 签署状态: {contract.signing_status}, 合同状态: {contract.status}"
            )
                

        if "notes" in data:
            original_note = contract.notes or ""
            appended_note = data["notes"] or ""
            separator = "\\n\\n--- 运营备注 ---\\n"
            if separator in original_note:
                base_note = original_note.split(separator)[0]
                new_full_note = f"{base_note}{separator}{appended_note}"
            else:
                new_full_note = f"{original_note}{separator}{appended_note}"
            if contract.notes != new_full_note:
                log_details['备注'] = {'from': contract.notes, 'to': new_full_note}
                contract.notes = new_full_note

        if log_details:
            any_bill = CustomerBill.query.filter_by(contract_id=contract.id).first()
            any_payroll = EmployeePayroll.query.filter_by(contract_id=contract.id).first()
            _log_activity(any_bill, any_payroll, "更新了合同详情", details=log_details)

        db.session.commit()

        if should_generate_bills:
            current_app.logger.info(f"为合同 {contract.id} 触发统一的后台账单生成任务...")
            generate_all_bills_for_contract_task.delay(str(contract.id))
            return jsonify({"message": "合同信息更新成功，并已在后台开始生成所有相关账单。"})
        
        # --- BEGIN: 我们新增的重算逻辑 ---
        if should_recalculate_first_bill:
            # 找到这份合同的第一个账单,在首月账单中新增”介绍费“的财务调整项
            first_bill = CustomerBill.query.filter_by(
                contract_id=contract.id,
                is_substitute_bill=False
            ).order_by(CustomerBill.cycle_start_date.asc()).first()

            if first_bill:
                # 如果找到了，就为它触发一个后台重算任务
                calculate_monthly_billing_task.delay(
                    year=first_bill.year,
                    month=first_bill.month,
                    contract_id=str(contract.id),
                    force_recalculate=True
                )
                current_app.logger.info(f"介绍费更新后，已为合同 {contract.id} 的首月账单 {first_bill.id} 触发重算。")
                return jsonify({"message": "介绍费更新成功，首月账单已提交重算。"})
        # --- END: 新增逻辑结束 ---

        return jsonify({"message": "合同信息更新成功"})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新合同 {contract_id} 失败: {e}", exc_info=True)
        return jsonify({"error": "更新合同失败，服务器内部错误"}), 500

def _apply_all_adjustments_and_settlements(adjustments_data, bill, payroll):
    user_id = get_jwt_identity()
    old_adjustments = FinancialAdjustment.query.filter(
        or_(FinancialAdjustment.customer_bill_id == bill.id,FinancialAdjustment.employee_payroll_id == payroll.id)
    ).all()
    old_adjustments_map = {str(adj.id): adj for adj in old_adjustments}
    new_adjustments_ids = {str(adj.get('id')) for adj in adjustments_data if adj.get('id')}

    was_deletion = False # 初始化标志

    # Deletion Logic
    for old_id, old_adj in old_adjustments_map.items():
        if old_id not in new_adjustments_ids:
            was_deletion = True # 在实际删除时设置标志
            # !!! 新增日志 !!!
            current_app.logger.info(f"!!! DELETION DETECTED !!! Deleting adj {old_adj.id} ('{old_adj.description}'). Flag was_deletion is now True.")

            if old_adj.is_settled and old_adj.details and'linked_record' in old_adj.details:
                linked_record_info = old_adj.details['linked_record']
                record_id, record_type =linked_record_info.get('id'),linked_record_info.get('type')
                if record_id and record_type == 'payment':
                    db.session.query(PaymentRecord).filter_by(id=record_id).delete()
                elif record_id and record_type == 'payout':
                    db.session.query(PayoutRecord).filter_by(id=record_id).delete()
            db.session.delete(old_adj)

    # 关键修复：在处理新增/更新前，先将删除操作刷新到会话中
    db.session.flush()
    current_app.logger.info(f"处理财务调整项，共 {len(adjustments_data)} 项。")
    for adj_data in adjustments_data:
        adj_id = str(adj_data.get('id', ''))
        adj_type = AdjustmentType(adj_data["adjustment_type"])
        adj_amount = D(adj_data["amount"])
        adj_description = adj_data["description"]
        is_settled = adj_data.get('is_settled', False)
        settlement_date_str = adj_data.get('settlement_date')
        settlement_date = date_parse(settlement_date_str).date() if settlement_date_str else date.today()
        settlement_details = adj_data.get('settlement_details')
        current_app.logger.info(f"处理调整项，ID: {adj_id or '新建'}, 类型: {adj_type.name}, 金额: {adj_amount}, 结算状态: {is_settled}")
        if adj_id and adj_id in old_adjustments_map:
            current_app.logger.info(f"=== 发现已有调整项，准备更新 ID: {adj_id} ===")
            existing_adj = old_adjustments_map[adj_id]
            if existing_adj.details is None: existing_adj.details = {}

            settlement_method = (settlement_details or {}).get('notes') or'settlement'
            current_app.logger.info(f"处理调整项更新，ID: {adj_id}, 类型: {adj_type.name}, 金额:{adj_amount}, 结算状态: {is_settled}")
            # 状态变更1：从未结算 -> 已结算 (创建记录)
            if is_settled and not existing_adj.is_settled:
                # current_app.logger.info(f"[SETTLEMENT_DEBUG] Condition MET: is_settled and notexisting_adj.is_settled. Creating record for adj_type: {adj_type.value}")
                new_record, record_type = None, None
                record_notes = f"[来自结算调整项] {adj_description}"
                settlement_method = (settlement_details or {}).get('notes') or 'settlement'

                customer_increase_types =[AdjustmentType.CUSTOMER_INCREASE,AdjustmentType.INTRODUCTION_FEE, AdjustmentType.DEPOSIT,AdjustmentType.COMPANY_PAID_SALARY,AdjustmentType.SUBSTITUTE_MANAGEMENT_FEE,AdjustmentType.DEFERRED_FEE, AdjustmentType.COMPANY_PAID_SALARY]
                customer_decrease_types =[AdjustmentType.CUSTOMER_DECREASE,AdjustmentType.CUSTOMER_DISCOUNT]
                employee_increase_types =[AdjustmentType.EMPLOYEE_INCREASE,AdjustmentType.EMPLOYEE_CLIENT_PAYMENT,AdjustmentType.DEPOSIT_PAID_SALARY]
                employee_decrease_types = [AdjustmentType.EMPLOYEE_DECREASE]  # EMPLOYEE_COMMISSION 首月返佣 不计入已发总额计算

                if adj_type in customer_increase_types:
                    new_record = PaymentRecord(customer_bill_id=bill.id, amount=abs(adj_amount),payment_date=settlement_date, method=settlement_method,notes=record_notes,created_by_user_id=user_id)
                    record_type = 'payment'
                elif adj_type in customer_decrease_types:
                    new_record = PaymentRecord(customer_bill_id=bill.id, amount=abs(adj_amount)*-1, payment_date=settlement_date, method=settlement_method, notes=f"[客户退款]{adj_description}", created_by_user_id=user_id)
                    record_type = 'payment'
                elif adj_type in employee_increase_types:
                    new_record = PayoutRecord(employee_payroll_id=payroll.id, amount=abs(adj_amount), payout_date=settlement_date, method=settlement_method, notes=record_notes, payer='公司', created_by_user_id=user_id)
                    record_type = 'payout'
                elif adj_type in employee_decrease_types:
                    payer_name = payroll.contract.service_personnel.name
                    new_record = PayoutRecord(employee_payroll_id=payroll.id, amount=abs (adj_amount) * -1, payout_date=settlement_date, method=settlement_method, notes=f"[员工缴款] {adj_description}", payer=payer_name, created_by_user_id=user_id)
                    record_type = 'payout'

                if new_record:
                    # current_app.logger.info(f"[SETTLEMENT_DEBUG] new_record created. Type: {record_type}, Amount: {new_record.amount}")
                    db.session.add(new_record)
                    db.session.flush()
                    existing_adj.details['linked_record'] = {'id': str(new_record.id), 'type':record_type}


            # 状态变更2：从已结算 -> 未结算 (删除记录)
            elif not is_settled and existing_adj.is_settled:
                linked_record_info = existing_adj.details.pop('linked_record',None)
                if linked_record_info:
                    record_id, record_type = linked_record_info.get('id'),linked_record_info.get('type')
                    if record_id and record_type == 'payment':
                        db.session.query(PaymentRecord).filter_by(id=record_id).delete()
                    elif record_id and record_type == 'payout':
                        db.session.query(PayoutRecord).filter_by(id=record_id).delete()

            # 状态变更3：已结算 -> 已结算 (更新记录)
            elif is_settled and existing_adj.is_settled:
                linked_record_info = existing_adj.details.get('linked_record')
                if linked_record_info:
                    record_id, record_type = linked_record_info.get('id'),linked_record_info.get('type')
                    record_to_update = None
                    if record_type == 'payment':
                        record_to_update = db.session.get(PaymentRecord,record_id)
                        if record_to_update: record_to_update.payment_date =settlement_date
                    elif record_type == 'payout':
                        record_to_update = db.session.get(PayoutRecord, record_id)
                        if record_to_update: record_to_update.payout_date =settlement_date

                    if record_to_update:
                        record_to_update.method = settlement_method
                        # 根据类型决定金额是正是负
                        if adj_type in[AdjustmentType.CUSTOMER_DECREASE,AdjustmentType.CUSTOMER_DISCOUNT,AdjustmentType.EMPLOYEE_DECREASE, AdjustmentType.EMPLOYEE_COMMISSION]:
                             record_to_update.amount = abs(adj_amount) * -1
                        else:
                             record_to_update.amount = abs(adj_amount)

            # 更新 adjustment 自身的信息
            existing_adj.amount = abs(adj_amount)
            existing_adj.description = adj_description
            existing_adj.adjustment_type = adj_type
            existing_adj.is_settled = is_settled
            existing_adj.settlement_date = settlement_date
            existing_adj.settlement_details = settlement_details
            attributes.flag_modified(existing_adj, "details")
            # --- 日志记录：更新 ---
            change_details = {}
            if existing_adj.amount != abs(adj_amount):
                change_details['amount'] = {'from': str(existing_adj.amount), 'to': str(abs(adj_amount))}
            if existing_adj.description != adj_description:
                change_details['description'] = {'from':existing_adj.description, 'to':adj_description}
            if existing_adj.is_settled != is_settled:
                change_details['is_settled'] = {'from':existing_adj.is_settled, 'to':is_settled}

            if change_details:
                # 加上调整项的基础信息，方便识别
                change_details['type']=ADJUSTMENT_TYPE_LABELS.get(existing_adj.adjustment_type,existing_adj.adjustment_type.name)
                change_details['original_description'] =existing_adj.description
                _log_activity(bill, payroll,"更新了财务调整项", details=change_details)
            # --- 日志结束 ---

        elif not adj_id or 'temp' in adj_id:
            # 新增调整项的逻辑
            current_app.logger.info(f"[ADJ_CREATE_DEBUG] Preparing to create new adjustment with data: {adj_data}")

            new_adj = FinancialAdjustment(
                adjustment_type=adj_type, amount=abs(adj_amount),description=adj_description,
                date=bill.cycle_start_date,is_settled=is_settled,
                settlement_date=settlement_date,settlement_details=settlement_details,
                details={} # 初始化 details 字典
            )

            customer_types = [
                AdjustmentType.CUSTOMER_INCREASE,
                AdjustmentType.CUSTOMER_DECREASE,
                AdjustmentType.CUSTOMER_DISCOUNT,
                AdjustmentType.INTRODUCTION_FEE,
                AdjustmentType.DEFERRED_FEE,
                AdjustmentType.DEPOSIT,
                AdjustmentType.COMPANY_PAID_SALARY # <--- 修正：将公司代付工资归类为客户账单侧的调整
            ]
            if adj_type in customer_types:
                new_adj.customer_bill_id = bill.id
            else:
                new_adj.employee_payroll_id = payroll.id
            db.session.add(new_adj)

            # --- 日志记录：新增 ---
            log_details = {
                "type":ADJUSTMENT_TYPE_LABELS.get(new_adj.adjustment_type,new_adj.adjustment_type.name),
                "amount": str(new_adj.amount),
                "description": new_adj.description,
                "is_settled": new_adj.is_settled
            }
            _log_activity(bill, payroll, "新增了财务调整项",details=log_details)
            # --- 日志结束 ---

            # --- 【核心修复】如果新增项同时被结算，立即创建对应的收付款记录 ---
            if is_settled:
                new_record, record_type = None, None
                settlement_method = (settlement_details or{}).get('notes') or 'settlement'
                record_notes = f"[来自结算调整项] {adj_description}"

                if adj_type ==AdjustmentType.CUSTOMER_INCREASE:
                    new_record =PaymentRecord(customer_bill_id=bill.id, amount=abs(adj_amount),payment_date=settlement_date,method=settlement_method,notes=record_notes,created_by_user_id=user_id)
                    record_type = 'payment'
                elif adj_type in[AdjustmentType.CUSTOMER_DECREASE,AdjustmentType.CUSTOMER_DISCOUNT]:
                    new_record =PaymentRecord(customer_bill_id=bill.id, amount=abs(adj_amount)*-1, payment_date=settlement_date, method=settlement_method,notes=f"[客户退款]{adj_description}",created_by_user_id=user_id)
                    record_type = 'payment'
                elif adj_type ==AdjustmentType.EMPLOYEE_INCREASE:
                    new_record =PayoutRecord(employee_payroll_id=payroll.id, amount=abs(adj_amount), payout_date=settlement_date,method=settlement_method, notes=record_notes, payer='公司',created_by_user_id=user_id)
                    record_type = 'payout'
                elif adj_type ==AdjustmentType.EMPLOYEE_DECREASE:
                    new_record =PayoutRecord(employee_payroll_id=payroll.id, amount=abs(adj_amount) * -1, payout_date=settlement_date,method=settlement_method, notes=f"[员工缴款]{adj_description}", payer=employee_name, created_by_user_id=user_id)
                    record_type = 'payout'
                elif adj_type ==AdjustmentType.EMPLOYEE_COMMISSION:
                    new_record =PayoutRecord(employee_payroll_id=payroll.id, amount=abs(adj_amount), payout_date=settlement_date,method=settlement_method, notes=f"[首月返佣]{adj_description}", payer='公司',created_by_user_id=user_id)
                    record_type = 'payout'

                if new_record:
                    db.session.add(new_record)
                    db.session.flush() # 刷新以获取 new_record.id
                    new_adj.details['linked_record'] = {'id':str(new_record.id), 'type':record_type}
                    attributes.flag_modified(new_adj,"details")
            # --- 修复结束 ---
    if was_deletion:
        current_app.logger.info(f"!!! FLAG CHECK !!! Deletion occurred in _apply_all_adjustments_and_settlements for bill {bill.id}. Returning was_deletion=True.")
    return was_deletion

@billing_bp.route("/batch-update", methods=["POST"])
@admin_required
def batch_update_billing_details():
    # First, check if the request has the correct Content-Type header
    content_type = request.headers.get('Content-Type', '')

    # If no Content-Type header or not application/json, try to detect and handle accordingly
    if not content_type or not content_type.startswith('application/json'):
        # Try to parse as JSON anyway, in case the client omitted Content-Type
        # but still sent JSON data (common frontend behavior)
        try:
            if request.data:
                data = request.get_data(as_text=True)
                if data.strip():
                    import json
                    data = json.loads(data)
                else:
                    data = {}
            else:
                data = {}
        except Exception as e:
            return jsonify({"error": f"请求体解析失败: {str(e)}"}), 400
    else:
        # Normal JSON request with proper Content-Type header
        try:
            data = request.get_json()
        except Exception as e:
            return jsonify({"error": f"JSON解析失败: {str(e)}"}), 400

    bill_id = data.get("bill_id")
    if not bill_id:
        return jsonify({"error": "请求体中缺少 bill_id"}), 400

    try:
        bill = db.session.get(CustomerBill, bill_id)
        if not bill:
            return jsonify({"error": "账单未找到"}), 404

        payroll = EmployeePayroll.query.filter_by(
            contract_id=bill.contract_id,
            cycle_start_date=bill.cycle_start_date,
            is_substitute_payroll=bill.is_substitute_bill,
        ).first()

        if not payroll:
            return jsonify({"error": f"未找到与账单ID {bill_id} 关联的薪酬单"}), 404

        engine = BillingEngine()

        # 【V2.0 重构】
        # 这个函数现在只负责处理非支付类的更新，例如：
        # - 调整项 (通过 _apply_all_adjustments_and_settlements)
        # - 发票记录
        # - 加班天数
        # - 实际工作天数
        # - 账单是否需要开票
        #
        # 所有关于“支付/结算”状态的逻辑都已被移除，它们现在由 add_payment_record API 统一处理。

        # --- 1. 处理发票记录 ---
        # --- 0. 准备日志记录 ---
        log_details = {}

        # 记录原始加班天数
        original_overtime_days = D("0")
        attendance_record=AttendanceRecord.query.filter_by(contract_id=bill.contract_id,cycle_start_date=bill.cycle_start_date).first()
        if attendance_record:
            original_overtime_days =attendance_record.overtime_days

        # 记录原始实际工作天数
        original_actual_work_days = bill.actual_work_days

        # 记录原始发票需求状态
        original_invoice_needed = bill.invoice_needed

        # 记录原始发票信息
        original_invoices_info = [f"发票号: {inv.invoice_number}, 金额: {inv.amount}, 日期:{inv.issue_date}" for inv in bill.invoices]
        invoices_from_frontend = data.get('invoices', [])
        if invoices_from_frontend:
            existing_invoices_map = {str(inv.id): inv for inv in bill.invoices}
            new_invoice_ids = {str(inv_data.get('id')) for inv_data in invoices_from_frontend if inv_data.get('id')}
            for inv_data in invoices_from_frontend:
                inv_id = inv_data.get('id')
                if inv_id and str(inv_id) in existing_invoices_map:
                    invoice_to_update = existing_invoices_map[str(inv_id)]
                    invoice_to_update.amount = D(inv_data.get('amount', '0'))
                    invoice_to_update.issue_date = date_parse(inv_data.get('issue_date')).date()if inv_data.get('issue_date')else None
                    invoice_to_update.invoice_number = inv_data.get('invoice_number')
                    invoice_to_update.notes = inv_data.get('notes')
                elif not inv_id or 'temp' in str(inv_id):
                    db.session.add(InvoiceRecord(
                        customer_bill_id=bill.id,
                        amount=D(inv_data.get('amount', '0')),
                        issue_date=date_parse(inv_data.get('issue_date')).date() if inv_data.get('issue_date') else None,
                        invoice_number=inv_data.get('invoice_number'),
                        notes=inv_data.get('notes')
                    ))
            for old_id, old_invoice in existing_invoices_map.items():
                if old_id not in new_invoice_ids:
                    db.session.delete(old_invoice)

        # --- 2. 处理其他可编辑字段 ---
        new_actual_work_days = D(str(data['actual_work_days'])) if 'actual_work_days' in data and data['actual_work_days'] is not None and data['actual_work_days'] != '' else None
        if new_actual_work_days is not None and bill.actual_work_days != new_actual_work_days:
            bill.actual_work_days = new_actual_work_days
            payroll.actual_work_days = new_actual_work_days

        new_overtime_decimal = D(str(data.get("overtime_days", "0")))
        attendance_record=AttendanceRecord.query.filter_by(contract_id=bill.contract_id,cycle_start_date=bill.cycle_start_date).first()
        if attendance_record:
            if attendance_record.overtime_days.quantize(D('0.01'))!=new_overtime_decimal.quantize(D('0.01')):
                attendance_record.overtime_days = new_overtime_decimal
        elif new_overtime_decimal > 0:
            employee_id = bill.contract.service_personnel_id
            if employee_id:
                db.session.add(AttendanceRecord(
                    employee_id=employee_id, contract_id=bill.contract_id,
                    cycle_start_date=bill.cycle_start_date, cycle_end_date=bill.cycle_end_date,
                    overtime_days=new_overtime_decimal, total_days_worked=0
                ))

        if 'invoice_needed' in data:
            bill.invoice_needed = data.get('invoice_needed', False)

        # --- 3. 处理财务调整项 ---
        was_deletion = False
        if 'adjustments' in data:
            adjustments_data = data.get('adjustments', [])
            was_deletion = _apply_all_adjustments_and_settlements(adjustments_data,bill,payroll)

        # !!! 新增日志 !!!
        current_app.logger.info(f"!!! FLAG CHECK !!! In batch_update_billing_details, was_deletion is: {was_deletion}")

        # --- 3.5. 汇总并记录日志 ---
        if original_overtime_days.quantize(D('0.01')) != new_overtime_decimal.quantize(D('0.01')):
            log_details['加班天数'] = {'from': str(original_overtime_days), 'to': str(new_overtime_decimal)}

        if new_actual_work_days is not None and original_actual_work_days != new_actual_work_days:
            log_details['实际工作天数'] = {'from': str(original_actual_work_days), 'to': str(new_actual_work_days)}

        new_invoice_needed = data.get('invoice_needed', False)
        if original_invoice_needed != new_invoice_needed:
            log_details['是否需要开票'] = {'from':original_invoice_needed, 'to':new_invoice_needed}

        new_invoices_info = [f"发票号: {inv.get('invoice_number')}, 金额: {inv.get('amount')}, 日期: {inv.get('issue_date')}" for inv in data.get('invoices', [])]
        if set(original_invoices_info) != set(new_invoices_info):
            log_details['发票记录'] = {'from':original_invoices_info, 'to': new_invoices_info}

               # 如果有任何变动，则记录日志
        if log_details:
            _log_activity(bill, payroll, "更新账单详情",details=log_details)

        # --- 核心修复：根据账单类型，调用不同的重算逻辑 ---
        if bill.is_substitute_bill:
            current_app.logger.info(f"重算替班账单 {bill.id}，使用覆盖值...")
            engine.calculate_for_substitute(
                substitute_record_id=bill.source_substitute_record_id,
                commit=False,
                overrides={
                    'actual_work_days': new_actual_work_days,
                    'overtime_days': new_overtime_decimal
                }
            )
        else:
            current_app.logger.info(f"重算主账单 {bill.id}，使用覆盖值...")
            engine.calculate_for_month(
                year=bill.year, month=bill.month, contract_id=bill.contract_id,
                force_recalculate=True, actual_work_days_override=new_actual_work_days,
                cycle_start_date_override=bill.cycle_start_date
            )

        # --- 核心修复：在重算后，检查是否为最终账单，若是则更新最终薪资调整项 ---
        contract = bill.contract
        if contract.status in ['terminated', 'finished']:
            last_bill_in_db = CustomerBill.query.filter(
                CustomerBill.contract_id == contract.id,
                CustomerBill.is_substitute_bill == False
            ).order_by(CustomerBill.cycle_end_date.desc()).first()

            if last_bill_in_db and last_bill_in_db.id == bill.id:
                # !!! 新增日志 !!!
                current_app.logger.info(f"!!! CALLING FINAL ADJUSTMENTS !!! Bill {bill.id} is afinal bill. Calling create_final_salary_adjustments with allow_creation={not was_deletion}.")
                engine.create_final_salary_adjustments(bill.id, allow_creation=not was_deletion)
        # --- 修复结束 ---


        # --- 4.5. 在重算总额后，更新最终状态 ---
        _update_bill_payment_status(bill)
        _update_payroll_payout_status(payroll)
        # --- 5. 提交所有更改 ---
        db.session.commit()


        # --- 6. 返回最新数据 ---
        latest_details = get_billing_details_internal(bill_id=bill.id)

        # 【核心修复】手动为返回的数据附加最新的发票余额信息
        if latest_details:
            engine = BillingEngine()
            invoice_balance = engine.calculate_invoice_balance(bill.id)
            latest_details["invoice_balance"] = invoice_balance

        return jsonify({"message": "账单已更新并成功重算！", "latest_details": latest_details})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量更新失败 (bill_id: {bill_id}): {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500


@billing_bp.route("/logs", methods=["GET"])
@admin_required
def get_activity_logs():
    bill_id = request.args.get("bill_id")
    payroll_id = request.args.get("payroll_id")

    # 至少需要一个ID
    if not bill_id and not payroll_id:
        return jsonify({"error": "缺少 bill_id 或 payroll_id 参数"}), 400

    # 构建一个灵活的查询过滤器
    # **核心修正**: 使用 or_() 来查询关联到任意一个ID的日志
    from sqlalchemy import or_

    filters = []
    if bill_id:
        filters.append(FinancialActivityLog.customer_bill_id == bill_id)
    if payroll_id:
        filters.append(FinancialActivityLog.employee_payroll_id == payroll_id)

    logs = (
        FinancialActivityLog.query.filter(or_(*filters))
        .order_by(FinancialActivityLog.created_at.desc())
        .all()
    )

    results = [
        {
            "id": str(log.id),
            "user": log.user.username if log.user else "未知用户",
            "action": log.action,
            "details": log.details,
            "created_at": log.created_at.isoformat(),
        }
        for log in logs
    ]
    return jsonify(results)


# Helper function to get contract type label
# def get_contract_type_details(contract_type):
#     if contract_type == "nanny":
#         return "育儿嫂"
#     elif contract_type == "maternity_nurse":
#         return "月嫂"
#     elif contract_type == "external_substitution":
#         return "临时替工"
#     elif contract_type == "nanny_trial":
#         return "育儿嫂试工"
#     return "未知类型"

@billing_bp.route("/contracts/eligible-for-transfer", methods=["GET"])
@admin_required
def get_eligible_contracts_for_transfer():
    """
    获取用于保证金转移的目标合同列表。
    V2: 支持基于 family_id 的跨客户转移。
    列表包含：
    1. 同一客户名下的其他合同
    2. 同一 family_id 下的其他合同（不同客户名）
    """
    customer_name = request.args.get("customer_name")
    exclude_contract_id = request.args.get("exclude_contract_id")
    family_id = request.args.get("family_id")

    if not customer_name or not exclude_contract_id:
        return jsonify({"error": "缺少 customer_name 或 exclude_contract_id 参数"}), 400

    try:
        # 构建查询条件：同一客户名 OR 同一家庭ID
        filter_conditions = [BaseContract.customer_name == customer_name]

        # 如果有 family_id，添加家庭匹配条件
        family_members = []
        if family_id:
            filter_conditions.append(BaseContract.family_id == family_id)
            # 查询同一家庭下的所有不同客户名
            family_customer_names = (
                db.session.query(BaseContract.customer_name)
                .filter(BaseContract.family_id == family_id)
                .distinct()
                .all()
            )
            family_members = [name[0] for name in family_customer_names if name[0] != customer_name]

        eligible_contracts = (
            BaseContract.query.join(
                ServicePersonnel, BaseContract.service_personnel_id == ServicePersonnel.id
            )
            .filter(
                or_(*filter_conditions),  # 使用 OR 连接条件
                BaseContract.id != exclude_contract_id,
            )
            .order_by(BaseContract.start_date.desc())
            .all()
        )

        results = []
        for contract in eligible_contracts:
            # 如果是不同客户名（通过 family_id 匹配），在标签中显示客户名
            if contract.customer_name != customer_name:
                label = f"{get_contract_type_details(contract.type)} - {contract.service_personnel.name} ({contract.start_date.strftime('%Y-%m-%d')}生效) [客户: {contract.customer_name}]"
            else:
                label = f"{get_contract_type_details(contract.type)} - {contract.service_personnel.name} ({contract.start_date.strftime('%Y-%m-%d')}生效)"

            results.append(
                {
                    "id": str(contract.id),
                    "label": label,
                    "customer_name": contract.customer_name,
                    "is_same_customer": contract.customer_name == customer_name,
                }
            )

        # 返回结果，包含家庭成员信息
        return jsonify({
            "contracts": results,
            "family_members": family_members
        })

    except Exception as e:
        current_app.logger.error(f"获取可转移合同列表失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500


@billing_bp.route("/contracts/substitute-bills", methods=["GET"])
@admin_required
def get_substitute_bills_for_customer():
    """
    获取指定客户名下的所有替班账单列表。
    """
    customer_name = request.args.get("customer_name")
    if not customer_name:
        return jsonify({"error": "缺少 customer_name 参数"}), 400

    try:
        # 通过关联 CustomerBill, BaseContract 和 SubstituteRecord 来获取所需信息
        substitute_bills = (
            db.session.query(CustomerBill)
            .join(BaseContract, CustomerBill.contract_id == BaseContract.id)
            .filter(
                BaseContract.customer_name == customer_name,
                CustomerBill.is_substitute_bill == True,
            )
            .options(
                db.selectinload(CustomerBill.source_substitute_record).selectinload(
                    SubstituteRecord.substitute_user
                ),
                db.selectinload(CustomerBill.source_substitute_record).selectinload(
                    SubstituteRecord.substitute_personnel
                ),
            )
            .order_by(CustomerBill.cycle_start_date.desc())
            .all()
        )

        results = []
        for bill in substitute_bills:
            sub_record = bill.source_substitute_record
            employee_name = "未知员工"
            if sub_record:
                sub_employee = (
                    sub_record.substitute_user or sub_record.substitute_personnel
                )
                if sub_employee:
                    employee_name = getattr(
                        sub_employee, "username", getattr(sub_employee, "name", "未知员工")
                    )

            label = f"替班账单 - {employee_name} ({bill.cycle_start_date.strftime('%Y-%m-%d')} 生效)"
            results.append({"id": str(bill.id), "label": label})

        return jsonify(results)

    except Exception as e:
        current_app.logger.error(
            f"获取客户 {customer_name} 的替班账单列表失败: {e}", exc_info=True
        )
        return jsonify({"error": "服务器内部错误"}), 500

@billing_bp.route("/contracts", methods=["GET"])
@admin_required
def get_all_contracts():
    try:
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 100, type=int)
        search_term = request.args.get("search", "").strip()
        contract_type = request.args.get("type", "")
        status = request.args.get("status", "active")
        employee_id = request.args.get("employee_id")
        customer_name = request.args.get("customer_name")
        deposit_status_filter = request.args.get("deposit_status", "")
        sort_by = request.args.get("sort_by", None)
        sort_order = request.args.get("sort_order", "asc")
        current_app.logger.info(f"Fetching contracts - Page: {page}, Per Page: {per_page}, Search: '{search_term}', Type: '{contract_type}', Status: '{status}', Employee ID: '{employee_id}', Deposit Status: '{deposit_status_filter}', Sort By: '{sort_by}', Sort Order: '{sort_order}'")

        # 2. 【关键修正】最终版的定金子查询，包含了所有支付逻辑
        deposit_adjustment_sq = db.session.query(
            FinancialAdjustment.contract_id,
            FinancialAdjustment.amount,
            FinancialAdjustment.is_settled, # <-- 需要查询 is_settled 字段
            case(
                (
                    # 情况1: 调整项自身状态就是 PAID
                    FinancialAdjustment.status == 'PAID', 'PAID'
                ),
                (
                    # 情况2: 状态是 BILLED 并且 is_settled 标志为 True
                    and_(FinancialAdjustment.status == 'BILLED',FinancialAdjustment.is_settled == True), 'PAID'
                ),
                (
                    # 情况3: 调整项关联的账单状态是 PAID 或 OVERPAID
                    CustomerBill.payment_status.in_([PaymentStatus.PAID,PaymentStatus.OVERPAID]), 'PAID'
                ),
                # 其他情况，使用调整项自身的状态
                else_=FinancialAdjustment.status
            ).label('effective_status')
        ).outerjoin(
            CustomerBill, FinancialAdjustment.customer_bill_id ==CustomerBill.id
        ).filter(
            FinancialAdjustment.adjustment_type == AdjustmentType.DEPOSIT
        ).subquery('deposit_adjustment')

        # 3. 主查询，使用新的子查询
        query = db.session.query(
            BaseContract,
            deposit_adjustment_sq.c.amount.label('deposit_amount'),
            deposit_adjustment_sq.c.effective_status.label('deposit_status')
        ).outerjoin(
            deposit_adjustment_sq, BaseContract.id ==deposit_adjustment_sq.c.contract_id
        ).join( # 统一只通过 service_personnel_id 关联 ServicePersonnel
            ServicePersonnel,
            BaseContract.service_personnel_id == ServicePersonnel.id,
            isouter=True,
        )

        if status and status != "all":
            query = query.filter(BaseContract.status == status)

        if contract_type:
            types_to_filter = [t.strip() for t in contract_type.split(',')]
            query = query.filter(BaseContract.type.in_(types_to_filter))

        if search_term:
            query = query.filter(
                db.or_(
                    BaseContract.customer_name.ilike(f"%{search_term}%"),
                    BaseContract.customer_name_pinyin.ilike(f"%{search_term}%"),
                    ServicePersonnel.name.ilike(f"%{search_term}%"),
                    ServicePersonnel.name_pinyin.ilike(f"%{search_term}%"),
                )
            )

        if employee_id:
            # 统一按 service_personnel_id 筛选
            query = query.filter(BaseContract.service_personnel_id == employee_id)
        if customer_name:
            query = query.filter(
                db.or_(
                    BaseContract.customer_name == customer_name,
                    BaseContract.contact_person == customer_name
                )
            )
        # --- 【新增逻辑】处理定金状态筛选 ---
        if deposit_status_filter == 'paid':
            query = query.filter(deposit_adjustment_sq.c.effective_status =='PAID')
        elif deposit_status_filter == 'unpaid':
            # 未支付包含两种情况：1. 没有定金记录 2. 有定金记录但状态不是PAID
            query = query.filter(
                db.or_(
                    deposit_adjustment_sq.c.effective_status == None,
                    deposit_adjustment_sq.c.effective_status != 'PAID'
                ),
                deposit_adjustment_sq.c.amount > 0
            )

        today = date.today()

        if sort_by == "remaining_days":
            end_date_expr = case(
                (
                    MaternityNurseContract.expected_offboarding_date is not None,
                    MaternityNurseContract.expected_offboarding_date,
                ),
                else_=BaseContract.end_date,
            )
            order_expr = case(
                (NannyContract.is_monthly_auto_renew, date(9999, 12, 31)),
                else_=end_date_expr,
            )
            if sort_order == "desc":
                query = query.order_by(db.desc(order_expr))
            else:
                query = query.order_by(db.asc(order_expr))
        else:
            query = query.order_by(BaseContract.start_date.desc())

        paginated_contracts = query.paginate(
            page=page, per_page=per_page, error_out=False
        )




        results = []
        for contract, deposit_amount, deposit_status in paginated_contracts.items:
            # 统一从 ServicePersonnel 获取员工姓名
            employee_name = (
                contract.service_personnel.name
                if contract.service_personnel
                else "未知员工"
            )

            # 你的原始“剩余有效期”计算逻辑，完全保留
            remaining_months_str = "N/A"
            highlight_remaining = False
            start_date_dt = contract.actual_onboarding_date or contract.start_date
            start_date_for_calc = None
            if isinstance(start_date_dt, datetime):
                start_date_for_calc = start_date_dt.date()
            elif isinstance(start_date_dt, date):
                start_date_for_calc = start_date_dt
            end_date_dt = None
            if contract.type == "maternity_nurse":
                end_date_dt = (
                    contract.expected_offboarding_date or contract.end_date
                )
            else:
                end_date_dt = contract.end_date
            end_date_for_calc = None
            if isinstance(end_date_dt, datetime):
                end_date_for_calc = end_date_dt.date()
            elif isinstance(end_date_dt, date):
                end_date_for_calc = end_date_dt
            today = date.today()
            if isinstance(contract, NannyContract) and getattr(
                contract, "is_monthly_auto_renew", False
            ):
                remaining_months_str = "月签"
            elif start_date_for_calc and end_date_for_calc:
                if start_date_for_calc > today:
                    remaining_months_str = "合同未开始"
                elif end_date_for_calc > today:
                    total_days_remaining = (end_date_for_calc - today).days
                    if contract.type == "nanny" and total_days_remaining < 30:
                        highlight_remaining = True
                    if total_days_remaining >= 365:
                        years = total_days_remaining // 365
                        months = (total_days_remaining % 365) // 30
                        remaining_months_str = f"约{years}年{months}个月"
                    elif total_days_remaining >= 30:
                        months = total_days_remaining // 30
                        remaining_months_str = f"{months}个月"
                    elif total_days_remaining >= 0:
                        remaining_months_str = f"{total_days_remaining}天"
                    else:
                        remaining_months_str = "已结束"
                else:
                    remaining_months_str = "已结束"

        

            final_deposit_amount = deposit_amount if deposit_amount is not None else 0

            # 检查 试工合同 是否可以被标记为“试工成功", 即是否有后续合作的合同
            can_convert = False
            if contract.type == 'nanny_trial':
                can_convert = _check_can_convert(contract)

            
            results.append(
                {
                    "id": str(contract.id),
                    "customer_name": contract.customer_name,
                    "employee_name": employee_name,
                    "contract_type_value": contract.type,
                    "contract_type_label": get_contract_type_details(contract.type),
                    "status": contract.status,
                    "employee_level": contract.employee_level,
                    "start_date": contract.start_date.isoformat() if contract.start_date else None,
                    "end_date": contract.end_date.isoformat() if contract.end_date else None,
                    "actual_onboarding_date": getattr(contract,"actual_onboarding_date", None),
                    "provisional_start_date": getattr(contract,"provisional_start_date", None),
                    "remaining_months": remaining_months_str,
                    "highlight_remaining": highlight_remaining,
                    "deposit_amount": str(final_deposit_amount),
                    "deposit_paid": (deposit_status == 'PAID'),
                    "can_convert_to_formal": can_convert,
                }
            )

        return jsonify(
            {
                "items": results,
                "total": paginated_contracts.total,
                "page": paginated_contracts.page,
                "per_page": paginated_contracts.per_page,
                "pages": paginated_contracts.pages,
            }
        )

    except Exception as e:
        current_app.logger.error(f"获取所有合同列表失败: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

def _check_can_convert(trial_contract):
    """
    检查一个试工合同是否有资格转换为正式合同。
    资格标准：客户名下存在与该试工员工匹配的、状态为'active'的正式育儿嫂合同。
    """
    if not trial_contract or trial_contract.type != 'nanny_trial':
        return False

    # 统一使用 service_personnel_id
    employee_sp_id = trial_contract.service_personnel_id
    if not employee_sp_id:
        return False

    # 构建查询，查找匹配的正式合同
    # exists() 是一个高效的查询，只要找到一个匹配项就会立即返回
    has_eligible_formal_contract = db.session.query(
        NannyContract.query.filter(
            NannyContract.customer_name == trial_contract.customer_name,
            NannyContract.service_personnel_id == employee_sp_id,
            NannyContract.status == 'active'
        ).exists()
    ).scalar()

    return has_eligible_formal_contract

# 你可能需要一个辅助函数，如果还没有的话
def get_contract_type_details(type_string):
    # 这个函数只是一个示例，你需要根据你的实际情况来实现
    type_map = {
        "nanny": "育儿嫂",
        "maternity_nurse": "月嫂",
        "nanny_trial": "育儿嫂试工",
        "external_substitution": "外部替班"
    }
    return type_map.get(type_string, type_string)

@billing_bp.route("/personnel/search", methods=["GET"])
@admin_required
def search_personnel():
    """
    用于前端自动补全，仅搜索服务人员(ServicePersonnel)。
    """
    query_str = request.args.get("q", "").strip()
    if not query_str:
        return jsonify([])

    search_term = f"%{query_str}%"

    # 仅查询 ServicePersonnel 表
    service_personnel = ServicePersonnel.query.filter(
        or_(
            ServicePersonnel.name.ilike(search_term),
            ServicePersonnel.name_pinyin.ilike(search_term)
        )
    ).limit(15).all()

    results = [
        {"id": str(sp.id), "name": sp.name}
        for sp in service_personnel
    ]

    return jsonify(results)


@billing_bp.route("/customers/search", methods=["GET"])
@admin_required
def search_customers():
    """
    用于前端自动补全，搜索已存在的客户名称。
    """
    query_str = request.args.get("q", "").strip()
    if not query_str:
        return jsonify([])

    search_term = f"%{query_str}%"

    # 在合同表中去重查找
    customers_query = db.session.query(BaseContract.customer_name).filter(
        or_(
            BaseContract.customer_name.ilike(search_term),
            BaseContract.customer_name_pinyin.ilike(search_term)
        )
    ).distinct().limit(10)

    results = [item[0] for item in customers_query.all()]

    return jsonify(results)

def _get_or_create_customer(name: str, phone: str = None):
    """
    根据姓名和（可选）手机号，查找或创建客户。
    返回客户ID。
    """
    # 1. 按手机号在 Customer 表中精确查找
    if phone:
        customer = Customer.query.filter_by(phone_number=phone).first()
        if customer:
            return customer.id

    # 2. 按姓名在 Customer 表中查找
    customer = Customer.query.filter_by(name=name).first()
    if customer:
        return customer.id

    # 3. 如果都找不到，则创建新的 Customer
    name_pinyin_full = "".join(lazy_pinyin(name))
    name_pinyin_initials = "".join(item[0] for item in pinyin(name, style=Style.FIRST_LETTER))
    name_pinyin_combined = f"{name_pinyin_full}{name_pinyin_initials}"

    new_customer = Customer(
        name=name,
        phone_number=phone,
        name_pinyin=name_pinyin_combined
    )
    db.session.add(new_customer)
    db.session.flush()
    current_app.logger.info(f"创建了新的客户: {name} (Pinyin: {name_pinyin_combined})")
    return new_customer.id

@billing_bp.route("/contracts/virtual", methods=["POST"])
@admin_required
def create_virtual_contract():
    """手动创建虚拟合同。"""
    data = request.get_json()
    if not data:
        return jsonify({"error": "请求体不能为空"}), 400

    required_fields = ["contract_type", "customer_name", "customer_phone", "employee_name","employee_level", "start_date", "end_date"]
    if not all(field in data for field in required_fields):
        return jsonify({"error": f"缺少必要字段: {', '.join(required_fields)}"}),400

    try:
        employee_ref = _get_or_create_personnel_ref(data["employee_name"])
        customer_id = _get_or_create_customer(data["customer_name"], data.get("customer_phone"))

        customer_name = data["customer_name"]
        cust_pinyin_full = "".join(item[0] for item in pinyin(customer_name,style=Style.NORMAL))
        cust_pinyin_initials = "".join(item[0] for item in pinyin(customer_name,style=Style.FIRST_LETTER))
        customer_name_pinyin = f"{cust_pinyin_full}{cust_pinyin_initials}"
        management_fee_rate=D(data.get("management_fee_rate", 0.10)),

        common_params = {
            "customer_id": customer_id,
            "customer_name": data["customer_name"],
            "customer_name_pinyin": customer_name_pinyin,
            "contact_person": data.get("contact_person"),
            "employee_level": D(data["employee_level"]),
            # 【关键修改】直接解析为 DateTime 对象，不再取 .date()
            "start_date": date_parse(data["start_date"]),
            "end_date": date_parse(data["end_date"]),
            "notes": data.get("notes"),
            "source": "virtual",
            "management_fee_rate": management_fee_rate,
        }
        # 统一将员工ID赋值给 service_personnel_id
        common_params["service_personnel_id"] = employee_ref["id"]
        contract_type = data["contract_type"]
        new_contract = None

        if contract_type == "maternity_nurse":
            new_contract = MaternityNurseContract(
                **common_params,
                status="active",
                # 关键修改：直接使用 date_parse 解析为 datetime 对象
                provisional_start_date=date_parse(data["provisional_start_date"]) if data.get("provisional_start_date") else None,
                security_deposit_paid=D(data.get("security_deposit_paid") or 0),
                management_fee_amount=D(data.get("management_fee_amount") or 0),
                introduction_fee=D(data.get("introduction_fee") or 0),
                deposit_amount=D(data.get("deposit_amount") or 0)
            )
        elif contract_type == "nanny":
            new_contract = NannyContract(
                **common_params,
                status="active",
                is_monthly_auto_renew=data.get("is_monthly_auto_renew", False),
                management_fee_amount=D(data.get("management_fee_amount") or 0),
                introduction_fee=D(data.get("introduction_fee") or 0),
                security_deposit_paid=D(data["employee_level"])
            )
        elif contract_type == "nanny_trial":
            # --- 【核心修改】处理管理费率和介绍费的互斥逻辑 ---
            intro_fee = D(data.get("introduction_fee") or 0)
            rate = D(data.get("management_fee_rate") or 0)

            if intro_fee > 0 and rate > 0:
                return jsonify({"error": "介绍费和管理费率不能同时填写。"}), 400

            new_contract = NannyTrialContract(
                **common_params,
                status="trial_active",
                introduction_fee=intro_fee,
                management_fee_rate=rate
            )
        # 【新增分支】
        elif contract_type == "external_substitution":
            new_contract = ExternalSubstitutionContract(
                **common_params,
                status="active", # 外部替班合同默认为 active
                management_fee_rate=D(data.get("management_fee_rate", 0.20)),
                management_fee_amount=D(data.get("management_fee_amount") or 0),
                security_deposit_paid=D(data["employee_level"]), # 保证金 = 员工级别
            )
        else:
            return jsonify({"error": "无效的合同类型"}), 400

        db.session.add(new_contract)
        # --- BEGIN: 我们新增的核心业务逻辑 ---
        # 先 flush 一次，让 new_contract 对象获得数据库的 ID
        db.session.flush()

        if isinstance(new_contract, MaternityNurseContract):
            # 如果是月嫂合同，调用专属函数处理定金和介绍费
            create_maternity_nurse_contract_adjustments(new_contract)
        # else:
        #     # 对于其他合同（育儿嫂、试工），只处理介绍费
        #     upsert_introduction_fee_adjustment(new_contract)
        # --- END: 新增逻辑结束 ---
        db.session.commit()
        current_app.logger.info(f"成功创建虚拟合同，ID: {new_contract.id}, 类型: {contract_type}")

        # --- 新增：为符合条件的虚拟合同创建薪酬历史 ---
        update_salary_history_on_contract_activation(new_contract)
        db.session.commit()
        # --- 新增结束 ---

        # 根据合同类型，触发正确的后续处理任务
        if contract_type in ["nanny", "external_substitution"]:
            post_virtual_contract_creation_task.delay(str(new_contract.id))
            current_app.logger.info(f"已为合同 {new_contract.id} (类型: {contract_type}) 提交通用的后续处理任务。")
        elif contract_type == "maternity_nurse":
            generate_all_bills_for_contract_task.delay(str(new_contract.id))
            current_app.logger.info(f"已为月嫂合同 {new_contract.id} 提交账单生成任务。")

        return jsonify({
            "message": "虚拟合同创建成功！",
            "contract_id": str(new_contract.id)
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"创建虚拟合同失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误，创建失败"}), 500

# 批量生成月嫂合同账单
@billing_bp.route(
    "/contracts/<string:contract_id>/generate-all-bills", methods=["POST"]
)
@admin_required
def generate_all_bills_for_contract(contract_id):
    current_app.logger.info(
        f"--- [API START] 开始为合同 {contract_id} 批量生成账单 ---"
    )
    contract = db.session.get(MaternityNurseContract, contract_id)
    if (
        not contract
        or not contract.actual_onboarding_date
        or not contract.expected_offboarding_date
    ):
        return jsonify({"error": "合同未找到或缺少必要日期"}), 404

    try:
        engine = BillingEngine()
        cycle_start = contract.actual_onboarding_date
        processed_months = set()
        cycle_count = 0

        current_app.logger.info(
            f"[API] 准备进入循环，预计下户日期为: {contract.expected_offboarding_date}"
        )

        while cycle_start <= contract.expected_offboarding_date:
            cycle_count += 1
            cycle_end = cycle_start + timedelta(days=26)
            if cycle_end > contract.expected_offboarding_date:
                cycle_end = contract.expected_offboarding_date

            settlement_month_key = (cycle_end.year, cycle_end.month)

            if settlement_month_key not in processed_months:
                current_app.logger.info(
                    f"[API] 正在为周期 {cycle_start} ~ {cycle_end} 创建后台计算任务 (结算月: {settlement_month_key[0]}-{settlement_month_key[1]})"
                )
                engine.calculate_for_month(
                    year=settlement_month_key[0],
                    month=settlement_month_key[1],
                    contract_id=str(contract.id),
                    force_recalculate=True,
                )
                processed_months.add(settlement_month_key)

            if cycle_end >= contract.expected_offboarding_date:
                break
            cycle_start = cycle_end + timedelta(days=1)

        db.session.commit()

        newly_generated_bills = CustomerBill.query.filter_by(
            contract_id=contract_id
        ).order_by(CustomerBill.cycle_start_date.asc()).all()

        bills_data = [{
            "id": str(bill.id),
            "billing_period": f"{bill.year}-{str(bill.month).zfill(2)}",
            "cycle_start_date": bill.cycle_start_date.isoformat(),
            "cycle_end_date": bill.cycle_end_date.isoformat(),
            "base_work_days": (bill.calculation_details or {}).get("base_work_days", "0"),
            "overtime_days": (bill.calculation_details or {}).get("overtime_days", "0"),
            "total_due": str(bill.total_due),
            "status": bill.payment_status.value,
        } for bill in newly_generated_bills]

        

        if cycle_end == contract.expected_offboarding_date:
            current_app.logger.info(
                f"[API LOOP-{cycle_count}] 已到达预计下户日期，循环结束。"
            )

        current_app.logger.info(
            f"--- [API END] 所有周期处理完毕，共涉及月份: {processed_months} ---"
        )
        current_app.logger.info(
            f"--- [API END] 所有周期处理完毕，共循环 {cycle_count} 次。---"
        )

        return jsonify({
            "message": f"已为合同 {contract.id} 成功生成 {len(bills_data)} 个账单。",
            "bills": bills_data
        })

    except Exception as e:
        # 这里的 rollback 可能不是必须的，因为引擎内部已经处理了
        current_app.logger.error(
            f"为合同 {contract_id} 批量生成账单时发生顶层错误: {e}", exc_info=True
        )
        return jsonify({"error": "批量生成账单失败"}), 500


@billing_bp.route("/contracts/<string:contract_id>/bills", methods=["GET"])
@admin_required
def get_bills_for_contract(contract_id):
    bills = (
        CustomerBill.query.filter_by(contract_id=contract_id)
        .order_by(CustomerBill.cycle_start_date.asc())
        .all()
    )

    results = []
    engine = BillingEngine()

    # 定义状态映射
    status_map = {
        PaymentStatus.PAID: "已支付",
        PaymentStatus.UNPAID: "未支付",
        PaymentStatus.PARTIALLY_PAID: "部分支付",
        PaymentStatus.OVERPAID: "超额支付",
    }


    for bill in bills:
        calc = bill.calculation_details or {}
        invoice_balance = engine.calculate_invoice_balance(str(bill.id))

        results.append(
            {
                "id": str(bill.id),
                "billing_period": f"{bill.year}-{str(bill.month).zfill(2)}",
                "cycle_start_date": bill.cycle_start_date.isoformat() if bill.cycle_start_date else "N/A",
                "cycle_end_date": bill.cycle_end_date.isoformat() if bill.cycle_end_date else "N/A",
                "total_due": str(bill.total_due),
                "total_paid": str(bill.total_paid),
                "status": status_map.get(bill.payment_status, "未知"), # <-- V2 修改
                "payment_status_label": status_map.get(bill.payment_status, "未知"), # <-- V2 新增
                "customer_is_paid": bill.payment_status == PaymentStatus.PAID, # <-- V2 新增
                "overtime_days": calc.get("overtime_days", "0"),
                "base_work_days": calc.get("base_work_days", "0"),
                "is_substitute_bill": bill.is_substitute_bill,
                "invoice_needed": (bill.payment_details or {}).get("invoice_needed", False),
                "remaining_invoice_amount": str(invoice_balance.get("remaining_un_invoiced", "0.00"))
            }
        )
    return jsonify(results)


# backend/api/billing_api.py (添加新端点)
# 更新月嫂账单周期，后续账单向后顺延
@billing_bp.route("/bills/<string:bill_id>/postpone", methods=["POST"])
@admin_required
def postpone_subsequent_bills(bill_id):
    data = request.get_json()
    new_end_date_str = data.get("new_end_date")
    if not new_end_date_str:
        return jsonify({"error": "缺少新的结束日期 (new_end_date)"}), 400

    try:
        new_end_date = datetime.strptime(new_end_date_str, "%Y-%m-%d").date()

        # 1. 找到当前被修改的账单及其关联的考勤记录
        target_bill = db.session.get(CustomerBill, bill_id)
        if not target_bill:
            return jsonify({"error": "目标账单未找到"}), 404

        target_attendance = AttendanceRecord.query.filter_by(
            contract_id=target_bill.contract_id,
            year=target_bill.year,
            month=target_bill.month,
        ).first()  # 假设通过年月能找到唯一考勤

        if not target_attendance:
            return jsonify({"error": "关联的考勤记录未找到，无法顺延"}), 404

        # 2. 更新当前考勤周期的结束日期
        # original_end_date = target_attendance.cycle_end_date
        target_attendance.cycle_end_date = new_end_date

        # 3. 找出所有后续的考勤记录
        subsequent_attendances = (
            AttendanceRecord.query.filter(
                AttendanceRecord.contract_id == target_bill.contract_id,
                AttendanceRecord.cycle_start_date > target_attendance.cycle_start_date,
            )
            .order_by(AttendanceRecord.cycle_start_date)
            .all()
        )

        # 4. 循环顺延更新后续所有周期
        current_start_date = new_end_date + timedelta(days=1)
        for attendance in subsequent_attendances:
            attendance.cycle_start_date = current_start_date
            attendance.cycle_end_date = current_start_date + timedelta(days=26)

            # 同时更新关联账单的年月
            related_bill = CustomerBill.query.filter_by(
                contract_id=attendance.contract_id,
                # 这里需要一个更可靠的方式来找到关联账单，例如通过考勤ID关联
            ).first()
            if related_bill:
                related_bill.year = attendance.cycle_end_date.year
                related_bill.month = attendance.cycle_end_date.month

            current_start_date = attendance.cycle_end_date + timedelta(days=1)

        db.session.commit()
        return jsonify({"message": "后续所有账单周期已成功顺延更新。"})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"顺延账单失败: {e}", exc_info=True)
        return jsonify({"error": "顺延操作失败"}), 500


# backend/api/billing_api.py (添加新端点)


@billing_bp.route("/bills/<string:bill_id>/update-cycle", methods=["POST"])
@admin_required
def update_bill_cycle_and_cascade(bill_id):
    data = request.get_json()
    new_start_str = data.get("new_start_date")
    new_end_str = data.get("new_end_date")

    if not all([new_start_str, new_end_str]):
        return jsonify({"error": "必须提供新的开始和结束日期"}), 400

    try:
        with db.session.begin():
            new_start_date = datetime.strptime(new_start_str, "%Y-%m-%d").date()
            new_end_date = datetime.strptime(new_end_str, "%Y-%m-%d").date()

            # 1. 找到当前被修改的账单及其关联考勤
            target_bill = db.session.get(CustomerBill, bill_id)
            if not target_bill:
                return jsonify({"error": "目标账单未找到"}), 404

            # **核心修正**: 使用周期重叠的方式查找关联的考勤记录
            year, month = target_bill.year, target_bill.month
            first_day_of_month = date(year, month, 1)
            last_day_of_month = (
                first_day_of_month.replace(day=28) + timedelta(days=4)
            ).replace(day=1) - timedelta(days=1)

            target_attendance = AttendanceRecord.query.filter(
                AttendanceRecord.contract_id == target_bill.contract_id,
                AttendanceRecord.cycle_end_date >= first_day_of_month,
                AttendanceRecord.cycle_start_date <= last_day_of_month,
            ).first()

            if not target_attendance:
                return jsonify({"error": "关联的考勤记录未找到"}), 404

            # 2. 更新当前周期的起止日期
            original_cycle_start = target_attendance.cycle_start_date
            target_attendance.cycle_start_date = new_start_date
            target_attendance.cycle_end_date = new_end_date

            # 3. 找出所有后续的考勤记录
            subsequent_attendances = (
                AttendanceRecord.query.filter(
                    AttendanceRecord.contract_id == target_bill.contract_id,
                    AttendanceRecord.cycle_start_date > original_cycle_start,
                )
                .order_by(AttendanceRecord.cycle_start_date)
                .all()
            )

            # 4. 循环顺延更新后续所有周期
            current_next_start_date = new_end_date + timedelta(days=1)
            for attendance in subsequent_attendances:
                # 获取旧周期的天数
                cycle_duration = (
                    attendance.cycle_end_date - attendance.cycle_start_date
                ).days

                # 设置新的起止日期
                attendance.cycle_start_date = current_next_start_date
                attendance.cycle_end_date = current_next_start_date + timedelta(
                    days=cycle_duration
                )

                # 找到并更新关联账单的年月（这依然是潜在的问题点）
                related_bill = CustomerBill.query.filter(
                    CustomerBill.contract_id == attendance.contract_id,
                    # ...需要一个可靠的方式找到bill
                ).first()
                if related_bill:
                    related_bill.year = attendance.cycle_end_date.year
                    related_bill.month = attendance.cycle_end_date.month

                current_next_start_date = attendance.cycle_end_date + timedelta(days=1)

        return jsonify({"message": "当前周期已更新，且所有后续账单已成功顺延。"})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新并顺延账单失败: {e}", exc_info=True)
        return jsonify({"error": "操作失败"}), 500


@billing_bp.route("/contracts/<string:contract_id>/details", methods=["GET"])
@admin_required
def get_single_contract_details(contract_id):
    current_app.logger.info(f"获取合同详情，ID: {contract_id}")
    try:
        query = db.session.query(with_polymorphic(BaseContract, "*"))
        contract = query.filter(BaseContract.id == contract_id).first()

        if not contract:
            return jsonify({"error": "合同未找到"}), 404
        
        # 【新增逻辑开始】
        # 1. 在获取合同后，单独查询该合同的定金信息
        deposit_info = db.session.query(
            FinancialAdjustment.amount,
            FinancialAdjustment.status,
            FinancialAdjustment.is_settled,
            CustomerBill.payment_status
        ).outerjoin(
            CustomerBill, FinancialAdjustment.customer_bill_id ==CustomerBill.id
        ).filter(
            FinancialAdjustment.contract_id == contract_id,
            FinancialAdjustment.adjustment_type == AdjustmentType.DEPOSIT
        ).first()
        is_monthly_auto_renew = getattr(contract, 'is_monthly_auto_renew', False)
        # 2. 在 Python 中应用我们最终确定的业务逻辑
        final_deposit_amount = 0
        is_deposit_paid = False

        # 优先使用 FinancialAdjustment 中的金额，如果不存在，再尝试从合同自身获取 (兼容旧数据)
        if deposit_info and deposit_info.amount is not None:
            final_deposit_amount = deposit_info.amount
        elif hasattr(contract, 'deposit_amount'):
             final_deposit_amount = contract.deposit_amount or 0

        if deposit_info:
            amount, status, is_settled, bill_payment_status = deposit_info
            current_app.logger.debug(f"定金信息 - 金额: {amount}, 状态: {status}, is_settled: {is_settled}, 账单支付状态: {bill_payment_status}")
            if status == 'PAID':
                is_deposit_paid = True
            elif status == 'BILLED' and is_settled:
                is_deposit_paid = True
            elif bill_payment_status in [PaymentStatus.PAID,PaymentStatus.OVERPAID]:
                is_deposit_paid = True
        # 【新增逻辑结束】

        employee_name = (
            contract.service_personnel.name
            if contract.service_personnel
            else "未知员工"
        )

        remaining_months_str = "N/A"
        highlight_remaining = False
        today = date.today()

        start_date_obj = contract.actual_onboarding_date or contract.start_date
        start_date_for_calc = start_date_obj.date() if isinstance(start_date_obj, datetime) else start_date_obj

        end_date_obj = None
        if contract.type == "maternity_nurse":
            end_date_obj = contract.expected_offboarding_date or contract.end_date
        else:
            end_date_obj = contract.end_date

        end_date_for_calc = end_date_obj.date() if isinstance(end_date_obj, datetime) else end_date_obj

        if isinstance(contract, NannyContract) and getattr(
            contract, "is_monthly_auto_renew", False
        ):
            remaining_months_str = "月签"
        elif start_date_for_calc and end_date_for_calc:
            if start_date_for_calc > today:
                remaining_months_str = "合同未开始"
            elif end_date_for_calc > today:
                total_days_remaining = (end_date_for_calc - today).days
                if contract.type == "nanny" and total_days_remaining < 30:
                    highlight_remaining = True
                if total_days_remaining >= 365:
                    years = total_days_remaining // 365
                    months = (total_days_remaining % 365) // 30
                    remaining_months_str = f"约{years}年{months}个月"
                elif total_days_remaining >= 30:
                    months = total_days_remaining // 30
                    days = total_days_remaining % 30
                    remaining_months_str = f"{months}个月"
                    if days > 0:
                        remaining_months_str += f" {days}天"
                elif total_days_remaining >= 0:
                    remaining_months_str = f"{total_days_remaining}天"
                else:
                    remaining_months_str = "已结束"
            else:
                remaining_months_str = "已结束"

        def safe_isoformat(dt_obj):
            if not dt_obj:
                return None
            if isinstance(dt_obj, datetime):
                return dt_obj.date().isoformat()
            return dt_obj.isoformat()
        
        # 检查 试工合同 是否可以被标记为“试工成功", 即是否有后续合作的合同
        can_convert = False
        if contract.type == 'nanny_trial':
            can_convert = _check_can_convert(contract)

        result = {
            "id": str(contract.id),
            "source_trial_contract_id": str(contract.source_trial_contract_id) if hasattr(contract, 'source_trial_contract_id') and contract.source_trial_contract_id else None,
            'service_personnel_id': str(contract.service_personnel_id) if contract.service_personnel_id else None,
            "customer_name": contract.customer_name,
            "contact_person": contract.contact_person,
            "family_id": contract.family_id,
            "introduction_fee": contract.introduction_fee,
            "deposit_amount": getattr(contract, "deposit_amount", None),
            "management_fee_amount": contract.management_fee_amount,
            "management_fee_rate": getattr(contract, "management_fee_rate", None),
            "employee_name": employee_name,
            "contract_type_value": contract.type,
            "contract_type_label": get_contract_type_details(contract.type),
            "status": contract.status,
            "employee_level": contract.employee_level,
            "start_date": safe_isoformat(contract.start_date),
            "end_date": safe_isoformat(contract.end_date),
            "actual_onboarding_date": safe_isoformat(getattr(contract, "actual_onboarding_date",None)),
            "provisional_start_date": safe_isoformat(getattr(contract, "provisional_start_date",None)),
            "remaining_months": remaining_months_str,
            "highlight_remaining": highlight_remaining,
            "notes": contract.notes,
            "attachment_content": contract.attachment_content,
            "is_monthly_auto_renew":is_monthly_auto_renew,
            # "is_monthly_auto_renew": getattr(contract,'is_monthly_auto_renew', None),
            "deposit_amount": str(final_deposit_amount),
            "deposit_paid": is_deposit_paid,
            "can_convert_to_formal": can_convert,
            "termination_date": safe_isoformat(contract.termination_date),
            "previous_contract_id": str(contract.previous_contract_id) if contract.previous_contract_id else None,
            "successor_contract_id": str(contract.next_contracts[0].id) if contract.next_contracts else None,
            # --- 签署相关字段 ---
            "signing_status": contract.signing_status.value if contract.signing_status else 'UNSIGNED',
            "requires_signature": contract.requires_signature,
        }
        if contract.type == 'nanny_trial':
            result['trial_outcome'] = contract.trial_outcome.value if contract.trial_outcome else None
            # 同时，把转换后的正式合同ID也加上，供前端使用
            if hasattr(contract, 'converted_to_formal_contract') and contract.converted_to_formal_contract:
                result['converted_to_formal_contract_id'] = str (contract.converted_to_formal_contract.id)

        
        return jsonify(result)

    except Exception as e:
        current_app.logger.error(f"获取合同详情 {contract_id} 失败: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@billing_bp.route("/contracts/<uuid:contract_id>/terminate", methods=["POST"])
@jwt_required()
def terminate_contract(contract_id):
    current_app.logger.debug(f"=========开始终止合同(billing_api)========={contract_id}")
    data = request.get_json()
    termination_date_str = data.get("termination_date")
    transfer_options = data.get("transfer_options")
    charge_on_termination_date = data.get("charge_on_termination_date", True)

    if not termination_date_str:
        return jsonify({"error": "Termination date is required"}), 400

    try:
        termination_date = datetime.strptime(termination_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD."}), 400

    contract = db.session.get(BaseContract, str(contract_id))
    if not contract:
        return jsonify({"error": "Contract not found"}), 404

    contract.termination_date = termination_date
    original_end_date = contract.end_date.date() if isinstance(contract.end_date, datetime) else contract.end_date
    contract_start_date = contract.start_date.date() if isinstance(contract.start_date,datetime) else contract.start_date

    if isinstance(contract, NannyTrialContract):
        contract.trial_outcome = TrialOutcome.FAILURE
        current_app.logger.info(f"试工合同 {contract.id} 终止，结果设置为“失败”。")
        try:
            if termination_date < contract_start_date:
                return jsonify({"error": "终止日期不能早于合同开始日期"}), 400
            actual_trial_days = (termination_date - contract_start_date).days
            engine = BillingEngine()
            engine.process_trial_termination(contract, actual_trial_days)
            contract.status = "terminated"
            contract.end_date = termination_date
            db.session.commit()
            return jsonify({"message": "育儿嫂试工合同已成功结算并终止。"})
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"处理试工合同终止失败: {e}", exc_info=True)
            return jsonify({"error": f"处理试工失败结算时发生错误: {e}"}), 500

    try:
        # 步骤 1: 清理未来账单、更新合同状态
        termination_year = termination_date.year
        termination_month = termination_date.month

        bills_to_delete_query = CustomerBill.query.with_entities(CustomerBill.id).filter(
            CustomerBill.contract_id == str(contract.id),
            ((CustomerBill.year == termination_year) & (CustomerBill.month > termination_month))|
            (CustomerBill.year > termination_year)
        )
        bill_ids_to_delete = [item[0] for item in bills_to_delete_query.all()]
        if bill_ids_to_delete:
            FinancialActivityLog.query.filter(FinancialActivityLog.customer_bill_id.in_(bill_ids_to_delete)).delete(synchronize_session=False)
            CustomerBill.query.filter(CustomerBill.id.in_(bill_ids_to_delete)).delete(synchronize_session=False)

        payrolls_to_delete_query = EmployeePayroll.query.with_entities(EmployeePayroll.id).filter(
            EmployeePayroll.contract_id == str(contract_id),
            ((EmployeePayroll.year == termination_year) & (EmployeePayroll.month >termination_month)) |
            (EmployeePayroll.year > termination_year)
        )
        payroll_ids_to_delete = [item[0] for item in payrolls_to_delete_query.all()]
        if payroll_ids_to_delete:
            FinancialActivityLog.query.filter(FinancialActivityLog.employee_payroll_id.in_(payroll_ids_to_delete)).delete(synchronize_session=False)
            EmployeePayroll.query.filter(EmployeePayroll.id.in_(payroll_ids_to_delete)).delete(synchronize_session=False)

        contract.status = "terminated"
        if hasattr(contract, 'expected_offboarding_date'):
            contract.expected_offboarding_date = termination_date

        # 步骤 2: 重算最后一期账单，得到干净的应付总额
        engine = BillingEngine()
        engine.calculate_for_month(
            year=termination_year,
            month=termination_month,
            contract_id=str(contract.id),
            force_recalculate=True,
            end_date_override=termination_date
        )
        db.session.commit()

        # 步骤 3: 基于干净的账单，调用新函数来创建最终的薪资调整项
        final_bill = CustomerBill.query.filter(CustomerBill.contract_id == str(contract.id)).order_by(CustomerBill.cycle_end_date.desc()).first()
        if final_bill:
            engine.create_final_salary_adjustments(final_bill.id)
            db.session.commit()
        else:
            raise ValueError("无法找到用于处理最终调整项的账单。")

        # 步骤 4: 手动计算管理费退款，并添加到最终账单
        if isinstance(contract, NannyContract):
            monthly_management_fee = contract.management_fee_amount if contract.management_fee_amount and contract.management_fee_amount > 0 else (D(contract.employee_level or '0') * D("0.1"))
            if monthly_management_fee > 0 and (contract.is_monthly_auto_renew or termination_date < original_end_date):
                daily_management_fee = (monthly_management_fee / D(30)).quantize(D("0.0001"))
                management_refund_item = None
                if not contract.is_monthly_auto_renew:
                    # 非月签合同的管理费退款逻辑
                    if termination_date.year == original_end_date.year and termination_date.month == original_end_date.month:
                        # 同月终止的情况
                        days_to_refund = (original_end_date - termination_date).days if charge_on_termination_date else (original_end_date - termination_date).days + 1
                        if days_to_refund > 0:
                            amount = (daily_management_fee * D(days_to_refund)).quantize(D("0.01"))
                            if amount > 0:
                                management_refund_item = {'amount': amount, 'description': f"[系统] 单月合同提前终止退款: 日管理费({daily_management_fee:.4f}) * 退款天数({days_to_refund}) = {amount:.2f}元"}
                    else:
                        # 跨月终止的情况 - 使用新的非月签逻辑
                        description_parts = ["合同提前终止，管理费退款计算如下："]
                        total_refund_amount = D(0)
                        
                        # 对于x月11日开始y月11日结束的非月签合同
                        # 按 x月11日~x+1月10日算一个整月
                        contract_start_day = contract_start_date.day
                        contract_end_day = original_end_date.day
                        
                        # 计算从终止日到当前周期结束日的天数
                        if contract_start_day == contract_end_day:
                            # 例如：17日开始17日结束的合同，周期是 X月17日 ~ X+1月16日
                            # 需要找到终止日所在周期的结束日
                            if termination_date.day >= contract_start_day:
                                # 终止日在当前周期内（例如：1月17日~1月31日之间）
                                # 当前周期结束日是下个月的 contract_start_day - 1
                                if termination_date.month == 12:
                                    current_cycle_end = date(termination_date.year + 1, 1, contract_start_day - 1)
                                else:
                                    current_cycle_end = date(termination_date.year, termination_date.month + 1, contract_start_day - 1)
                            else:
                                # 终止日在上个周期的后半段（例如：1月1日~1月16日之间）
                                # 当前周期结束日是本月的 contract_start_day - 1
                                current_cycle_end = date(termination_date.year, termination_date.month, contract_start_day - 1)
                            
                            next_month_10th = current_cycle_end  # 保持变量名兼容性
                            
                            # 用n+1月10日与终止日n月m日之间的差距来计算待退还管理费的天数
                            # 根据是否收取终止日管理费来决定退款起始日期
                            if charge_on_termination_date:
                                # 收取终止日管理费，从终止日的下一天开始退款
                                refund_start_date = termination_date + timedelta(days=1)
                                days_to_next_cycle_end = (next_month_10th - refund_start_date).days + 1
                                description_suffix = f"(收取终止日管理费，从{refund_start_date.month}月{refund_start_date.day}日开始退款)"
                            else:
                                # 不收取终止日管理费，从终止日当天开始退款
                                refund_start_date = termination_date
                                days_to_next_cycle_end = (next_month_10th - refund_start_date).days + 1
                                description_suffix = f"(不收取终止日管理费，从{refund_start_date.month}月{refund_start_date.day}日开始退款)"
                            
                            if days_to_next_cycle_end > 0:
                                partial_refund = (daily_management_fee * D(days_to_next_cycle_end)).quantize(D("0.01"))
                                total_refund_amount += partial_refund
                                description_parts.append(f"  - 退款天数: {days_to_next_cycle_end}天 * {daily_management_fee:.4f} = {partial_refund:.2f}元 {description_suffix}")
                            
                            # 计算剩余的完整周期数
                            remaining_cycles = 0
                            current_cycle_start = next_month_10th + timedelta(days=1)  # 从n+1月11日开始
                            
                            while current_cycle_start < original_end_date:
                                remaining_cycles += 1
                                # 移动到下一个周期开始
                                if current_cycle_start.month == 12:
                                    current_cycle_start = date(current_cycle_start.year + 1, 1, contract_start_day)
                                else:
                                    try:
                                        current_cycle_start = date(current_cycle_start.year, current_cycle_start.month + 1, contract_start_day)
                                    except ValueError:
                                        # 处理月末日期不存在的情况（如31日）
                                        current_cycle_start = date(current_cycle_start.year, current_cycle_start.month + 1, min(contract_start_day, calendar.monthrange(current_cycle_start.year, current_cycle_start.month + 1)[1]))
                            
                            if remaining_cycles > 0:
                                full_cycles_refund = monthly_management_fee * D(remaining_cycles)
                                total_refund_amount += full_cycles_refund
                                description_parts.append(f"  - 剩余完整周期: {remaining_cycles}个 * {monthly_management_fee:.2f} = {full_cycles_refund:.2f}元")
                        else:
                            # 其他情况使用原有逻辑
                            term_month_refund_amount = D(0)
                            _, days_in_month = calendar.monthrange(termination_date.year, termination_date.month)
                            refund_days_term = (days_in_month - termination_date.day) if charge_on_termination_date else (days_in_month - termination_date.day + 1)
                            
                            if refund_days_term > 0:
                                term_month_refund_amount = (D(refund_days_term) * daily_management_fee).quantize(D("0.01"))
                            
                            full_months_count = 0
                            full_months_total_amount = D(0)
                            current_month_start = termination_date.replace(day=1) + relativedelta(months=1)
                            while current_month_start.year < original_end_date.year or (current_month_start.year == original_end_date.year and current_month_start.month < original_end_date.month):
                                full_months_count += 1
                                current_month_start += relativedelta(months=1)
                            if full_months_count > 0:
                                full_months_total_amount = monthly_management_fee * D(full_months_count)
                            
                            original_end_month_refund_amount = D(0)
                            if not (termination_date.year == original_end_date.year and termination_date.month == original_end_date.month):
                                refund_days_original_end = original_end_date.day
                                if refund_days_original_end > 0:
                                    original_end_month_refund_amount = (D(refund_days_original_end) * daily_management_fee).quantize(D("0.01"))
                            
                            total_refund_amount = term_month_refund_amount + full_months_total_amount + original_end_month_refund_amount
                            
                            if term_month_refund_amount > 0:
                                description_parts.append(f"  - 终止月({termination_date.month}月{termination_date.day}日)剩余 {refund_days_term} 天: {term_month_refund_amount:.2f}元")
                            if full_months_count > 0 and full_months_total_amount > 0:
                                description_parts.append(f"  - 完整月份: {full_months_count}个月 * {monthly_management_fee:.2f} = {full_months_total_amount:.2f}元")
                            if original_end_month_refund_amount > 0:
                                description_parts.append(f"  - 原始末月({original_end_date.month}月)部分 {original_end_date.day} 天: {original_end_month_refund_amount:.2f}元")
                        
                        if total_refund_amount > 0:
                            description_parts.append(f"  - 总计：{total_refund_amount:.2f}元")
                            management_refund_item = {'amount': total_refund_amount, 'description': '\n'.join(description_parts)}
                elif contract.is_monthly_auto_renew:
                    daily_fee = (monthly_management_fee / D(30)).quantize(D("0.0001"))

                    _, num_days_in_month = calendar.monthrange(termination_date.year, termination_date.month)
                    original_cycle_end = date(termination_date.year, termination_date.month, num_days_in_month)

                    remaining_days = (original_cycle_end - termination_date).days if charge_on_termination_date else (original_cycle_end - termination_date).days + 1

                    if remaining_days > 0:
                        amount = (daily_fee * D(remaining_days)).quantize(D("0.01"))
                        if amount > 0:
                            management_refund_item = {'amount': amount, 'description': f"[系统] 月签合同终止退管理费: 日管理费({daily_fee:.4f}) * 退款天数({remaining_days}) = {amount:.2f}元"}
                if management_refund_item:
                    existing_refund = FinancialAdjustment.query.filter_by(customer_bill_id=final_bill.id,description=management_refund_item['description']).first()
                    if not existing_refund:
                        db.session.add(FinancialAdjustment(customer_bill_id=final_bill.id,adjustment_type=AdjustmentType.CUSTOMER_DECREASE, amount=management_refund_item['amount'],description=management_refund_item['description'], date=termination_date))

        # 步骤 5: 查找最终账单上所有的退款项（包括引擎生成的和我们刚加的）
        all_refund_items = FinancialAdjustment.query.filter(
            FinancialAdjustment.customer_bill_id == final_bill.id,
            FinancialAdjustment.adjustment_type == AdjustmentType.CUSTOMER_DECREASE
        ).all()

        message = f"合同 {contract.id} 已成功终止。"

        # 步骤 6: 如果是转签，则处理转移和冲抵
        if transfer_options and all_refund_items:
            new_contract_id = transfer_options.get("new_contract_id")
            new_contract = db.session.get(BaseContract, new_contract_id)
            if not new_contract: raise ValueError(f"转签失败：新合同 {new_contract_id} 未找到。")

            first_bill_of_new_contract = CustomerBill.query.filter_by(contract_id=new_contract.id,is_substitute_bill=False).order_by(CustomerBill.cycle_start_date.asc()).first()
            if not first_bill_of_new_contract: raise ValueError(f"新合同 {new_contract.id} 尚未生成账单，无法应用冲抵额度。")

            for item in all_refund_items:
                if item.amount > 0:
                    db.session.add(FinancialAdjustment(
                        customer_bill_id=first_bill_of_new_contract.id,
                        adjustment_type=AdjustmentType.CUSTOMER_DECREASE,
                        amount=item.amount,
                        description=f"[从前合同转入] {item.description}",
                        date=termination_date, status='PENDING',
                        details={'source_contract_id': str(contract.id), 'source_adjustment_id': str(item.id)}
                    ))
                    db.session.add(FinancialAdjustment(
                        customer_bill_id=final_bill.id,
                        adjustment_type=AdjustmentType.CUSTOMER_INCREASE,
                        amount=item.amount,
                        description=f"[冲抵] {item.description} (已转至新合同)",
                        date=termination_date,
                    ))

            substitute_user_id = transfer_options.get("substitute_user_id")
            if substitute_user_id:
                substitute_user = db.session.get(User, substitute_user_id)
                if substitute_user: cancel_substitute_bill_due_to_transfer(db.session, old_contract=contract,new_contract=new_contract, substitute_user=substitute_user,termination_date=termination_date)

            message = f"合同 {contract.id} 已成功终止，并完成费用转签。"
        elif all_refund_items:
             message = f"合同 {contract.id} 已终止，并已在最终账单生成退款项。"

        # 【新增逻辑】检查并生成提前终止合同产生的替班管理费
        for sub_record in contract.substitute_records:
            current_app.logger.debug(f"[API-Terminate] Processing sub_record: {sub_record.id}, start_date: {sub_record.start_date}, end_date: {sub_record.end_date}")
            total_fee = calculate_substitute_management_fee(
                sub_record, contract, contract_termination_date=termination_date
            )
            current_app.logger.debug(f"[API-Terminate] Calculated total_fee for sub_record {sub_record.id}: {total_fee}")
            if total_fee > 0:
                sub_record.substitute_management_fee = total_fee
                db.session.add(sub_record)
                current_app.logger.info(
                    f"因合同 {contract.id} 提前终止，为替班记录 {sub_record.id} 更新了 {total_fee} 元的管理费。"
                )
            else:
                current_app.logger.info(
                    f"因合同 {contract.id} 提前终止，替班记录 {sub_record.id} 未产生管理费 (total_fee: {total_fee})."
                )
        db.session.commit()
        current_app.logger.info(message)
        return jsonify({"message": message}), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error during contract termination for {contract_id}: {e}", exc_info=True)
        return jsonify({"error": "operation_failed", "message": str(e)}), 500





    


@billing_bp.route("/contracts/<uuid:contract_id>/succeed", methods=["POST"])
def succeed_trial_contract(contract_id):
    contract = BaseContract.query.get_or_404(contract_id)

    if contract.type != "nanny_trial":
        return jsonify({"error": "Only trial contracts can succeed."}), 400

    if contract.status != "trial_active":
        return jsonify(
            {
                "error": f"Contract is not in trial_active state, but in {contract.status}."
            }
        ), 400

    contract.status = "trial_succeeded"
    db.session.commit()

    current_app.logger.info(
        f"Trial contract {contract_id} has been marked as 'trial_succeeded'."
    )
    return jsonify({"message": "Trial contract marked as succeeded."})


@billing_bp.route("/bills/find", methods=["GET"])
@admin_required
def find_bill_and_its_page():
    bill_id = request.args.get("bill_id")
    per_page = request.args.get("per_page", 100, type=int)

    if not bill_id:
        return jsonify({"error": "bill_id is required"}), 400

    target_bill = db.session.get(CustomerBill, bill_id)
    if not target_bill:
        return jsonify({"error": "Bill not found"}), 404

    billing_year = target_bill.year
    billing_month = target_bill.month
    # 不要直接用 target_bill.contract，而是用 with_polymorphic 重新查询
    contract_poly = with_polymorphic(BaseContract, "*")
    contract = db.session.query(contract_poly).filter(contract_poly.id == target_bill.contract_id).first()

    if not contract:
        return jsonify({"error": "Associated contract not found"}), 404

    query = (
        CustomerBill.query.join(
            BaseContract, CustomerBill.contract_id == BaseContract.id
        )
        .filter(CustomerBill.year == billing_year, CustomerBill.month == billing_month)
        .order_by(BaseContract.customer_name, CustomerBill.cycle_start_date)
    )

    all_bill_ids_in_month = [str(b.id) for b in query.all()]

    try:
        position = all_bill_ids_in_month.index(bill_id)
        page_number = (position // per_page) + 1 # 页码从1开始
    except ValueError:
        return jsonify(
            {"error": "Bill found but could not determine its position."}
        ), 500

    details = get_billing_details_internal(bill_id=bill_id)

    # --- Modification: Add necessary contract info to the response ---
    employee_name = (
        contract.service_personnel.name
        if contract.service_personnel
        else "未知员工"
    )

    return jsonify(
        {
            "bill_details": details,
            "page": page_number,
            "billing_month": f"{billing_year}-{str(billing_month).zfill(2)}",
            # Add a 'context' object that mimics the list's bill structure
            "context": {
                "id": str(target_bill.id),
                "contract_id": str(contract.id),
                "customer_name": contract.customer_name,
                "employee_name": employee_name,
                "contract_type_value": contract.type,
                "status": contract.status,
                "employee_level": contract.employee_level,
                "is_monthly_auto_renew": getattr(contract, 'is_monthly_auto_renew', False),
            },
        }
    )

@billing_bp.route("/bills/<string:bill_id>/extend", methods=["POST"])
@admin_required
def extend_single_bill(bill_id):
    data = request.get_json()
    new_end_date_str = data.get("new_end_date")
    if not new_end_date_str:
        return jsonify({"error": "必须提供新的结束日期 (new_end_date)"}), 400

    try:
        bill = db.session.get(CustomerBill, bill_id)
        if not bill:
            return jsonify({"error": "账单未找到"}), 404

        new_end_date = date_parse(new_end_date_str).date()
        contract = bill.contract

        # --- 核心修复：确保所有变更和计算都在一个事务内，并使用正确的逻辑 ---

        # 1. 更新合同的权威结束日期
        #    我们统一使用 expected_offboarding_date 来存储延长后的日期
        contract.expected_offboarding_date = new_end_date

        # 2. 更新当前账单的结束日期以匹配
        bill.cycle_end_date = new_end_date

        db.session.add(contract)
        db.session.add(bill)

        # 3. 调用计算引擎，并把新的结束日期作为最高优先级传入
        engine = BillingEngine()
        engine.calculate_for_month(
            year=bill.year,
            month=bill.month,
            contract_id=bill.contract_id,
            force_recalculate=True,
            cycle_start_date_override=bill.cycle_start_date,
            end_date_override=new_end_date
        )

        # 4. 提交所有更改（日期变更和计算结果）
        db.session.commit()

        # 5. 返回最新的、正确的数据
        latest_details = get_billing_details_internal(bill_id=bill.id)

        # current_app.logger.info(f"账单 {bill_id} 已成功延latest_details长至 {latest_details}")
        return jsonify(latest_details)

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"延长账单 {bill_id} 失败: {e}", exc_info=True)
        return jsonify({"error": f"延长服务失败: {str(e)}"}), 500

@billing_bp.route('/customer-bills/<uuid:bill_id>/defer', methods=['POST'])
@admin_required
def defer_customer_bill(bill_id):
    """
    将指定账单中未支付的客户应付款，顺延到下一个月的账单。
    """
    bill_to_defer = db.session.get(CustomerBill, str(bill_id))
    if not bill_to_defer:
        return jsonify({"error": "要顺延的账单未找到"}), 404

    # 关键修复 1: 使用新的 payment_status 字段来判断
    if bill_to_defer.payment_status == PaymentStatus.PAID:
        return jsonify({"error": "已结清的账单不能顺延"}), 400

    # 关键修复 2: 检查是否已存在顺延记录，防止重复操作
    existing_deferral = FinancialAdjustment.query.filter_by(
        description=f"[系统添加] 从账单 {bill_to_defer.id} 顺延"
    ).first()
    if existing_deferral:
        return jsonify({"error": "此账单已被顺延，不能重复操作"}), 400

    # 找到下一个月的账单
    next_month_date = (date(bill_to_defer.year, bill_to_defer.month, 1) + relativedelta(months=1))
    next_bill = CustomerBill.query.filter_by(
        contract_id=bill_to_defer.contract_id,
        year=next_month_date.year,
        month=next_month_date.month,
        is_substitute_bill=False
    ).first()

    if not next_bill:
        return jsonify({"error": "未找到下一个月的账单可供顺延"}), 404

    # 计算待付金额
    amount_to_defer = bill_to_defer.total_due - bill_to_defer.total_paid
    if amount_to_defer <= 0:
        return jsonify({"error": "没有待付金额可以顺延"}), 400

    try:
        # 在下个账单中创建一个“上期顺延费用”的增款项
        deferred_fee_adjustment = FinancialAdjustment(
            customer_bill_id=next_bill.id,
            adjustment_type=AdjustmentType.DEFERRED_FEE,
            amount=amount_to_defer,
            description=f"[系统添加] 从 {bill_to_defer.month} 月账单顺延到本月",
            date=next_bill.cycle_start_date
        )
        db.session.add(deferred_fee_adjustment)

        # 在当前账单中创建一个等额的减款项，使其结清
        offsetting_adjustment = FinancialAdjustment(
            customer_bill_id=bill_to_defer.id,
            adjustment_type=AdjustmentType.CUSTOMER_DECREASE,
            amount=amount_to_defer,
            description=f"[系统添加] 顺延至{next_bill.month}月账单",
            date=bill_to_defer.cycle_end_date
        )
        db.session.add(offsetting_adjustment)

        # 重新计算两个账单
        engine = BillingEngine()
        engine.calculate_for_month(year=bill_to_defer.year, month=bill_to_defer.month,contract_id=bill_to_defer.contract_id, force_recalculate=True)
        engine.calculate_for_month(year=next_bill.year, month=next_bill.month,contract_id=next_bill.contract_id, force_recalculate=True)

        db.session.commit()
        return jsonify({"message": "费用已成功顺延到下个账单"})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"顺延账单 {bill_id} 失败: {e}", exc_info=True)
        return jsonify({"error": "顺延操作失败"}), 500
    
@billing_bp.route("/batch-settle", methods=["POST"])
@admin_required
def batch_settle():
    """
    批量更新账单和薪酬单的支付/发放状态。
    """
    data = request.get_json()
    updates = data.get("updates")
    if not updates:
        return jsonify({"error": "缺少更新数据"}), 400

    try:
        # 【关键修复】：在函数开头获取当前操作用户的ID
        user_id = get_jwt_identity()
        updated_count = 0

        for update_data in updates:
            bill_id = update_data.get("bill_id")
            customer_is_paid = update_data.get("customer_is_paid")
            employee_is_paid = update_data.get("employee_is_paid")

            bill = db.session.get(CustomerBill, bill_id)
            if not bill:
                continue

            # --- 客户账单结算 ---
            if customer_is_paid is not None:
                current_customer_paid_status = (bill.payment_status == PaymentStatus.PAID)
                if current_customer_paid_status != customer_is_paid:
                    if customer_is_paid:
                        pending_amount = bill.total_due - bill.total_paid
                        if pending_amount > 0:
                             db.session.add(PaymentRecord(
                                customer_bill_id=bill.id,
                                amount=pending_amount,
                                payment_date=date.today(),
                                method='batch_settle',
                                notes='[系统] 批量结算抹平差额',
                                created_by_user_id=user_id  # <--- 使用获取到的用户ID
                            ))
                    else:
                        PaymentRecord.query.filter_by(
                            customer_bill_id=bill.id,
                            method='batch_settle'
                        ).delete(synchronize_session=False)

                    _update_bill_payment_status(bill)
                    updated_count += 1

            # --- 员工薪酬结算 ---
            payroll = EmployeePayroll.query.filter_by(
                contract_id=bill.contract_id,
                cycle_start_date=bill.cycle_start_date
            ).first()

            if payroll and employee_is_paid is not None:
                current_employee_paid_status = (payroll.payout_status == PayoutStatus.PAID)
                if current_employee_paid_status != employee_is_paid:
                    if employee_is_paid:
                        pending_amount = payroll.total_due - payroll.total_paid_out
                        if pending_amount > 0:
                            db.session.add(PayoutRecord(
                                employee_payroll_id=payroll.id,
                                amount=pending_amount,
                                payout_date=date.today(),
                                method='batch_settle',
                                notes='[系统] 批量结算抹平差额',
                                created_by_user_id=user_id  # <--- 在这里也使用获取到的用户ID
                            ))
                    else:
                        PayoutRecord.query.filter_by(
                            employee_payroll_id=payroll.id,
                            method='batch_settle'
                        ).delete(synchronize_session=False)

                    _update_payroll_payout_status(payroll)
                    updated_count += 1

        if updated_count > 0:
            db.session.commit()
            return jsonify({"message": f"成功更新 {updated_count} 条记录的结算状态。"})
        else:
            return jsonify({"message": "没有记录的状态需要更新。"})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"批量结算失败: {e}", exc_info=True)
        return jsonify({"error": "批量结算时发生服务器内部错误"}), 500

@billing_bp.route("/dashboard/summary", methods=["GET"])
@admin_required
def get_dashboard_summary():
    """
    获取用于仪表盘显示的核心业务摘要数据。
    """
    try:
        today = date.today()

        # === 1. KPIs ===
        active_contracts_count = db.session.query(func.count(BaseContract.id)).filter(BaseContract.status == 'active').scalar() or 0
        active_employees_count =db.session.query(func.count(distinct(BaseContract.service_personnel_id))).filter(
            BaseContract.status == 'active',
            BaseContract.service_personnel_id.isnot(None)
        ).scalar() or 0

        pending_deposit_query = db.session.query(
            func.count(distinct(FinancialAdjustment.contract_id))
        ).join(
            BaseContract, FinancialAdjustment.contract_id == BaseContract.id
        ).outerjoin(
            CustomerBill, FinancialAdjustment.customer_bill_id ==CustomerBill.id
        ).filter(
            BaseContract.status == 'active',
            BaseContract.type.in_(['nanny', 'maternity_nurse']),
            FinancialAdjustment.adjustment_type == AdjustmentType.DEPOSIT,
            FinancialAdjustment.amount > 0,
            not_(or_(
                FinancialAdjustment.status == 'PAID',
                and_(FinancialAdjustment.status == 'BILLED',FinancialAdjustment.is_settled == True),
                CustomerBill.payment_status.in_([PaymentStatus.PAID,PaymentStatus.OVERPAID])
            ))
        )
        pending_deposit_count = pending_deposit_query.scalar() or 0

        yearly_management_fee_received = db.session.query(
            func.sum(case((CustomerBill.payment_status == PaymentStatus.PAID,CustomerBill.calculation_details['management_fee'].as_float()), else_=0))
        ).filter(
            CustomerBill.year == today.year
        ).scalar() or 0

        yearly_management_fee_total = db.session.query(
            func.sum(CustomerBill.calculation_details['management_fee'].as_float())
        ).filter(
            CustomerBill.year == today.year
        ).scalar() or 0

        kpis = {
            "monthly_management_fee_received": f"{D(yearly_management_fee_received).quantize(D('0.01'))}",
            "monthly_management_fee_total": f"{D(yearly_management_fee_total).quantize(D('0.01'))}",
            "active_contracts_count": active_contracts_count,
            "active_employees_count": active_employees_count,
            "pending_deposit_count": pending_deposit_count,
        }

        # === 2. TODO Lists ===
        two_weeks_later = today + timedelta(weeks=2)
        expiring_contracts_query = BaseContract.query.filter(
            BaseContract.status == 'active',
            func.date(BaseContract.end_date).between(today, today +timedelta(days=30))
        ).order_by(BaseContract.end_date.asc()).limit(5)

        expiring_contracts = [{
            "id": str(c.id),
            "customer_name": c.customer_name,
            "employee_name": (c.service_personnel.name if c.service_personnel else '未知'),
            "end_date": (c.end_date.date() if isinstance(c.end_date, datetime)else c.end_date).isoformat(),
            "expires_in_days": ((c.end_date.date() if isinstance(c.end_date,datetime) else c.end_date) - today).days
        } for c in expiring_contracts_query.all()]

        upcoming_maternity_contracts_query = MaternityNurseContract.query.filter(
            MaternityNurseContract.status == 'active',
            MaternityNurseContract.provisional_start_date.isnot(None),
func.date(MaternityNurseContract.provisional_start_date).between(today, today+ timedelta(weeks=2))
        ).order_by(MaternityNurseContract.provisional_start_date.asc()).limit(5)

        upcoming_maternity_contracts = [{
            "id": str(c.id),
            "customer_name": c.customer_name,
            "provisional_start_date": (c.provisional_start_date.date() if isinstance(c.provisional_start_date, datetime) else c.provisional_start_date).isoformat(),
            "days_until": ((c.provisional_start_date.date() if isinstance(c.provisional_start_date, datetime) else c.provisional_start_date) -today).days
        } for c in upcoming_maternity_contracts_query.all()]

        pending_payments_query = CustomerBill.query.filter(
            CustomerBill.year == today.year,
            CustomerBill.month == today.month,
            CustomerBill.payment_status.in_([PaymentStatus.UNPAID,PaymentStatus.PARTIALLY_PAID]),
            CustomerBill.total_due > CustomerBill.total_paid
        ).order_by(CustomerBill.total_due.desc()).limit(5)

        pending_payments = [{
            "bill_id": str(p.id),
            "customer_name": p.contract.customer_name,
            "contract_type": get_contract_type_details(p.contract.type),
            "amount": f"{(p.total_due - p.total_paid).quantize(D('0.01'))}"
        } for p in pending_payments_query.all()]

        todo_lists = {
            "expiring_contracts": expiring_contracts,
            "approaching_provisional": upcoming_maternity_contracts,
            "pending_payments": pending_payments
        }

        # === 3. 应收款构成 (Receivables Summary for This Year) (V2 - 修正版) ===
        this_year = today.year

        receivables_summary = {
            "management_fee": D(0),
            "introduction_fee": D(0),
            "employee_commission": D(0),
            "other_receivables": D(0)
        }

        customer_linked_adjustments = db.session.query(
            FinancialAdjustment.adjustment_type,
            func.sum(FinancialAdjustment.amount)
        ).join(
            CustomerBill, FinancialAdjustment.customer_bill_id ==CustomerBill.id
        ).filter(
            CustomerBill.year == this_year,
            FinancialAdjustment.adjustment_type.in_([
                AdjustmentType.INTRODUCTION_FEE,
                AdjustmentType.CUSTOMER_INCREASE
            ])
        ).group_by(FinancialAdjustment.adjustment_type).all()

        for adj_type, total_amount in customer_linked_adjustments:
            if adj_type == AdjustmentType.INTRODUCTION_FEE:
                receivables_summary["introduction_fee"] = total_amount or D(0)
            elif adj_type == AdjustmentType.CUSTOMER_INCREASE:
                receivables_summary["other_receivables"] = total_amount or D(0)

        employee_commission_total = db.session.query(
            func.sum(FinancialAdjustment.amount)
        ).join(
            EmployeePayroll, FinancialAdjustment.employee_payroll_id ==EmployeePayroll.id
        ).filter(
            EmployeePayroll.year == this_year,
            FinancialAdjustment.adjustment_type ==AdjustmentType.EMPLOYEE_COMMISSION
        ).scalar()
        receivables_summary["employee_commission"] =D(employee_commission_total or 0)

        total_management_fee = db.session.query(
            func.sum(CustomerBill.calculation_details['management_fee'].as_float())
        ).filter(
            CustomerBill.year == this_year,
            CustomerBill.calculation_details.has_key('management_fee')
        ).scalar() or 0
        receivables_summary["management_fee"] = D(total_management_fee)

        final_receivables_summary = {
            "management_fee": str(receivables_summary["management_fee"].quantize(D('0.01'))),
            "introduction_fee": str(receivables_summary["introduction_fee"].quantize(D('0.01'))),
            "employee_first_month_fee": str(receivables_summary["employee_commission"].quantize(D('0.01'))),
            "other_receivables": str(receivables_summary["other_receivables"].quantize(D('0.01')))
        }

        # === 4. Charts ===
        last_12_months_dates = [today - relativedelta(months=i) for i in range(12)]

        due_revenue_data = db.session.query(
            CustomerBill.year,
            CustomerBill.month,
            func.sum(CustomerBill.calculation_details['management_fee'].as_float())
        ).filter(
            func.date_trunc('month',func.to_date(func.concat(cast(CustomerBill.year, String), '-',cast(CustomerBill.month, String)), 'YYYY-MM')) >= func.date_trunc('month',today - relativedelta(months=11))
        ).group_by(CustomerBill.year, CustomerBill.month).all()
        due_revenue_by_month = {f"{r.year}-{r.month}": float(r[2] or 0) for r in due_revenue_data}

        paid_revenue_data = db.session.query(
            CustomerBill.year,
            CustomerBill.month,
            func.sum(case((CustomerBill.payment_status == PaymentStatus.PAID,CustomerBill.calculation_details['management_fee'].as_float()), else_=0))
        ).filter(
            func.date_trunc('month',func.to_date(func.concat(cast(CustomerBill.year, String), '-',cast(CustomerBill.month, String)), 'YYYY-MM')) >= func.date_trunc('month',today - relativedelta(months=11))
        ).group_by(CustomerBill.year, CustomerBill.month).all()
        paid_revenue_by_month = {f"{r.year}-{r.month}": float(r[2] or 0) for r in paid_revenue_data}

        revenue_trend = {
            "categories": [],
            "series": [{"name": "应收管理费", "data": []}],
            "paid_data": []
        }
        for dt in sorted(last_12_months_dates, key=lambda d: (d.year,d.month)):
            month_key = f"{dt.year}-{dt.month}"
            revenue_trend["categories"].append(f"{dt.month}月")
            revenue_trend["series"][0]["data"].append(due_revenue_by_month.get(month_key, 0))
            revenue_trend["paid_data"].append(paid_revenue_by_month.get(month_key, 0))

        distribution_query = db.session.query(
            BaseContract.type,
            func.sum(CustomerBill.calculation_details['management_fee'].as_float())
        ).join(CustomerBill).filter(
            CustomerBill.year == today.year
        ).group_by(BaseContract.type).all()

        management_fee_distribution = {
            "this_year": {"labels": [], "series": []},
            "last_12_months": {"labels": [], "series": []}
        }
        for contract_type, total_fee in distribution_query:
            management_fee_distribution["this_year"]["labels"].append(get_contract_type_details(contract_type))
            management_fee_distribution["this_year"]["series"].append(float(total_fee or 0))

        # === Final Assembly ===
        summary = {
            "kpis": kpis,
            "todo_lists": todo_lists,
            "revenue_trend": revenue_trend,
            "management_fee_distribution": management_fee_distribution,
            "receivables_summary": final_receivables_summary
        }
        return jsonify(summary)

    except Exception as e:
        current_app.logger.error(f"获取仪表盘数据失败: {e}", exc_info=True)
        return jsonify({"error": "获取仪表盘数据时发生服务器内部错误"}), 500

@billing_bp.route("/export-management-fees", methods=["GET"])
@admin_required
def export_management_fees():
    """
    导出管理费明细为 Excel 文件 (最终排序修正版)。
    """
    try:
        # 1. 获取筛选参数 (逻辑不变)
        search_term = request.args.get("search", "").strip()
        contract_type = request.args.get("type", "")
        status = request.args.get("status", "")
        billing_month_str = request.args.get("billing_month")
        payment_status_filter = request.args.get("payment_status", "")
        payout_status_filter = request.args.get("payout_status", "")

        if not billing_month_str:
            return jsonify({"error": "必须提供账单月份 (billing_month) 参数"}), 400

        billing_year, billing_month = map(int, billing_month_str.split("-"))

        # 2. 构建查询 (逻辑不变, 已包含拼音搜索)
        contract_poly = with_polymorphic(BaseContract, "*")
        query = (
            db.session.query(CustomerBill, contract_poly)
            .select_from(CustomerBill)
            .join(contract_poly, CustomerBill.contract_id == contract_poly.id)
            .outerjoin(User, contract_poly.user_id == User.id)
            .outerjoin(ServicePersonnel, contract_poly.service_personnel_id ==ServicePersonnel.id)
        )
        query = query.filter(
            CustomerBill.year == billing_year, CustomerBill.month ==billing_month,
            CustomerBill.calculation_details.has_key('management_fee')
        )
        if status:
            query = query.filter(contract_poly.status == status)
        if contract_type:
            query = query.filter(contract_poly.type == contract_type)
        if search_term:
            query = query.filter(
                db.or_(
                    contract_poly.customer_name.ilike(f"%{search_term}%"),
                    contract_poly.customer_name_pinyin.ilike(f"%{search_term}%"),
                    User.username.ilike(f"%{search_term}%"),
                    User.name_pinyin.ilike(f"%{search_term}%"),
                    ServicePersonnel.name.ilike(f"%{search_term}%"),
                    ServicePersonnel.name_pinyin.ilike(f"%{search_term}%"),
                )
            )
        if payment_status_filter:
            query = query.filter(CustomerBill.payment_status ==PaymentStatus(payment_status_filter))
        if payout_status_filter:
            query = query.join(EmployeePayroll,and_(EmployeePayroll.contract_id == CustomerBill.contract_id,EmployeePayroll.cycle_start_date == CustomerBill.cycle_start_date))
            query = query.filter(EmployeePayroll.payout_status ==PayoutStatus(payout_status_filter))

        bills_to_export = query.all()

        # --- 核心重构：使用字典列表来存储数据 ---
        data_rows = []
        for bill, contract in bills_to_export:
            original_management_fee = D(bill.calculation_details.get('management_fee', 0))

            employee_payable = D(0)
            payroll = EmployeePayroll.query.filter_by(
                contract_id=bill.contract_id,
                cycle_start_date=bill.cycle_start_date,
                is_substitute_payroll=bill.is_substitute_bill
            ).first()
            if payroll:
                payable_adjustments = FinancialAdjustment.query.filter(
                    FinancialAdjustment.employee_payroll_id == payroll.id,
                    FinancialAdjustment.adjustment_type.in_([
                        AdjustmentType.EMPLOYEE_DECREASE,
                        AdjustmentType.EMPLOYEE_COMMISSION
                    ])
                ).all()
                if payable_adjustments:
                    employee_payable = sum(adj.amount for adj in payable_adjustments)

            total_management_fee = original_management_fee + employee_payable

            employee_name = ""
            if bill.is_substitute_bill:
                sub_record = bill.source_substitute_record
                if sub_record:
                    sub_employee = sub_record.substitute_user or sub_record.substitute_personnel
                    employee_name = getattr(sub_employee, "username", getattr(sub_employee, "name", "未知替班员工"))
                else:
                    employee_name = "替班(记录丢失)"
            else:
                original_employee = contract.service_personnel
                employee_name = getattr(original_employee, "name", "未知员工")

            # 将计算结果存入字典
            data_rows.append({
                "customer_name": contract.customer_name,
                "employee_name": employee_name,
                "contract_type": get_contract_type_details(contract.type),
                "original_fee": float(original_management_fee),
                "employee_payable": float(employee_payable),
                "total_fee": float(total_management_fee)
            })

        # --- 核心重构：按字典的 'total_fee' 键进行排序 ---
        data_rows.sort(key=lambda x: x['total_fee'], reverse=True)

        # 3. 创建 Excel 工作簿并写入数据
        wb = Workbook()
        ws = wb.active
        ws.title = f"{billing_month_str} 管理费明细"

        headers = ["客户姓名", "服务人员", "合同类型", "原始管理费","员工应缴款", "合计管理费"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # --- 核心重构：从字典中按顺序提取值并写入 ---
        for row_dict in data_rows:
            ws.append([
                row_dict['customer_name'],
                row_dict['employee_name'],
                row_dict['contract_type'],
                row_dict['original_fee'],
                row_dict['employee_payable'],
                row_dict['total_fee']
            ])

        # 4. 保存到内存并发送 (逻辑不变)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{billing_month_str}_management_fees_v3.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        current_app.logger.error(f"导出管理费失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误，导出失败"}), 500
    
@billing_bp.route("/conflicts", methods=['GET'])
@admin_required
def get_billing_conflicts():
    """
    检测指定月份的合同冲突。
    """
    billing_month_str = request.args.get("billing_month")
    if not billing_month_str:
        return jsonify({"error": "必须提供 billing_month 参数 (格式: YYYY-MM)"}), 400

    try:
        year, month = map(int, billing_month_str.split('-'))
    except ValueError:
        return jsonify({"error": "无效的 billing_month 格式"}), 400

    # 1. 获取目标月份的所有有效账单
    bills = db.session.query(
        CustomerBill
    ).join(BaseContract, CustomerBill.contract_id == BaseContract.id).filter(
        CustomerBill.year == year,
        CustomerBill.month == month,
        BaseContract.status == 'active'
    ).all()

    # 2. 员工冲突检测
    employee_bills = defaultdict(list)
    for bill in bills:
        # 确保合同关联了员工
        if bill.contract and bill.contract.service_personnel_id:
            employee_id = bill.contract.service_personnel_id
            employee_bills[employee_id].append(bill)

    employee_conflicts = []
    processed_employee_pairs = set()

    for employee_id, bill_list in employee_bills.items():
        if len(bill_list) > 1:
            sorted_bills = sorted(bill_list, key=lambda b: b.cycle_start_date)
            for i in range(len(sorted_bills)):
                for j in range(i + 1, len(sorted_bills)):
                    bill_a = sorted_bills[i]
                    bill_b = sorted_bills[j]

                    # 创建一个唯一的键来标识这对账单
                    pair_key = tuple(sorted( (str(bill_a.id), str(bill_b.id)) ))
                    if pair_key in processed_employee_pairs:
                        continue

                    if bill_a.cycle_start_date <= bill_b.cycle_end_date and bill_a.cycle_end_date >= bill_b.cycle_start_date:
                        employee = bill_a.contract.service_personnel
                        if not employee: continue # 如果找不到员工信息，则跳过

                        # 查找是否已存在该员工的冲突记录
                        existing_conflict = next((c for c in employee_conflicts if c[ 'identifier_id'] == employee_id), None)

                        formatted_bill_a = format_bill_for_conflict_response(bill_a)
                        formatted_bill_b = format_bill_for_conflict_response(bill_b)

                        if existing_conflict:
                            # 如果已存在，将新的冲突账单添加进去
                            if formatted_bill_a not in existing_conflict['conflicts']:
                                existing_conflict['conflicts'].append(formatted_bill_a)
                            if formatted_bill_b not in existing_conflict['conflicts']:
                                existing_conflict['conflicts'].append(formatted_bill_b)
                        else:
                            # 否则，创建新的冲突记录
                            employee_conflicts.append({
                                "type": "employee",
                                "identifier_id": employee_id,
                                "identifier_name": employee.name if employee else "N/A",
                                "conflicts": [formatted_bill_a, formatted_bill_b]
                            })

                        processed_employee_pairs.add(pair_key)

    # 3. 客户冲突检测
    customer_bills = defaultdict(list)
    for bill in bills:
        if bill.customer_name:
            customer_bills[bill.customer_name].append(bill)

    customer_conflicts = []
    processed_customer_pairs = set()

    for customer_name, bill_list in customer_bills.items():
        if len(bill_list) > 1:
            sorted_bills = sorted(bill_list, key=lambda b: b.cycle_start_date)
            for i in range(len(sorted_bills)):
                for j in range(i + 1, len(sorted_bills)):
                    bill_a = sorted_bills[i]
                    bill_b = sorted_bills[j]

                    pair_key = tuple(sorted((str(bill_a.id), str(bill_b.id))))
                    if pair_key in processed_customer_pairs:
                        continue

                    if bill_a.cycle_start_date <= bill_b.cycle_end_date and bill_a.cycle_end_date >= bill_b.cycle_start_date:
                        existing_conflict = next((c for c in customer_conflicts if c['identifier_name'] == customer_name),None)

                        formatted_bill_a = format_bill_for_conflict_response(bill_a)
                        formatted_bill_b = format_bill_for_conflict_response(bill_b)

                        if existing_conflict:
                            if formatted_bill_a not in existing_conflict['conflicts']:
                                existing_conflict['conflicts'].append(formatted_bill_a)
                            if formatted_bill_b not in existing_conflict['conflicts']:
                                existing_conflict['conflicts'].append(formatted_bill_b)
                        else:
                            customer_conflicts.append({
                                "type": "customer",
                                "identifier_name": customer_name,
                                "conflicts": [formatted_bill_a, formatted_bill_b]
                            })

                        processed_customer_pairs.add(pair_key)

    return jsonify({
        "employee_conflicts": employee_conflicts,
        "customer_conflicts": customer_conflicts
    })

def format_bill_for_conflict_response(bill):
    """
    辅助函数，用于格式化冲突检测接口中的账单信息。
    """
    employee = None
    contract_start_date = None
    contract_end_date = None
    management_fee = None

    if bill.contract:
        employee = bill.contract.service_personnel
        # 如果不是替班账单，则获取合同的起止日期
        if not bill.is_substitute_bill:
            contract_start_date = bill.contract.start_date.isoformat() if bill.contract.start_date else None
            contract_end_date = bill.contract.end_date.isoformat() if bill.contract.end_date else None

    # 从 calculation_details 中获取管理费
    if bill.calculation_details and 'management_fee' in bill.calculation_details:
        management_fee = bill.calculation_details['management_fee']

    return {
        "bill_id": str(bill.id),
        "contract_id": str(bill.contract.id),
        "customer_name": bill.contract.customer_name,
        "employee_name": employee.name if employee else "N/A",
        "contract_type": get_contract_type_details(bill.contract.type) if bill.contract else "N/A",
        "cycle_start_date": bill.cycle_start_date.isoformat(),
        "cycle_end_date": bill.cycle_end_date.isoformat(),
        "is_substitute_bill": bill.is_substitute_bill,
        "contract_start_date": contract_start_date,
        "contract_end_date": contract_end_date,
        "total_due": str(bill.total_due or 0),
        "management_fee": str(management_fee or 0)
    }

@billing_bp.route("/financial-adjustments/<uuid:adjustment_id>/transfer", methods=["POST"])
@admin_required
def transfer_financial_adjustment(adjustment_id):
    """
    将一个财务调整项转移到下一个账单周期。
    V2: 增加自动寻找续约合同的逻辑。
    """
    data = request.get_json()
    destination_contract_id = data.get("destination_contract_id")

    try:
        # 1. 数据校验
        source_adj = db.session.get(FinancialAdjustment, adjustment_id)
        if not source_adj:
            return jsonify({"error": "需要转移的财务调整项未找到"}), 404

        if source_adj.details and source_adj.details.get('status') in ['transferred_out', 'offsetting_transfer']:
            return jsonify({"error": "该款项是转出或冲账条目，无法再次转移"}), 400

        source_bill_or_payroll = None
        is_employee_adj = False
        if source_adj.customer_bill_id:
            source_bill_or_payroll = db.session.get(CustomerBill, source_adj.customer_bill_id)
            source_contract = source_bill_or_payroll.contract
        elif source_adj.employee_payroll_id:
            source_bill_or_payroll = db.session.get(EmployeePayroll, source_adj.employee_payroll_id)
            source_contract = source_bill_or_payroll.contract
            is_employee_adj = True
        else:
            # 如果调整项没有关联账单或薪酬单，但有关联合同，则可能是待处理的调整项
            if source_adj.contract_id:
                source_contract = db.session.get(BaseContract, source_adj.contract_id)
                # 对于待处理项，我们需要一个“虚拟”的账单周期来比较，就用合同开始日期
                source_bill_or_payroll = source_contract 
            else:
                return jsonify({"error": "源财务调整项未关联到任何合同、账单或薪酬单"}), 500

        if not source_bill_or_payroll:
             return jsonify({"error": "找不到源财务调整项所属的上下文"}), 500

        # 2. 智能确定目标账单
        dest_bill = None
        dest_contract = None

        # 策略一：查找同一合同的下一期账单
        # 使用 source_contract 的开始日期或账单周期开始日期作为比较基准
        base_date = getattr(source_bill_or_payroll, 'cycle_start_date', source_contract.start_date)
        next_bill_in_contract = CustomerBill.query.filter(
            CustomerBill.contract_id == source_contract.id,
            CustomerBill.cycle_start_date > base_date,
            CustomerBill.is_substitute_bill == False
        ).order_by(CustomerBill.cycle_start_date.asc()).first()

        if next_bill_in_contract:
            dest_bill = next_bill_in_contract
            dest_contract = source_contract
            current_app.logger.info(f"策略1: 找到同一合同的下一期账单 {dest_bill.id} 作为转移目标。")
        else:
            # 策略二：自动寻找续约合同
            successor_contract = _find_successor_contract_internal(str(source_contract.id))
            if successor_contract:
                dest_contract = successor_contract
                current_app.logger.info(f"策略2: 自动找到续约合同 {dest_contract.id}。")
            # 策略三：使用前端手动指定的合同
            elif destination_contract_id:
                dest_contract = db.session.get(BaseContract, destination_contract_id)
                if not dest_contract:
                    return jsonify({"error": "指定的目标合同未找到"}), 404
                current_app.logger.info(f"策略3: 使用前端指定的目标合同 {dest_contract.id}。")

        # 如果通过策略2或3找到了目标合同，现在查找它的第一个账单
        if dest_contract and not dest_bill:
            # V2: 支持基于 family_id 的跨客户转移
            is_same_customer = dest_contract.customer_name == source_contract.customer_name
            is_same_family = (
                source_contract.family_id 
                and dest_contract.family_id 
                and source_contract.family_id == dest_contract.family_id
            )
            
            if not is_same_customer and not is_same_family:
                return jsonify({"error": "只能在同一客户或同一家庭下的不同合同间转移"}), 400

            dest_bill = CustomerBill.query.filter(
                CustomerBill.contract_id == dest_contract.id,
                CustomerBill.is_substitute_bill == False
            ).order_by(CustomerBill.cycle_start_date.asc()).first()

        # 最终检查：如果所有策略都失败了
        if not dest_bill:
            return jsonify({"error": "未指定可供转移的目标合同"}), 400

        # 3. 执行转移和冲抵操作
        with db.session.begin_nested():
            original_description = source_adj.description

            # 3a. 在目标账单下创建“转入”项
            original_type_label = ADJUSTMENT_TYPE_LABELS.get(source_adj.adjustment_type, source_adj.adjustment_type.name)

            source_customer_bill_id = None
            if is_employee_adj:
                source_customer_bill = CustomerBill.query.filter_by(
                    contract_id=source_contract.id,
                    cycle_start_date=source_bill_or_payroll.cycle_start_date,
                    is_substitute_bill=source_bill_or_payroll.is_substitute_payroll
                ).with_entities(CustomerBill.id).first()
                if source_customer_bill:
                    source_customer_bill_id = str(source_customer_bill.id)
            else:
                source_customer_bill_id = str(source_bill_or_payroll.id) if hasattr(source_bill_or_payroll, 'id') else None

            new_adj_in = FinancialAdjustment(
                adjustment_type=source_adj.adjustment_type,
                amount=source_adj.amount,
                description=f"[转入] {original_type_label}",
                date=dest_bill.cycle_start_date,
                details={
                    "status": "transferred_in",
                    "transferred_from_adjustment_id": str(source_adj.id),
                    "transferred_from_bill_id": source_customer_bill_id,
                    "transferred_from_bill_info": f"{getattr(source_bill_or_payroll, 'year', 'N/A')}-{getattr(source_bill_or_payroll, 'month', 'N/A')}"
                }
            )

            dest_payroll = None
            if is_employee_adj:
                dest_payroll = EmployeePayroll.query.filter_by(
                    contract_id=dest_bill.contract_id,
                    cycle_start_date=dest_bill.cycle_start_date
                ).first()
                if not dest_payroll:
                    raise Exception("目标账单没有对应的薪酬单，无法转移员工调整项")
                new_adj_in.employee_payroll_id = dest_payroll.id
            else:
                new_adj_in.customer_bill_id = dest_bill.id

            db.session.add(new_adj_in)
            db.session.flush()

            # 3b. 在源位置创建冲抵项
            offsetting_type = None
            source_type = source_adj.adjustment_type

            if source_type == AdjustmentType.EMPLOYEE_COMMISSION:
                offsetting_type = AdjustmentType.EMPLOYEE_COMMISSION_OFFSET
            elif source_type in [AdjustmentType.CUSTOMER_INCREASE, AdjustmentType.EMPLOYEE_DECREASE]:
                offsetting_type = AdjustmentType.CUSTOMER_DECREASE if not is_employee_adj else AdjustmentType.EMPLOYEE_INCREASE
            elif source_type in [AdjustmentType.CUSTOMER_DECREASE, AdjustmentType.EMPLOYEE_INCREASE]:
                offsetting_type = AdjustmentType.CUSTOMER_INCREASE if not is_employee_adj else AdjustmentType.EMPLOYEE_DECREASE
            else:
                offsetting_type = AdjustmentType.CUSTOMER_DECREASE if not is_employee_adj else AdjustmentType.EMPLOYEE_INCREASE

            offsetting_adj = FinancialAdjustment(
                adjustment_type=offsetting_type,
                amount=source_adj.amount,
                description=f"[冲账] {original_description}",
                date=getattr(source_bill_or_payroll, 'cycle_end_date', source_contract.end_date),
                details={
                    "status": "offsetting_transfer",
                    "offset_for_adjustment_id": str(source_adj.id),
                    "linked_adjustment_id": str(new_adj_in.id),
                    "transferred_to_bill_id": str(dest_bill.id),
                    "transferred_to_bill_info": f"{dest_bill.year}-{dest_bill.month}"
                }
            )
            if is_employee_adj:
                offsetting_adj.employee_payroll_id = source_bill_or_payroll.id
            elif source_adj.customer_bill_id: # 只有当源是账单时才关联
                offsetting_adj.customer_bill_id = source_bill_or_payroll.id
            else: # 如果源是合同，冲抵项也关联到合同
                offsetting_adj.contract_id = source_contract.id


            db.session.add(offsetting_adj)

            # 3c. 更新源调整项
            source_adj.details = {
                "status": "transferred_out",
                "transferred_to_contract_id": str(dest_contract.id),
                "transferred_to_bill_id": str(dest_bill.id),
                "transferred_to_bill_info": f"{dest_bill.year}-{dest_bill.month}",
                "offsetting_adjustment_id": str(offsetting_adj.id)
            }
            source_adj.description = f"{original_description} [已转移]"
            attributes.flag_modified(source_adj, "details")

            # 3d. 记录日志
            log_to_customer = dest_contract.customer_name
            _log_activity(source_bill_or_payroll if not is_employee_adj else None, source_bill_or_payroll if is_employee_adj else None, "执行款项转移(转出)", details={"amount": str(source_adj.amount), "description": original_description, "to_customer": log_to_customer})
            _log_activity(dest_bill if not is_employee_adj else None, dest_payroll, "接收转移款项(转入)", details={"amount": str(new_adj_in.amount), "description": original_description, "from_customer": source_contract.customer_name})

        db.session.commit()

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"款项转移操作失败: {e}", exc_info=True)
        return jsonify({"error": f"服务器内部错误，转移操作失败: {e}"}), 500

    # 4. 触发重算
    recalculation_error = None
    try:
        engine = BillingEngine()
        if hasattr(source_bill_or_payroll, 'year'):
            engine.calculate_for_month(year=source_bill_or_payroll.year, month=source_bill_or_payroll.month, contract_id=source_contract.id, force_recalculate=True, cycle_start_date_override=source_bill_or_payroll.cycle_start_date)
        
        engine.calculate_for_month(year=dest_bill.year, month=dest_bill.month, contract_id=dest_contract.id, force_recalculate=True, cycle_start_date_override=dest_bill.cycle_start_date)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"款项转移后重算账单失败: {e}", exc_info=True)
        recalculation_error = f"自动重算失败: {e}"
    
    # 5. 获取并返回更新后的源账单完整信息
    refresh_bill_id = None
    if not is_employee_adj and hasattr(source_bill_or_payroll, 'id'):
        refresh_bill_id = source_bill_or_payroll.id
    else:
        source_customer_bill = CustomerBill.query.filter_by(
            contract_id=source_contract.id,
            cycle_start_date=getattr(source_bill_or_payroll, 'cycle_start_date', source_contract.start_date),
            is_substitute_bill=getattr(source_bill_or_payroll, 'is_substitute_payroll', False)
        ).first()
        if source_customer_bill:
            refresh_bill_id = source_customer_bill.id

    latest_details = get_billing_details_internal(bill_id=refresh_bill_id) if refresh_bill_id else None

    response_message = "款项转移成功。"
    if recalculation_error:
        response_message += f" 但{recalculation_error}，请尝试手动重算。"
    else:
        response_message += " 账单已自动重算。"

    return jsonify({
        "message": response_message,
        "latest_details": latest_details
    })
    





@billing_bp.route("/payable-details/<item_id>", methods=["GET"])
@admin_required
def get_payable_details(item_id):
    item_type = request.args.get("item_type")
    if not item_type:
        return jsonify({"error": "item_type query parameter is required"}), 400

    service = BankStatementService()
    details = service.get_payable_details(item_id, item_type)

    if details is None:
        return jsonify({"error": "Details not found for the given item"}), 404

    return jsonify(details)


@billing_bp.route("/bills/<string:bill_id>/payments", methods=["POST", "OPTIONS"])
@admin_required
def add_payment_record(bill_id):
    if request.method == 'OPTIONS':
        return jsonify({'status': 'ok'}), 200

    bill = db.session.get(CustomerBill, bill_id)
    if not bill:
        return jsonify({"error": "账单未找到"}), 404

    data = request.form
    if not data or not data.get('amount') or not data.get('payment_date'):
        return jsonify({"error": "必须提供支付金额和支付日期"}), 400


    # ==================== 新的代码 ====================
    db_image_path = None
    full_url_for_frontend = None

    # 检查 'image' 是否存在，并且用户确实上传了文件（文件名不为空）
    if 'image' in request.files and request.files['image'].filename != '':
        image_file = request.files['image']

        # 调用新函数，它会返回两个值
        db_path, full_url = _handle_image_upload(image_file)

        # 如果上传和处理成功
        if db_path:
            db_image_path = db_path
            full_url_for_frontend = full_url
    # ==================== 替换结束 ====================

    try:
        # ==================== 第一个修改点 ====================
        # 创建 PaymentRecord 对象时，使用 db_image_path
        new_payment = PaymentRecord(
            customer_bill_id=bill.id,
            amount=D(data['amount']),
            payment_date=date_parse(data['payment_date']).date(),
            method=data.get('method'),
            notes=data.get('notes'),
            image_url=db_image_path,  # <-- 使用 db_image_path
            created_by_user_id=get_jwt_identity()
        )
        db.session.add(new_payment)
        bill.total_paid += new_payment.amount
        _update_bill_payment_status(bill)

        log_action = f"新增了支付记录，金额: {new_payment.amount:.2f}"
        _log_activity(bill, None, log_action, details=dict(data))

        db.session.commit()

        # ==================== 在这里添加调试日志 ====================
        current_app.logger.info("================ DEBUGGING URL GENERATION ================")

        # 检查从配置中读取的值
        backend_host_from_config = current_app.config.get('BACKEND_BASE_URL', '!!! CONFIG KEY NOT FOUND !!!')
        current_app.logger.info(f"1. 从 app.config 读取的 BACKEND_BASE_URL: '{backend_host_from_config}'")

        # 检查 url_for 生成的路径部分
        # 我们假设 full_url_for_frontend 变量已经被正确设置
        if full_url_for_frontend:
            path_part = full_url_for_frontend.replace(backend_host_from_config, "")
            current_app.logger.info(f"2. 推断出的 URL 路径部分: '{path_part}'")
        else:
            current_app.logger.info("2. full_url_for_frontend 变量为空，无法推断路径！")

        # 检查最终要返回给前端的完整 URL
        current_app.logger.info(f"3. 最终准备在 JSON 中返回的 file_url: '{full_url_for_frontend}'")

        current_app.logger.info("========================================================")
        # ========================== 日志代码结束 ==========================

        # ==================== 第二个修改点 ====================
        # 在返回的 JSON 中，把完整的 URL (full_url_for_frontend) 发给前端
        return jsonify({
            "message": "支付记录添加成功",
            "file_url": full_url_for_frontend
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"添加支付记录失败 (bill_id: {bill_id}): {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500
    


@billing_bp.route("/payrolls/<string:payroll_id>/payouts", methods=["POST", "OPTIONS"])
@admin_required
def add_payout_record(payroll_id):
    if request.method == 'OPTIONS':
        return jsonify({'status': 'ok'}), 200

    payroll = db.session.get(EmployeePayroll, payroll_id)
    if not payroll:
        return jsonify({"error": "薪酬单未找到"}), 404

    data = request.form
    if not data or not data.get('amount') or not data.get('payout_date'):
        return jsonify({"error": "必须提供支付金额和支付日期"}), 400

    db_image_path = None
    full_url_for_frontend = None

    if 'image' in request.files and request.files['image'].filename != '':
        image_file = request.files['image']
        db_path, full_url = _handle_image_upload(image_file)
        if db_path:
            db_image_path = db_path
            full_url_for_frontend = full_url

    try:
        new_payout = PayoutRecord(
            employee_payroll_id=payroll.id,
            amount=D(data['amount']),
            payout_date=date_parse(data['payout_date']).date(),
            method=data.get('method'),
            notes=data.get('notes'),
            payer=data.get('payer'),
            image_url=db_image_path,
            created_by_user_id=get_jwt_identity()
        )
        db.session.add(new_payout)

        # --- 核心逻辑：处理公司代付 ---
        if new_payout.payer == '公司代付':
            customer_bill = CustomerBill.query.filter_by(
                contract_id=payroll.contract_id,
                cycle_start_date=payroll.cycle_start_date,
                is_substitute_bill=payroll.is_substitute_payroll
            ).first()

            if customer_bill:
                # --- 优化描述文本 ---
                employee = payroll.contract.service_personnel
                employee_name = employee.name if employee else '未知员工'
                payroll_month = payroll.month
                description = f"[系统] 公司代付员工:{employee_name} {payroll_month}月工资"
                # --- 优化结束 ---

                company_paid_adj = FinancialAdjustment(
                    customer_bill_id=customer_bill.id,
                    adjustment_type=AdjustmentType.COMPANY_PAID_SALARY,
                    amount=new_payout.amount,
                    description=description, # 使用新的描述
                    date=new_payout.payout_date
                )
                db.session.add(company_paid_adj)
                _log_activity(customer_bill, payroll, "系统自动创建客户增款", details={
                    "reason": "公司代付工资",
                    "amount": str(new_payout.amount),
                    "description": description
                })
        # --- 核心逻辑结束 ---

        _update_payroll_payout_status(payroll)

        log_action = f"新增了工资发放记录，金额: {new_payout.amount:.2f}"
        _log_activity(None, payroll, log_action, details=dict(data))

        db.session.commit()

        return jsonify({
            "message": "工资发放记录添加成功",
            "file_url": full_url_for_frontend
        }), 201

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"添加工资发放记录失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500

@billing_bp.route("/payments/<uuid:payment_id>", methods=["DELETE", "OPTIONS"])
@admin_required
def delete_payment_record(payment_id):
    if request.method == 'OPTIONS':
        return jsonify({'status': 'ok'}), 200

    user_id = get_jwt_identity()
    service = BankStatementService()
    result = service.delete_payment_record_and_reverse_allocation(str(payment_id), user_id)

    if "error" in result:
        # 根据错误类型返回不同的状态码
        if "not found" in result["error"].lower():
            return jsonify(result), 404
        else:
            return jsonify(result), 500
            
    return jsonify(result), 200

@billing_bp.route("/payouts/<uuid:payout_id>", methods=["DELETE", "OPTIONS"])
@admin_required
def delete_payout_record(payout_id):
    if request.method == 'OPTIONS':
        return jsonify({'status': 'ok'}), 200

    payout_record = db.session.get(PayoutRecord, payout_id)
    if not payout_record:
        return jsonify({"error": "工资发放记录未找到"}), 404

    payroll = payout_record.employee_payroll
    log_details = payout_record.to_dict()
    amount_to_revert = payout_record.amount

    try:
        # --- NEW: Reverse bank transaction allocation ---
        if payout_record.notes and '[银行流水分配:' in payout_record.notes:
            match = re.search(r'\[银行流水分配: (\S+)\]', payout_record.notes)
            if match:
                bank_txn_id = match.group(1)
                bank_txn = BankTransaction.query.filter_by(transaction_id=bank_txn_id).first()
                if bank_txn:
                    bank_txn.allocated_amount -= amount_to_revert
                    if bank_txn.allocated_amount <= 0:
                        bank_txn.status = BankTransactionStatus.UNMATCHED
                        bank_txn.allocated_amount = Decimal('0')
                    else:
                        bank_txn.status = BankTransactionStatus.PARTIALLY_ALLOCATED
                    db.session.add(bank_txn)
                    current_app.logger.info(f"Reverted {amount_to_revert} from BankTransaction {bank_txn.id}")

        db.session.delete(payout_record)
        _update_payroll_payout_status(payroll)
        log_action = f"删除了工资发放记录，金额: {log_details['amount']}"
        _log_activity(None, payroll, log_action, details=log_details)
        db.session.commit()
        return jsonify({"message": "工资发放记录删除成功"}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"删除工资发放记录失败 (id: {payout_id}): {e}",exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500
    
def _allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ==================== 新版本的文件处理函数 ====================
def _handle_image_upload(file_storage):
    """
    从 FileStorage 对象保存图片。
    返回两个值:
    1. db_path: 用于存入数据库的相对路径 (e.g., 'financial_records/...')
    2. full_url: 用于返回给前端的完整可访问 URL
    """
    if not file_storage or file_storage.filename == '':
        return None, None

    if _allowed_file(file_storage.filename):
        filename = secure_filename(file_storage.filename)
        # 我们只需要唯一文件名部分，用于URL生成和存储
        unique_filename = str(uuid.uuid4()) + "_" + filename

        # 构造物理保存路径
        upload_folder = os.path.join(current_app.instance_path, 'uploads', 'financial_records')
        os.makedirs(upload_folder, exist_ok=True)
        file_path = os.path.join(upload_folder, unique_filename)

        # 保存文件
        file_storage.save(file_path)

        # --- 这是核心修改点 ---

        # 1. 准备存入数据库的相对路径
        # 路径分隔符应该用 '/'，以保证跨平台兼容性
        db_path = f"financial_records/{unique_filename}"

        # 2. 为前端生成完整的、可访问的 URL
        # 'billing_api.serve_financial_record_upload' 是你上一步创建的那个路由函数名
        full_url = url_for(
            'billing_api.serve_financial_record_upload',
            filename=unique_filename,
            _external=True
        )

        # 返回两个值
        return db_path, full_url

    return None, None
# ========================== 函数结束 ==========================

@billing_bp.route("/contracts/<string:contract_id>/enable-auto-renew", methods=["POST", "OPTIONS"])
@admin_required
def enable_auto_renewal(contract_id):
    """
    为指定的育儿嫂合同开启自动续签，并触发账单延展。
    """
    # 使用 with_for_update 锁定合同记录，防止并发操作
    contract = db.session.query(NannyContract).filter_by(id=str(contract_id)).with_for_update().first()

    if not contract:
        return jsonify({"error": "育儿嫂合同未找到"}), 404

    if contract.status != 'active':
        return jsonify({"error": f"只有'服务中'的合同才能开启自动续签，当前状态为: {contract.status}"}), 400

    if contract.is_monthly_auto_renew:
        return jsonify({"message": "该合同已处于自动续签状态，无需重复操作。"}), 200

    try:
        # --- 关键修复：先设置为月签，再删除退款调整项 ---
        # 这样重算账单时不会重新添加退款项
        contract.is_monthly_auto_renew = True
        db.session.flush()  # 确保这个更改对后续查询可见
        current_app.logger.info(f"合同 {contract.id} 已被设置为自动续约（先设置，防止重算时重新添加退款项）。")

        # --- 删除最后一个月账单中的退款相关调整项 ---
        last_bill = db.session.query(CustomerBill).filter(
            CustomerBill.contract_id == str(contract_id)
        ).order_by(CustomerBill.cycle_end_date.desc()).first()

        if last_bill:
            current_app.logger.info(f"为合同 {contract.id} 查找最后一个账单: {last_bill.id}")
            
            # 收集需要删除的调整项
            adjustments_to_delete = []
            
            # 1. 查找保证金退款调整项
            deposit_refund_adj = FinancialAdjustment.query.filter(
                FinancialAdjustment.customer_bill_id == last_bill.id,
                FinancialAdjustment.adjustment_type == AdjustmentType.CUSTOMER_DECREASE,
                FinancialAdjustment.description.like('%保证金%')
            ).first()
            
            if deposit_refund_adj:
                adjustments_to_delete.append(deposit_refund_adj)
                current_app.logger.info(f"找到保证金退款调整项: {deposit_refund_adj.id}")
            
            # 2. 查找公司代付工资调整项
            company_paid_salary_adj = FinancialAdjustment.query.filter(
                FinancialAdjustment.customer_bill_id == last_bill.id,
                FinancialAdjustment.adjustment_type == AdjustmentType.COMPANY_PAID_SALARY,
                FinancialAdjustment.description == "[系统] 公司代付工资"
            ).first()
            
            if company_paid_salary_adj:
                adjustments_to_delete.append(company_paid_salary_adj)
                current_app.logger.info(f"找到公司代付工资调整项: {company_paid_salary_adj.id}")
                
                # 3. 如果存在镜像调整项（保证金支付工资），也一并删除
                if company_paid_salary_adj.mirrored_adjustment:
                    adjustments_to_delete.append(company_paid_salary_adj.mirrored_adjustment)
                    current_app.logger.info(f"找到镜像调整项（保证金支付工资）: {company_paid_salary_adj.mirrored_adjustment.id}")
            
            # 执行删除
            if adjustments_to_delete:
                for adj in adjustments_to_delete:
                    current_app.logger.info(f"删除调整项: {adj.id} ({adj.description})")
                    db.session.delete(adj)
            
            # --- 修复临界月的周期和补收管理费 ---
            # 获取原合同结束日期和该月最后一天
            contract_end_date = contract.end_date
            if isinstance(contract_end_date, datetime):
                contract_end_date = contract_end_date.date()
            
            # 计算该月的最后一天
            _, last_day_num = calendar.monthrange(contract_end_date.year, contract_end_date.month)
            last_day_of_month = date(contract_end_date.year, contract_end_date.month, last_day_num)
            
            # 1. 修复账单的 cycle_end_date（从原合同结束日改为月末）
            if last_bill.cycle_end_date != last_day_of_month:
                original_cycle_end = last_bill.cycle_end_date
                last_bill.cycle_end_date = last_day_of_month
                current_app.logger.info(f"修复临界月账单周期：{original_cycle_end} → {last_day_of_month}")
            
            # 2. 创建管理费调整项（抵消临界月已收部分）
            # 如果原合同不是月底结束，需要调整管理费
            if contract_end_date < last_day_of_month:
                days_already_charged = contract_end_date.day  # 原合同已收的天数（如7月4日 → 4天）
                max_charge_days = 30  # 月签合同每月最多收30天
                
                # 账单计算引擎会按整月（30天）收取管理费
                # 我们需要减去原合同已收的部分（如4天）
                # 最终客户应付：30天 - 4天 = 26天
                
                monthly_mgmt_fee = D(contract.management_fee_amount or 0)
                if monthly_mgmt_fee > 0 and days_already_charged > 0:
                    daily_rate = (monthly_mgmt_fee / D(30)).quantize(D("0.01"))
                    adjustment_amount = (daily_rate * D(days_already_charged)).quantize(D("0.01"))
                    
                    if adjustment_amount > 0:
                        # 创建减少管理费的调整项（抵消已收部分）
                        description = f"[系统] 转月签管理费调整: 原合同已收{days_already_charged}天管理费，本月减免 {adjustment_amount:.2f}元"
                        
                        # 检查是否已存在相同的调整项（避免重复）
                        existing_adj = FinancialAdjustment.query.filter(
                            FinancialAdjustment.customer_bill_id == last_bill.id,
                            FinancialAdjustment.description.like('%转月签管理费调整%')
                        ).first()
                        
                        if not existing_adj:
                            db.session.add(FinancialAdjustment(
                                customer_bill_id=last_bill.id,
                                adjustment_type=AdjustmentType.CUSTOMER_DECREASE,  # 减少客户应付
                                amount=adjustment_amount,
                                description=description,
                                date=last_day_of_month
                            ))
                            # 计算实际补收天数用于日志
                            actual_charge_days = max_charge_days - days_already_charged
                            current_app.logger.info(
                                f"创建转月签管理费调整项: 原合同已收{days_already_charged}天，"
                                f"本月应收{max_charge_days}天，实际补收{actual_charge_days}天，"
                                f"减免金额{adjustment_amount:.2f}元"
                            )
                        else:
                            current_app.logger.info(f"转月签管理费调整项已存在，跳过创建")
            
            # 3. 重新计算最后一个月账单（此时 is_monthly_auto_renew = True，不会重新添加退款项）
            current_app.logger.info(f"开始重算临界月账单 {last_bill.id}...")
            recalc_engine = BillingEngine()
            recalc_engine.calculate_for_month(
                year=last_bill.year,
                month=last_bill.month,
                contract_id=last_bill.contract_id,
                force_recalculate=True,
                cycle_start_date_override=last_bill.cycle_start_date
            )
            current_app.logger.info(f"临界月账单 {last_bill.id} 已重算完成。")
        # --- 逻辑结束 ---

        # 2. 调用引擎延展账单
        engine = BillingEngine()
        engine.extend_auto_renew_bills(contract.id)
        current_app.logger.info(f"已为合同 {contract.id} 触发账单自动延展。")

        # 3. 提交事务
        db.session.commit()

        return jsonify({"message": "合同已成功设置为自动续签，并已延展账单。"})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"开启合同 {contract_id} 自动续签失败: {e}", exc_info=True)
        return jsonify({"error": "处理失败，服务器内部错误"}), 500

@billing_bp.route("/contracts/<uuid:contract_id>/adjustments", methods=["GET"])
@admin_required
def get_contract_adjustments(contract_id):
    """
    获取单个合同关联的所有财务调整项。
    """
    try:
        # --- 核心修复：使用两次查询然后合并，避免复杂的JOIN错误 ---

        # 1. 查找所有直接关联到合同的调整项 (例如，待处理的定金)
        direct_adjustments =FinancialAdjustment.query.filter_by(contract_id=contract_id).all()

        # 2. 查找所有通过该合同的账单关联的调整项
        bill_adjustments = FinancialAdjustment.query.join(
            CustomerBill, FinancialAdjustment.customer_bill_id == CustomerBill.id
        ).filter(
            CustomerBill.contract_id == contract_id
        ).all()

        # 3. 合并并去重
        all_adjustments_map = {adj.id: adj for adj in direct_adjustments}
        for adj in bill_adjustments:
            all_adjustments_map[adj.id] = adj

        # 4. 按日期排序
        sorted_adjustments = sorted(
            all_adjustments_map.values(),
            key=lambda adj: adj.date or datetime.now().date(),
            reverse=True
        )

        return jsonify([adj.to_dict() for adj in sorted_adjustments])
        # --- 修复结束 ---

    except Exception as e:
        current_app.logger.error(f"获取合同 {contract_id} 的财务调整项失败: {e}",exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500

@billing_bp.route("/financial-adjustments/<uuid:adjustment_id>/record-payment",methods=["POST"])
@admin_required
def record_adjustment_payment(adjustment_id):
    data = request.get_json()
    paid_amount = data.get("paid_amount")
    paid_at_str = data.get("paid_at")
    settlement_notes = data.get("settlement_notes", "定金收款") # <-- 接收前端传来的备注

    if not paid_amount or not paid_at_str:
        return jsonify({"error": "必须提供支付金额和支付日期"}), 400

    try:
        adjustment = db.session.get(FinancialAdjustment, str(adjustment_id))
        if not adjustment:
            return jsonify({"error": "财务调整项未找到"}), 404

        if adjustment.status != 'PENDING':
            return jsonify({"error": f"该调整项状态为 {adjustment.status}，无法记录支付。"}), 400

        if adjustment.adjustment_type != AdjustmentType.DEPOSIT:
             return jsonify({"error": "此接口仅用于记录定金支付。"}), 400

        paid_at_date = date_parse(paid_at_str)

        # --- 核心修改：一次性更新所有相关状态 ---
        # 1. 更新支付信息
        adjustment.paid_amount = D(paid_amount)
        adjustment.paid_at = paid_at_date
        adjustment.paid_by_user_id = get_jwt_identity()
        adjustment.status = 'PAID'

        # 2. 同步更新结算信息
        adjustment.is_settled = True
        adjustment.settlement_date = paid_at_date.date()
        adjustment.settlement_details = { "notes": settlement_notes, "method":"定金" }
        # --- 修改结束 ---

        contract = db.session.get(BaseContract, adjustment.contract_id)
        log_details = {
            "adjustment_id": str(adjustment.id),
            "paid_amount": str(adjustment.paid_amount),
            "paid_at": adjustment.paid_at.strftime('%Y-%m-%d'),
            "settlement_notes": settlement_notes
        }
        _log_activity(bill=None, payroll=None, contract=contract, action="记录定金支付", details=log_details)

        db.session.commit()
        return jsonify({"message": "定金支付记录成功。", "adjustment":adjustment.to_dict()})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"记录定金支付失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500

@billing_bp.route("/contracts/<uuid:contract_id>/logs", methods=["GET"])
@admin_required
def get_contract_logs(contract_id):
    """
    获取单个合同关联的所有活动日志。
    """
    try:
        logs = FinancialActivityLog.query.filter_by(
            contract_id=contract_id
        ).order_by(FinancialActivityLog.created_at.desc()).all()

        results = [
            {
                "id": str(log.id),
                "user": log.user.username if log.user else "未知用户",
                "action": log.action,
                "details": log.details,
                "created_at": log.created_at.isoformat(),
            }
            for log in logs
        ]
        return jsonify(results)
    except Exception as e:
        current_app.logger.error(f"获取合同 {contract_id} 的活动日志失败: {e}",exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500

@billing_bp.route("/bills/<string:bill_id>/set-invoice-needed", methods=["POST"])
@admin_required
def set_invoice_needed_status(bill_id):
    """
    一个专门用于更新“是否需要开票”状态的轻量级API。
    """
    data = request.get_json()
    if 'invoice_needed' not in data:
        return jsonify({"error": "请求体中缺少 'invoice_needed' 字段"}), 400

    bill = db.session.get(CustomerBill, bill_id)
    if not bill:
        return jsonify({"error": "账单未找到"}), 404

    try:
        bill.invoice_needed = bool(data['invoice_needed'])
        db.session.commit()

        # 状态更新后，重新计算发票余额信息
        engine = BillingEngine()
        invoice_balance = engine.calculate_invoice_balance(bill_id)

        # 将最新的状态和计算结果返回给前端
        return jsonify({
            "message": "发票状态更新成功。",
            "invoice_needed": bill.invoice_needed,
            "invoice_balance": invoice_balance
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"更新发票状态失败 (bill_id: {bill_id}): {e}",exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500
    
@billing_bp.route("/export-receivables", methods=["GET"])
@admin_required
def export_receivables():
    """
    导出应收款项总览为 Excel 文件 (最终排序修正版)。
    """
    try:
        # 1. 获取筛选参数 (逻辑不变)
        search_term = request.args.get("search", "").strip()
        contract_type = request.args.get("type", "")
        status = request.args.get("status", "")
        billing_month_str = request.args.get("billing_month")
        payment_status_filter = request.args.get("payment_status", "")
        payout_status_filter = request.args.get("payout_status", "")

        if not billing_month_str:
            return jsonify({"error": "必须提供账单月份 (billing_month) 参数"}), 400

        billing_year, billing_month = map(int, billing_month_str.split("-"))

        # 2. 构建查询 (逻辑不变, 已包含拼音搜索)
        contract_poly = with_polymorphic(BaseContract, "*")
        query = (
            db.session.query(CustomerBill, contract_poly)
            .select_from(CustomerBill)
            .join(contract_poly, CustomerBill.contract_id == contract_poly.id)
            .outerjoin(
                ServicePersonnel,
                contract_poly.service_personnel_id == ServicePersonnel.id,
            )
        )
        query = query.filter(
            CustomerBill.year == billing_year, CustomerBill.month ==billing_month
        )
        if status:
            query = query.filter(contract_poly.status == status)
        if contract_type:
            query = query.filter(contract_poly.type == contract_type)
        if search_term:
            query = query.filter(
                db.or_(
                    contract_poly.customer_name.ilike(f"%{search_term}%"),
                    contract_poly.customer_name_pinyin.ilike(f"%{search_term}%"),
                    ServicePersonnel.name.ilike(f"%{search_term}%"),
                    ServicePersonnel.name_pinyin.ilike(f"%{search_term}%"),
                )
            )
        if payment_status_filter:
            query = query.filter(CustomerBill.payment_status ==PaymentStatus(payment_status_filter))
        if payout_status_filter:
            query = query.join(EmployeePayroll,and_(EmployeePayroll.contract_id == CustomerBill.contract_id,EmployeePayroll.cycle_start_date == CustomerBill.cycle_start_date))
            query = query.filter(EmployeePayroll.payout_status ==PayoutStatus(payout_status_filter))

        bills_to_export = query.all()

        # --- 核心重构：使用字典列表来存储数据 ---
        data_rows = []
        for bill, contract in bills_to_export:
            management_fee = D(bill.calculation_details.get('management_fee',0) if bill.calculation_details else 0)
            deposit = D(0)
            introduction_fee = D(0)
            customer_increase = D(0)
            security_deposit = D(0)
            employee_payable = D(0)

            payroll = EmployeePayroll.query.filter_by(
                contract_id=bill.contract_id,
                cycle_start_date=bill.cycle_start_date,
                is_substitute_payroll=bill.is_substitute_bill
            ).first()

            customer_adjustments = FinancialAdjustment.query.filter(FinancialAdjustment.customer_bill_id == bill.id).all()
            employee_adjustments = []
            if payroll:
                employee_adjustments = FinancialAdjustment.query.filter(FinancialAdjustment.employee_payroll_id == payroll.id).all()

            adjustments = customer_adjustments + employee_adjustments

            for adj in adjustments:
                if adj.adjustment_type == AdjustmentType.DEPOSIT:
                    deposit += adj.amount
                elif adj.adjustment_type == AdjustmentType.INTRODUCTION_FEE:
                    introduction_fee += adj.amount
                elif adj.adjustment_type == AdjustmentType.CUSTOMER_INCREASE:
                    customer_increase += adj.amount
                elif adj.adjustment_type == AdjustmentType.EMPLOYEE_DECREASE:
                    employee_payable += adj.amount
                elif adj.adjustment_type == AdjustmentType.EMPLOYEE_COMMISSION:
                    employee_payable += adj.amount

                if adj.description and '保证金退款' in adj.description:
                    security_deposit -= adj.amount
                elif adj.description and '保证金' in adj.description:
                    security_deposit += adj.amount

            total_receivable = management_fee + deposit + introduction_fee +customer_increase + security_deposit + employee_payable

            employee_name = ""
            if bill.is_substitute_bill:
                sub_record = bill.source_substitute_record
                if sub_record:
                    sub_employee = sub_record.substitute_user or sub_record.substitute_personnel
                    employee_name = getattr(sub_employee, "username", getattr(sub_employee, "name", "未知替班员工"))
                else:
                    employee_name = "替班(记录丢失)"
            else:
                original_employee = contract.service_personnel
                employee_name = getattr(original_employee, "name", "未知员工")
            data_rows.append({
                "customer_name": contract.customer_name,
                "employee_name": employee_name,
                "contract_type": get_contract_type_details(contract.type),
                "management_fee": float(management_fee),
                "deposit": float(deposit),
                "introduction_fee": float(introduction_fee),
                "customer_increase": float(customer_increase),
                "security_deposit": float(security_deposit),
                "employee_payable": float(employee_payable),
                "total_receivable": float(total_receivable)
            })

        # --- 核心重构：按字典的 'total_receivable' 键进行排序 ---
        data_rows.sort(key=lambda x: x['total_receivable'], reverse=True)

        # 3. 创建 Excel 工作簿并写入数据
        wb = Workbook()
        ws = wb.active
        ws.title = f"{billing_month_str} 应收款明细"

        headers = [
            "客户姓名", "服务人员", "合同类型", "管理费", "定金",
            "介绍费", "客户增款", "保证金", "员工应缴款", "总计"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # --- 核心重构：从字典中按顺序提取值并写入 ---
        for row_dict in data_rows:
            ws.append([
                row_dict['customer_name'],
                row_dict['employee_name'],
                row_dict['contract_type'],
                row_dict['management_fee'],
                row_dict['deposit'],
                row_dict['introduction_fee'],
                row_dict['customer_increase'],
                row_dict['security_deposit'],
                row_dict['employee_payable'],
                row_dict['total_receivable']
            ])

        # 4. 保存到内存并发送 (逻辑不变)
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{billing_month_str}_receivables_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        current_app.logger.error(f"导出应收款失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误，导出失败"}), 500
    

@billing_bp.route('/pending_deposits', methods=['GET'])
@jwt_required()
def get_pending_deposits():
    """
    高效计算待收定金的月嫂合同数量。
    """
    try:
        # 子查询：找到所有已经支付了定金的合同ID
        paid_deposit_contract_ids =db.session.query(FinancialAdjustment.contract_id).filter(
            FinancialAdjustment.adjustment_type == AdjustmentType.DEPOSIT, # 修正字段名
            FinancialAdjustment.status == 'PAID' # <-- 【关键修正】
        ).distinct()

        # 主查询：计算所有状态不为“已终止”的月嫂合同中，ID不在上面子查询结果里的数量
        count = db.session.query(NannyContract.id).filter(
            NannyContract.status != 'terminated',
            not_(NannyContract.id.in_(paid_deposit_contract_ids))
        ).count()

        return jsonify({"pending_deposit_count": count})
    except Exception as e:
        current_app.logger.error(f"Failed to get pending deposits count: {e}",exc_info=True)
        return jsonify({"error": "Internal server error"}), 500

@billing_bp.route("/nanny-trial-contracts/<uuid:trial_contract_id>/convert", methods=["POST"])
@admin_required
def convert_nanny_trial_contract(trial_contract_id):
    """
    将一个成功的试工合同转换为正式合同。
    这会触发一个引擎服务，将试工期的费用和介绍费作为调整项附加到正式合同的第一个账单上。
    """
    data = request.get_json()
    if not data or "formal_contract_id" not in data:
        return jsonify({"error": "请求体中必须包含 'formal_contract_id'"}), 400

    formal_contract_id = data["formal_contract_id"]
    conversion_costs = data.get("conversion_costs")  # V12: 获取可能被覆盖的费用

    # 新增：获取当前操作员的ID
    operator_id = get_jwt_identity()

    try:
        engine = BillingEngine()
        # 修改：将 operator_id 和 conversion_costs 传递给引擎函数
        engine.process_trial_conversion(
            trial_contract_id=str(trial_contract_id),
            formal_contract_id=formal_contract_id,
            operator_id=operator_id,
            conversion_costs=conversion_costs
        )
        return jsonify({"message":"试工合同已成功转换，相关费用已附加到正式合同的第一期账单中。"}), 200
    except Exception as e:
        current_app.logger.error(f"转换试工合同 {trial_contract_id} 时发生错误:{e}", exc_info=True)
        # 返回一个更通用的错误信息给前端，具体的错误记录在日志中
        return jsonify({"error": f"处理转换失败: {str(e)}"}), 500

@billing_bp.route("/nanny-trial-contracts/<uuid:contract_id>/conversion-preview", methods=['GET', 'OPTIONS'])
@jwt_required()
def get_trial_conversion_preview(contract_id):
    """
    V6: (最终版) 精确预览试工转正的所有财务影响。
    """
    if request.method == 'OPTIONS':
        return jsonify({'status': 'ok'}), 200

    formal_contract_id = request.args.get("formal_contract_id")
    if not formal_contract_id:
        return jsonify({})

    trial_contract = db.session.get(NannyTrialContract, str(contract_id))
    if not trial_contract:
        return jsonify({"error": "试工合同未找到"}), 404

    formal_contract = db.session.get(NannyContract, formal_contract_id)
    if not formal_contract:
        return jsonify({"error": "正式合同未找到"}), 404

    try:
        engine = BillingEngine()
        trial_start = engine._to_date(trial_contract.start_date)
        trial_end = engine._to_date(trial_contract.end_date)
        formal_start = engine._to_date(formal_contract.start_date)

        # --- 1. 计算日期重叠和非重叠部分 ---
        non_overlap_days = 0
        if formal_start > trial_start:
            non_overlap_end = min(trial_end, formal_start - timedelta(days=1))
            if non_overlap_end >= trial_start:
                non_overlap_days = (non_overlap_end - trial_start).days
        
        overlap_days = 0
        if formal_start <= trial_end:
            overlap_start = max(trial_start, formal_start)
            overlap_end = trial_end
            current_app.logger.info(f"=====>overlap_end:{overlap_end}, overlap_start:{overlap_start}")
            if overlap_end >= overlap_start:
                overlap_days = (overlap_end - overlap_start).days

            current_app.logger.info(f"======>overlap_days:{overlap_days}")
        
        non_overlap_days = max(0, non_overlap_days)
        overlap_days = max(0, overlap_days)

        costs = {}

        # --- 2. 计算非重叠期的转移工资 ---
        if non_overlap_days > 0:
            trial_daily_rate = D(trial_contract.employee_level or '0')
            transfer_amount = (trial_daily_rate * D(non_overlap_days)).quantize(D("0.01"))
            if transfer_amount > 0:
                costs["non_overlap_salary"] = {
                    "amount": str(transfer_amount),
                    "description": f"纯试工期 {non_overlap_days} 天劳务费 (按试工标准)",
                    "editable": False
                }

        # --- 3. 计算重叠期的薪资和管理费差额 ---
        if overlap_days > 0:
            # a. 薪资差额
            trial_daily_rate = D(trial_contract.employee_level or '0')
            formal_daily_rate = D(formal_contract.employee_level or '0') / D(26)
            salary_diff = (formal_daily_rate - trial_daily_rate) * D(overlap_days)
            salary_diff = salary_diff.quantize(D("0.01"))
            if salary_diff != 0:
                costs["overlap_salary_adjustment"] = {
                    "amount": str(salary_diff),
                    "description": f"重叠期 {overlap_days} 天薪资差额 (员工侧)",
                    "editable": True
                }

            # b. 管理费差额
            formal_monthly_fee = D(formal_contract.management_fee_amount or '0')
            if formal_monthly_fee <= 0:
                 formal_monthly_fee = (D(formal_contract.employee_level or '0') * D('0.1')).quantize(D("0.01"))
            
            if formal_monthly_fee > 0:
                formal_daily_mgmt_fee = formal_monthly_fee / D(30)
                fee_charged_for_overlap = (formal_daily_mgmt_fee * D(overlap_days)).quantize(D("0.01"))
                
                fee_should_have_been = D(0)
                has_intro_fee = (trial_contract.introduction_fee or 0) > 0
                notes_has_mgmt_fee = "管理费" in (trial_contract.notes or "")
                if not (has_intro_fee and not notes_has_mgmt_fee):
                    trial_level = D(trial_contract.employee_level or '0')
                    fee_should_have_been = (trial_level * D('0.2') / D(30) * D(overlap_days)).quantize(D("0.01"))

                management_fee_diff = fee_charged_for_overlap - fee_should_have_been
                if management_fee_diff != 0:
                    costs["management_fee_adjustment"] = {
                        "amount": str(management_fee_diff),
                        "description": f"重叠期 {overlap_days} 天管理费差额 (客户侧)",
                        "editable": True
                    }

        # --- 4. 计算介绍费 ---
        introduction_fee = D(trial_contract.introduction_fee or '0')
        if introduction_fee > 0:
            costs["introduction_fee"] = {
                "amount": str(introduction_fee),
                "description": "介绍费 (一次性收取)",
                "editable": False
            }

        return jsonify(costs)
    except Exception as e:
        current_app.logger.error(f"预览试工合同费用失败: {e}", exc_info=True)
        return jsonify({"error": "计算预览费用时发生错误"}), 500

@billing_bp.route("/contracts/pending-trials", methods=["GET"])
@admin_required
def get_pending_trial_contracts():
    """
    V2: 返回完整的合同信息，包括 start_date 和 end_date。
    """
    try:
        today = date.today()

        # 查询时不再需要复杂的 add_columns，直接获取完整的合同对象 只查询2025年9月1日之后的数据
        pending_trials = NannyTrialContract.query.filter(
            NannyTrialContract.trial_outcome == TrialOutcome.PENDING,
            NannyTrialContract.end_date > date(2025, 9, 1)
            # NannyTrialContract.end_date < today
        ).order_by(NannyTrialContract.end_date.asc()).all()

        results = []
        for contract in pending_trials:
            employee_name = (contract.service_personnel.name if contract.service_personnel else '未知')
            results.append({
                "id": str(contract.id),
                "customer_name": contract.customer_name,
                "employee_name": employee_name,
                "message": f"{contract.customer_name} - {employee_name} 的试工合同待处理",
                "start_date": contract.start_date.isoformat() if contract.start_date else None,
                "end_date": contract.end_date.isoformat() if contract.end_date else None,
                "can_convert_to_formal": _check_can_convert(contract)
            })

        return jsonify(results)
    except Exception as e:
        current_app.logger.error(f"获取待处理试工合同列表失败: {e}", exc_info=True)
        return jsonify({"error": "获取待办事项列表时发生服务器内部错误"}), 500
    
# ==================== 第一步：创建文件服务路由 ====================
@billing_bp.route('/uploads/financial_records/<path:filename>')
@jwt_required() # 关键：保护这个端点，只有登录用户能访问
def serve_financial_record_upload(filename):
    """
    安全地提供 instance/uploads/financial_records 目录下的文件。
    """
    # 智能地构造到 instance/uploads/financial_records 的绝对路径
    directory = os.path.join(current_app.instance_path, 'uploads', 'financial_records')

    # 使用 send_from_directory 安全地发送文件
    try:
        return send_from_directory(directory, filename, as_attachment=False)
    except FileNotFoundError:
        return jsonify({"error": "文件未找到"}), 404
# ========================== 新增代码结束 ==========================




from backend.api.ai_generate import transform_text_with_llm

@billing_bp.route("/generate_payment_message", methods=["POST"])
@admin_required
def generate_payment_message():
    data = request.get_json()
    bill_ids = data.get("bill_ids")

    if not bill_ids:
        return jsonify({"error": "缺少 bill_ids 参数"}), 400

    try:
        generator = PaymentMessageGenerator()
        message_data = generator.generate_for_bills(bill_ids)
        return jsonify(message_data)
    except Exception as e:
        current_app.logger.error(f"生成催款消息失败: {e}", exc_info=True)
        return jsonify({"error": "生成消息时发生服务器内部错误"}), 500

@billing_bp.route("/beautify-message", methods=["POST"])
@admin_required
def beautify_payment_message():
    data = request.get_json()
    company_summary = data.get("company_summary", "")
    employee_summary = data.get("employee_summary", "")

    if not company_summary and not employee_summary:
        return jsonify({"error": "没有需要美化的内容"}), 400

    input_text = f"【应付公司款项】\n{company_summary}\n\n【应付员工款项】\n{employee_summary}"
    user_id = get_jwt_identity()

    try:
        beautified_data = transform_text_with_llm(
            input_text=input_text,
            prompt_identifier="SimplifiedBill",
            user_id=user_id
        )
        # 直接返回LLM生成的JSON对象
        return jsonify(beautified_data)
    except Exception as e:
        current_app.logger.error(f"美化账单信息失败: {e}", exc_info=True)
        return jsonify({"error": "AI美化失败，请稍后重试"}), 500

@billing_bp.route("/company_bank_accounts", methods=["GET"])
@admin_required
def get_company_bank_accounts():
    """
    获取所有启用的公司银行账户列表。
    """
    try:
        accounts = CompanyBankAccount.query.filter_by(is_active=True).order_by(CompanyBankAccount.is_default.desc()).all()
        results = [{
            "id": acc.id,
            "account_nickname": acc.account_nickname,
            "payee_name": acc.payee_name,
            "account_number": acc.account_number,
            "bank_name": acc.bank_name,
            "is_default": acc.is_default
        } for acc in accounts]
        return jsonify(results)
    except Exception as e:
        current_app.logger.error(f"获取公司银行账户列表失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500

@billing_bp.route("/search-unpaid-bills", methods=["GET"])
@jwt_required()
def search_unpaid_bills():
    search_term = request.args.get("search", "").strip()
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)

    # DEBUG: Log entry and parameters
    current_app.logger.info("--- [search_unpaid_bills] START ---")
    current_app.logger.info(f"Params: search='{search_term}', year={year}, month={month}")

    if not search_term or not year or not month:
        current_app.logger.info("Missing parameters, returning empty list.")
        current_app.logger.info("--- [search_unpaid_bills] END ---")
        return jsonify([])

    try:
        pinyin_search_term = search_term.replace(" ", "")
        results = []
        
        # 1. Search for Customers
        current_app.logger.info("Step 1: Searching for Customers...")
        customer_matches = db.session.query(BaseContract.customer_name).join(
            CustomerBill, BaseContract.id == CustomerBill.contract_id
        ).filter(
            CustomerBill.year == year,
            CustomerBill.month == month,
            or_(
                BaseContract.customer_name.ilike(f"%{search_term}%"),
                BaseContract.customer_name_pinyin.ilike(f"%{pinyin_search_term}%")
            )
        ).distinct().limit(10).all()
        current_app.logger.info(f"Found {len(customer_matches)} customer matches: {customer_matches}")

        for name, in customer_matches:
            results.append({
                "type": "customer",
                "name": name,
                "display": name
            })

        # 2. Search for Employees (ServicePersonnel)
        current_app.logger.info("Step 2: Searching for Employees (ServicePersonnel)...")
        personnel_matches = db.session.query(ServicePersonnel.name, BaseContract.customer_name).join(
            BaseContract, ServicePersonnel.id == BaseContract.service_personnel_id
        ).join(
            CustomerBill, BaseContract.id == CustomerBill.contract_id
        ).filter(
            CustomerBill.year == year,
            CustomerBill.month == month,
            or_(
                ServicePersonnel.name.ilike(f"%{search_term}%"),
                ServicePersonnel.name_pinyin.ilike(f"%{pinyin_search_term}%")
            )
        ).distinct().limit(10).all()
        current_app.logger.info(f"Found {len(personnel_matches)} ServicePersonnel matches: {personnel_matches}")

        for emp_name, cust_name in personnel_matches:
            results.append({
                "type": "employee",
                "name": emp_name,
                "customer_name": cust_name,
                "display": f"{emp_name} (员工, 客户: {cust_name})"
            })

        # 3. Search for Employees (internal Users)
        # This section is removed as BaseContract.user_id is being deprecated.
        # All employee searches should now go through ServicePersonnel.
        # current_app.logger.info("Step 3: Searching for Employees (Users)...")
        # user_matches = db.session.query(User.username, BaseContract.customer_name).join(
        #     BaseContract, User.id == BaseContract.user_id
        # ).join(
        #     CustomerBill, BaseContract.id == CustomerBill.contract_id
        # ).filter(
        #     CustomerBill.year == year,
        #     CustomerBill.month == month,
        #     or_(
        #         User.username.ilike(f"%{search_term}%"),
        #         User.name_pinyin.ilike(f"%{pinyin_search_term}%")
        #     )
        # ).distinct().limit(10).all()
        # current_app.logger.info(f"Found {len(user_matches)} User matches: {user_matches}")

        # for emp_name, cust_name in user_matches:
        #     results.append({
        #         "type": "employee",
        #         "name": emp_name,
        #         "customer_name": cust_name,
        #         "display": f"{emp_name} (员工, 客户: {cust_name})"
        #     })
        
        current_app.logger.info(f"Total results before deduplication: {len(results)}")

        # 4. Deduplicate and return
        final_results = list({item['display']: item for item in results}.values())
        current_app.logger.info(f"Final results after deduplication: {len(final_results)}")
        current_app.logger.info(f"Final results content: {final_results}")
        current_app.logger.info("--- [search_unpaid_bills] END ---")
        
        return jsonify(final_results)

    except Exception as e:
        current_app.logger.error(f"Failed to search unpaid bills: {e}", exc_info=True)
        current_app.logger.info("--- [search_unpaid_bills] END (with error) ---")
        return jsonify({"error": "Internal server error"}), 500
    
def _format_bill_for_reconciliation(bill, bank_transaction_id=None):
    """
    一个辅助函数，用于将单个账单对象格式化为对账页面所需的前端字典格式。
    """
    paid_by_this_txn = D('0')
    if bank_transaction_id:
        payments_from_this_txn = PaymentRecord.query.filter_by(
            customer_bill_id=bill.id,
            bank_transaction_id=bank_transaction_id
        ).all()
        if payments_from_this_txn:
            paid_by_this_txn = sum(p.amount for p in payments_from_this_txn)

    payments = []
    for pr in bill.payment_records:
        if pr.bank_transaction:
            payments.append({
                'payer_name': pr.bank_transaction.payer_name,
                'amount': str(pr.amount)
            })

    employee_name = "未知员工"
    if bill.is_substitute_bill:
        sub_record = bill.source_substitute_record
        if sub_record and (sub_record.substitute_user or sub_record.substitute_personnel):
            sub_employee = sub_record.substitute_user or sub_record.substitute_personnel
            employee_name = getattr(sub_employee, 'username', getattr(sub_employee, 'name', '未知替班员工'))
    elif bill.contract and bill.contract.service_personnel:
        original_employee = bill.contract.service_personnel
        employee_name = getattr(original_employee, 'name' , '未知员工')

    later_bill_exists = db.session.query(CustomerBill.query.filter(
        CustomerBill.contract_id == bill.contract_id,
        CustomerBill.is_substitute_bill == False,
        CustomerBill.cycle_start_date > bill.cycle_start_date
    ).exists()).scalar()
    is_last_bill = not later_bill_exists

    successor_contract_id = None
    if is_last_bill:
        successor = _find_successor_contract_internal(str(bill.contract_id))
        if successor:
            successor_contract_id = str(successor.id)

    is_balance_transferred = False
    if is_last_bill and successor_contract_id:
        # 查找是否有描述中包含“余额转出至”的财务调整项
        transfer_adjustment_exists = db.session.query(FinancialAdjustment.query.filter(
            FinancialAdjustment.customer_bill_id == bill.id,
            FinancialAdjustment.description.like('%转移至续约合同')
        ).exists()).scalar()
        is_balance_transferred = transfer_adjustment_exists

    return {
        "id": str(bill.id),
        "contract_id": str(bill.contract.id),
        "contract_status": bill.contract.status,
        "customer_name": bill.contract.customer_name,
        "contract_type": bill.contract.type,
        "employee_name": employee_name,
        "is_substitute_bill": bill.is_substitute_bill,
        "cycle": f"{bill.cycle_start_date.strftime('%Y-%m-%d')} to {bill.cycle_end_date.strftime('%Y-%m-%d')}",
        "bill_month": bill.month,
        "year": bill.year,
        "total_due": str(bill.total_due),
        "total_paid": str(bill.total_paid),
        "amount_remaining": str(bill.total_due - bill.total_paid),
        "payments": payments,
        "paid_by_this_txn": str(paid_by_this_txn),
        "is_merged": bill.is_merged,
        "is_last_bill": is_last_bill,
        "successor_contract_id": successor_contract_id,
        "is_balance_transferred": is_balance_transferred, 
        "merge_target_bill": None  # 默认设为 None
    }

@billing_bp.route("/bills-by-customer", methods=["GET"])
@jwt_required()
def get_bills_by_customer():
    """
    根据客户名和年月，获取该客户所有的账单，并计算每个账单从特定流水中已支付的金额。
    V6: 当发现可合并账单时，会主动附带上目标账单的信息。
    """
    customer_name_filter = request.args.get("customer_name")
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)
    bank_transaction_id = request.args.get("bank_transaction_id")

    if not year or not month:
        return jsonify({"error": "year and month are required."}), 400

    if not bank_transaction_id and not customer_name_filter:
        return jsonify({"error": "bank_transaction_id or customer_name is required."}), 400

    try:
        relevant_customer_names = set()
        if customer_name_filter:
            relevant_customer_names.add(customer_name_filter)
        elif bank_transaction_id:
            bank_txn = db.session.get(BankTransaction, bank_transaction_id)
            if bank_txn:
                relevant_customer_names.add(bank_txn.payer_name)
                aliases = PayerAlias.query.filter_by(payer_name=bank_txn.payer_name).all()
                for alias in aliases:
                    contract = db.session.get(BaseContract, alias.contract_id)
                    if contract:
                        relevant_customer_names.add(contract.customer_name)

        if not relevant_customer_names:
            return jsonify({"bills": [], "contracts_only": [], "closest_bill_period": None})

        all_relevant_contracts = BaseContract.query.filter(
            BaseContract.customer_name.in_(list(relevant_customer_names))
        ).all()
        all_relevant_contract_ids = [c.id for c in all_relevant_contracts]

        bills_in_period = CustomerBill.query.filter(
            CustomerBill.contract_id.in_(all_relevant_contract_ids),
            CustomerBill.year == year,
            CustomerBill.month == month
        ).all()

        all_bills_results = []
        customers_with_bills = set()

        for bill in bills_in_period:
            bill_data = _format_bill_for_reconciliation(bill, bank_transaction_id)

            if bill_data['is_last_bill'] and bill_data['successor_contract_id']:
                target_bill = CustomerBill.query.filter_by(
                    contract_id=bill_data['successor_contract_id'],
                    is_substitute_bill=False
                ).order_by(CustomerBill.cycle_start_date.asc()).first()

                if target_bill:
                    # 目标账单不需要计算银行流水分配，所以第二个参数传None
                    bill_data['merge_target_bill'] = _format_bill_for_reconciliation(target_bill, None)

            all_bills_results.append(bill_data)
            customers_with_bills.add(bill.contract.customer_name)

        contracts_only_results = []
        customers_without_bills = relevant_customer_names - customers_with_bills
        for customer_name in customers_without_bills:
            active_contract = BaseContract.query.filter(
                BaseContract.customer_name == customer_name,
            ).first()
            if active_contract:
                contracts_only_results.append({
                    'customer_name': customer_name,
                    'relevant_contract_id': str(active_contract.id)
                })

        sorted_bills = sorted(all_bills_results, key=lambda x: Decimal(x['amount_remaining']), reverse=True)

        return jsonify({
            "bills": sorted_bills,
            "contracts_only": contracts_only_results,
            "closest_bill_period": None
        })

    except Exception as e:
        current_app.logger.error(f"获取客户账单失败: {e}", exc_info=True)
        return jsonify({"error": "服务器内部错误"}), 500

@billing_bp.route("/payable-items-by-payee", methods=["GET"])
@jwt_required()
def get_payable_items_by_payee():
    payee_type = request.args.get("payee_type")
    payee_id = request.args.get("payee_id")
    year = request.args.get("year", type=int)
    month = request.args.get("month", type=int)

    if not all([payee_type, payee_id, year, month]):
        return jsonify({"error": "payee_type, payee_id, year, and month are required."}), 400

    try:
        service = BankStatementService()
        result = service.get_payable_items_for_payee(payee_type, payee_id, year, month)
        return jsonify(result)
    except Exception as e:
        current_app.logger.error(f"Failed to get payable items for payee {payee_id}: {e}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


@billing_bp.route("/bills/<uuid:bill_id>/transfer-balance", methods=["POST"])
@admin_required
def transfer_bill_balance(bill_id):
    """
    V4 - 支持将余额转移到指定合同或指定账单，并能正确触发重算。
    """
    data = request.get_json()
    destination_contract_id = data.get("destination_contract_id")
    destination_bill_id = data.get("destination_bill_id")

    if not destination_contract_id and not destination_bill_id:
        return jsonify({"error": "必须提供目标合同ID (destination_contract_id) 或目标账单ID (destination_bill_id)"}), 400

    try:
        engine = BillingEngine()
        with db.session.begin_nested():
            source_bill = db.session.get(CustomerBill, str(bill_id))
            if not source_bill:
                return jsonify({"error": "源账单未找到"}), 404

            dest_bill = None
            dest_contract = None

            if destination_bill_id:
                dest_bill = db.session.get(CustomerBill, destination_bill_id)
                if not dest_bill:
                    return jsonify({"error": "目标账单未找到"}), 404
                dest_contract = dest_bill.contract
            elif destination_contract_id:
                dest_contract = db.session.get(BaseContract, destination_contract_id)
                if not dest_contract:
                    return jsonify({"error": "目标合同未找到"}), 404

                dest_bill = CustomerBill.query.filter_by(contract_id=dest_contract.id, is_substitute_bill=False).order_by(CustomerBill.cycle_start_date.asc()).first()
                if not dest_bill:
                    return jsonify({"error": "目标合同尚未生成任何账单，无法接收余额。"}), 404

            if not dest_bill or not dest_contract:
                 return jsonify({"error": "无法确定转移目标"}), 500

            contract = source_bill.contract
            is_last_bill = not db.session.query(CustomerBill.id).filter(
                CustomerBill.contract_id == contract.id,
                CustomerBill.cycle_start_date > source_bill.cycle_start_date,
                CustomerBill.is_substitute_bill == False
            ).first()

            if contract.status not in ['terminated', 'finished'] or not is_last_bill:
                return jsonify({"error": "只能对已终止/已完成合同的最后一个账单执行余额转移"}), 400

            engine.calculate_for_month(
                year=source_bill.year, month=source_bill.month, contract_id=source_bill.contract_id,
                force_recalculate=True, cycle_start_date_override=source_bill.cycle_start_date
            )
            db.session.flush()
            db.session.refresh(source_bill)

            final_balance = source_bill.total_due - source_bill.total_paid
            if final_balance == D(0):
                return jsonify({"message": "账单余额为零，无需转移。"}), 200

            source_employee = source_bill.contract.service_personnel
            source_employee_name = source_employee.name

            dest_employee_name = "未知员工"
            if dest_bill.is_substitute_bill:
                sub_record = dest_bill.source_substitute_record
                if sub_record:
                    sub_employee = sub_record.substitute_user or sub_record.substitute_personnel
                    if sub_employee:
                        dest_employee_name = getattr(sub_employee, "username", getattr (sub_employee, "name", "未知替班员工"))
                dest_employee_name += " (替班)"
            else:
                dest_employee = dest_contract.service_personnel
                if dest_employee:
                    dest_employee_name = dest_employee.name

            if final_balance > 0:
                source_adj_type = AdjustmentType.CUSTOMER_DECREASE
                dest_adj_type = AdjustmentType.CUSTOMER_INCREASE
                amount = final_balance
            else:
                source_adj_type = AdjustmentType.CUSTOMER_INCREASE
                dest_adj_type = AdjustmentType.CUSTOMER_DECREASE
                amount = -final_balance

            source_adj = FinancialAdjustment(
                customer_bill_id=source_bill.id,
                amount=amount,
                adjustment_type=source_adj_type,
                description=f"[系统] 余额转出至: {dest_contract.customer_name}- {dest_employee_name}",
                date=source_bill.cycle_end_date.date(),
                details={"status": "transferred_out"}
            )
            db.session.add(source_adj)
            db.session.flush()

            dest_adj = FinancialAdjustment(
                customer_bill_id=dest_bill.id,
                amount=amount,
                adjustment_type=dest_adj_type,
                description=f"[系统] 余额从: {source_bill.contract.customer_name}- {source_employee_name} 转入",
                date=dest_bill.cycle_start_date.date(),
                details={
                    "status": "transferred_in",
                    "transferred_from_adjustment_id": str(source_adj.id),
                    "transferred_from_bill_id": str(source_bill.id)
                }
            )
            db.session.add(dest_adj)
            db.session.flush()

            source_adj.details["transferred_to_adjustment_id"] = str(dest_adj.id)
            source_adj.details["transferred_to_bill_id"] = str(dest_bill.id)
            attributes.flag_modified(source_adj, "details")

        # --- 核心修正：根据目标账单类型调用正确的重算方法 ---
        # 重算源账单
        engine.calculate_for_month(year=source_bill.year, month=source_bill.month, contract_id=source_bill.contract_id, force_recalculate=True, cycle_start_date_override=source_bill.cycle_start_date)

        # 重算目标账单
        if dest_bill.is_substitute_bill:
            engine.calculate_for_substitute(substitute_record_id=dest_bill.source_substitute_record_id, commit=False)
        else:
            engine.calculate_for_month(year=dest_bill.year, month=dest_bill.month, contract_id=dest_bill.contract_id, force_recalculate=True, cycle_start_date_override=dest_bill.cycle_start_date)

        db.session.commit()
        # --- 修正结束 ---

        latest_details = get_billing_details_internal(bill_id=str(source_bill.id))
        return jsonify({"message": "余额结转成功！", "latest_details": latest_details})

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"为账单 {bill_id} 执行余额转移时发生意外错误: {e}", exc_info= True)
        return jsonify({"error": "执行余额转移失败，服务器内部错误"}), 500